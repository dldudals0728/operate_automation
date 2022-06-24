import sys
import logging
from tkinter.tix import CheckList
from tokenize import String
from turtle import color
from typing import DefaultDict
# 표 생성 함수
# QMainWindow: 상태표시줄, 메뉴 추가 / QAction: 메뉴 액션 추가 / QMenu: menu sub group 추가 / qApp: 앱 종료 함수 사용
# from PyQt5.QtWidgets import (QApplication, QWidget, QPushButton, QMessageBox, QMainWindow, QAction, QMenu, qApp,
# QVBoxLayout, QHBoxLayout)
from PyQt5.QtWidgets import *
# 이벤트 처리. 슬롯으로 연결해줌(connect).
from PyQt5.QtCore import QCoreApplication, QLine, Qt, QDate
# table Read only mode
from PyQt5 import QtGui
from PyQt5.QtGui import *
# Tree view에 나열되는 내용을 담당.
# from PyQt5.QtGui import QStandardItemModel

from pymysql import NULL

from database import DB

from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
import datetime

import os
import shutil

# from PIL import Image

from automation import Automation

class LogIn(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.isLogin = False


    def initUI(self):
        self.setFixedSize(600, 300)
        self.setWindowTitle("남양노아요양보호사교육원 DBMS")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.main_box = QVBoxLayout()
        self.box_logo = QHBoxLayout()
        self.box_top = QHBoxLayout()
        self.box_middle = QHBoxLayout()
        self.box_bottom = QHBoxLayout()

        self.label_logo = QLabel("img", self)
        self.img_logo = QPixmap(r"D:\Master\PythonWorkspace\NYNOA\Icons\logo.png")
        self.img_logo.scaledToWidth(150)
        self.label_logo.setPixmap(self.img_logo)
        self.label_logo.setAlignment(Qt.AlignCenter)

        self.box_logo.addWidget(self.label_logo)
        
        self.main_box.addLayout(self.box_logo)
        self.main_box.addLayout(self.box_top)
        self.main_box.addLayout(self.box_middle)
        self.main_box.addLayout(self.box_bottom)

        self.box_top.addStretch(1)
        self.label_id = QLabel("아이디", self)
        self.label_id.setFixedWidth(100)
        self.label_id.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.input_id = QLineEdit(self)
        self.box_top.addWidget(self.label_id)
        self.box_top.addWidget(self.input_id)
        self.box_top.addStretch(1)

        self.box_middle.addStretch(1)
        self.label_pwd = QLabel("비밀번호", self)
        self.label_pwd.setFixedWidth(100)
        self.label_pwd.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.input_pwd = QLineEdit(self)
        self.input_pwd.setEchoMode(QLineEdit.Password)
        self.box_middle.addWidget(self.label_pwd)
        self.box_middle.addWidget(self.input_pwd)
        self.box_middle.addStretch(1)

        self.check_id_save = QCheckBox("아이디 저장", self)
        self.box_bottom.addWidget(self.check_id_save)
        self.box_bottom.addStretch(1)
        self.btn_login = QPushButton("로그인", self)
        self.btn_login.clicked.connect(self.logIn)
        self.box_bottom.addWidget(self.btn_login)

        self.setLayout(self.main_box)

        self.show()

    def logIn(self):
        user_id = self.input_id.text().strip()
        user_pwd = self.input_pwd.text().strip()

        if user_id == "" or user_pwd == "":
            QMessageBox.about(self, "입력 오류", "아이디와 비밀번호를 입력해주세요")
            return
        where = "id='{}' and password='{}'".format(user_id, user_pwd)
        res = db.main.dbPrograms.SELECT("id, password", "account", where=where)

        if not res:
            QMessageBox.about(self, "오류", "아이디 혹은 비밀번호 오류입니다.")
        else:
            self.isLogin = True
            idSave = self.check_id_save.isChecked()
            if idSave:
                current_id = self.input_id.text().strip()
                query = "savedId = '{}'".format(current_id)
            else:
                query = "savedId = NULL"

            if self.saved_id == NULL:
                where = "savedId IS null"
            else:
                where = "savedId='{}'".format(self.saved_id)

            db.main.dbPrograms.UPDATE("loginInfo", query, where)

            db.show()
            todo_list.show()
            db.main.dbPrograms.dumpDatabase(daily=True)
            self.close()


    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.logIn()

    def showEvent(self, QShowEvent):
        idIsSaved = db.main.dbPrograms.SELECT("savedId", "loginInfo")[0][0]
        if idIsSaved != None:
            self.check_id_save.setChecked(True)
            self.saved_id = idIsSaved
            self.input_id.setText(self.saved_id)
        else:
            self.check_id_save.setChecked(False)
            self.saved_id = NULL

    def closeEvent(self, QCloseEvent):
        # QMessageBox.question(인자, title, message, 버튼 추가(여러개 가능(|사용)), 버튼 기본값)
        if not self.isLogin:
            ans = QMessageBox.question(self, "종료", "DBMS를 종료하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if ans == QMessageBox.Yes:
                ####################### 여기다가 conn.close()추가하기 !!!!!!!!!!!!!!!!!!!!!
                try:
                    db.main.dbPrograms.conn.close()
                except:
                    print("DBGUI Exception: Database is already closed!")
                finally:
                    QCloseEvent.accept()
            else:
                QCloseEvent.ignore()
        else:
            pass
        

class WorkingInformation(QWidget):
    global db

    def __init__(self):
        super().__init__()

        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.setWindowTitle("업무 안내서")
        self.setStyleSheet("background-color: #FFFFF0;")
        self.setGeometry(400, 200, 1200, 800)

        self.main_box = QVBoxLayout()

        self.main_tab = QTabWidget()
        self.tab_list = []
        self.scroll_area_list = []
        self.tab_title_list = ["기관 운영", "개강과 종강", "대체실습", "국시원 - 응시원서 접수 및 응시표 출력", "국시원 - 합격자 명단"]

        # tab 생성
        for i in range(len(self.tab_title_list)):
            self.tab_list.append(QWidget())

        # main tab에 tab 추가 (with title)
        for tab, title in zip(self.tab_list, self.tab_title_list):
            self.main_tab.addTab(tab, title)

        # 각 tab의 layout을 QVBoxLayout으로 지정
        for tab in self.tab_list:
            tab.layout = QVBoxLayout(self)        


        title = "<b style='font-size: 30px; font-weight: bold;'>남양노아요양보호사교육원 기관장 업무</b>"
        self.label_title = QLabel(title, self)
        this_is_auto_program = """이 프로그램은 <b style='font-size: 15px; color: red; text-decoration: underline;'>데이터베이스에 입력된 정보를 기반으로 파일을 자동으로 생성하는 프로그램</b>입니다!<br>
        <b style='color: blue; font-size: 15px;'>데이터를 입력하실 때 오타가 나지 않도록 조심해주세요!</b>"""
        self.label_warning = QLabel(this_is_auto_program, self)

        self.main_box.addWidget(self.label_title)
        self.main_box.addWidget(self.label_warning)

        self.line_frame = QFrame()
        self.line_frame.setFrameShape(QFrame.HLine)
        self.line_frame.setFrameShadow(QFrame.Plain)
        self.line_frame.setLineWidth(2)

        self.main_box.addWidget(self.line_frame)        

        self.text_list = []
        self.page_list = []
        self.label_list = []

        for i in range(len(self.tab_list)):
            self.page_list.append("")
            self.label_list.append(QLabel(self))
            
        self.initUI()

        for i in range(len(self.tab_list)):
            self.scroll_area_list.append(QScrollArea())
            self.scroll_area_list[i].setStyleSheet("background-color: white;")
            self.label_list[i].setText(self.page_list[i])
            self.scroll_area_list[i].setWidget(self.label_list[i])
            self.tab_list[i].layout.addWidget(self.scroll_area_list[i])

            self.tab_list[i].setLayout(self.tab_list[i].layout)

        self.main_box.addWidget(self.main_tab)
        self.setLayout(self.main_box)
        # self.show()

    def initUI(self):
        rule_title = "<br><b style='font-size: 20px; font-weight: bolder; color: blue;'>교육기관장의 교육 운영</b><br>"

        rule_0_basic = """<b style='font-size: 20px; font-weight: bold; color: red; text-decoration: underline;'>업무 중 가장 우선순위는 \"원장님이 시키신 업무\"입니다!<br>
        급한 일이 있을 경우, 원장님께 업무 후에 하겠다고 미리 말씀드려야 합니다!<br>"""

        rule_1_summary = """<b><b>아침</b>
        <br>간판 불 OFF ➜ 전화 돌리기 & 출근부 작성, 발열체크 ➜ 강의실 시건 해제 ➜ 출석부 교체(야간 → 주간) ➜ 청소상태 & 쓰레기통 확인<br>
        ➜ 화장실 청소상태 & 휴지 여분 확인 ➜ 실내화 정리 ➜ 포털사이트 학원 검색 ➜ 프로그램 실행
        <h3>점심</h3>
        <br>강의실 전기장치 OFF(에어컨(선풍기) or 히터(온풍기)) ➜ 화장실 점검 ➜ 신발장 점검
        <h3>저녁</h3>
        <br>간판 불 ON ➜ 출석부 교체(주간 → 야간) ➜ 강의실 휴지통 및 책상 정리 & 환기 ➜ 화장실 점검 ➜ 신발장 점검<br>
        ➜ 컴퓨터 전원 OFF and 스위치(모니터 뒤) OFF ➜ 전화 돌리기 ➜ 책상 정리정돈 ➜ <span style='color: red;'>★퇴근!!!★</span><b>
        """

        rule_2_work_start = """<b>출근 시</b>
        <br>출근 후에는 <b>전화를 기관으로 돌려 놓고 출근부 작성과 발열체크</b>를 합니다.<br>
        먼저 <span style='color: blue;'>강의실</span>입니다.<br>
        그리고 강의실 문 <b>시건장치를 해제하여 문을 열고 주간 출석부를 교실에</b> 올려놓습니다.<br>
        <span style='color: blue; text-decoration: underline;'>(전날 야간 출석부가 교실에 있는 경우 사무실로 가져옵니다.)</span><br>
        그 후 <b>교실 쓰레기통과 청소상태를 확인</b>합니다.<br>
        다음은 <span style='color: blue;'>화장실과 현관</span>입니다.<br>
        화장실의 청소 상태와 <b>휴지 & 휴지 여분</b>이 충분한지 확인합니다.<br>
        현관에서는 <span style='text-decoration: underline;'>신발장의 실내화를 한 쪽으로</span> 몰아놓습니다.<br>
        컴퓨터를 키고 <b>남양노아간호학원</b>과 <b>남양노아요양보호사교육원</b>을 <span style='text-decoration: underline;'>검색</span>합니다.<br>
        그리고 프로그램을 실행시켜 <b style='color: red;'>오늘 해야 할 업무를 정리합니다.</b>"""

        rule_3_lunch = """<b>점심시간</b>
        <br>점심시간에는 강의실의 전기장치를 확인합니다.<br>
        강의실의 <b style='color: red;'>에어컨(혹은 히터)를 끄고</b>, <b>화장실과 신발장을 한번 더 점검</b>합니다.
        """

        rule_4_work_end = """<b>퇴근 전 시간</b>
        <br>날이 어두워지면 <span style='color: blue;'>간판의 불을 켭니다.</span><br>
        수업이 끝나고 모두 퇴근하시면, <span style='text-decoration: underline;'>출석부를 교체(주간 → 야간)하면서 강의실의 휴지통 및 책상 배열을 정리정돈하고, 강의실을 <b>환기</b></span>합니다.<br>
        마찬가지로 <b>화장실과 신발장을 점검</b>합니다.<br>
        퇴근 시에는 <b style='color: red; text-decoration: underline;'>컴퓨터 전원을 끈 후에 모니터 뒤 스위치를 OFF 합니다.<br>
        그리고 전화를 돌리고 책상 위에 개인정보 서류가 없도록 정리정돈한 후에 퇴근합니다.</b>
        """

        """
        대체실습 확인 서류(사진)
        """

        self.text_list = []

        self.text_list.append(rule_title)
        self.text_list.append(rule_0_basic)
        self.text_list.append(rule_1_summary)
        self.text_list.append(rule_2_work_start)
        self.text_list.append(rule_3_lunch)
        self.text_list.append(rule_4_work_end)

        for i in range(len(self.text_list)):
            self.text_list[i] = "<p style='font-size: 15px;'>" + self.text_list[i] + "</p>"
            self.page_list[0] += self.text_list[i]



        main_task_title = "<br><b style='font-size: 20px; font-weight: bolder; color: blue;'>개강과 종강</b>"

        file_lotation = """<br><b>파일 경로</b><span style='color: blue;'>기본 경로는 'D:\남양노아요양보호사교육원' 입니다.</span>
        <br>수강생: 교육생 관리 - 각 기수 / 시간표 및 출석부(저장): 경기도청 - 00_개강준비서류 - 기수 폴더 - 시간표<br>
        수강료 수납대장 & 사물함 주기: 교육생 관리 - 각 기수 / 커리큘럼: 경기도청 - 00_개강준비서류 / 개강보고서: 경기도청 - 01_개강보고 - 기수<br>
        """

        task_0_summary = """<b><b style='font-size: 18px;'>SUMMARY</b>
        <br>DBMS에 개강 반, 대체실습 반 데이터 입력 ➜ 상담 후 DBMS에 입력 ➜ 생성된 폴더에 스캔 자료 저장 ➜ 시간표 받기 & 검토<br>
        ➜ 수강료 수납대장[File - New - 수강료 수납대장] & 사물함 주기[File - New - 사물함 주기] ➜ 개강보고 명단 작성[File - 경기도청 - 개강보고]<br>
        ➜ 커리큘럼 만들기 ➜ 개강보고(D+2) ➜ 출석부 만들기[File - 경기도청 - 출석부]</b><br>
        """
        
        task_1_normal = """<b style='font-size: 18px;'>평상시(D-n)</b>
        <br><b>전담 암무: 입학생 면담 및 데이터 입력</b><br>
        <b style='color: red; text-decoration: underline;'>(필수!)우선적으로 개강이 예정되어 있는 반(ex. 10기 야간 또는 대체실습 12기)에 대한 데이터를 각각 기수 관리, 혹은 대체실습 탭에 자료를 입력합니다.</b><br> 
        면담 후 입학자(원서를 작성한 자)는 DBMS에 정보를 입력합니다. <b style='color: red; text-decoration: underline;'>(오타가 나오지 않게 주의!)</b><br>
        수강생이 가져온 서류(사진, 주민등록등본, 기본증명서, 자격증 등)가 있다면 <b style='color: red; text-decoration: underline;'>먼저 DBMS에 신상정보를 입력한 후에 scan을 떠서 해당 수강생 폴더로 옮겨 줍니다.</b><br>
        <b style='color: red; font-size: 17px;'>자격증은 추후에 경기도청에 보내야 할 서류이기 때문에, 흐릿하면 안되고, 선명한 스캔 파일이어야 합니다!!!</b><br>
        <span style='color: blue;'>(DBMS에 학생 정보를 입력하게되면 해당 학생의 폴더가 생성됩니다!)</span><br>
        <b style='color: red; text-decoration: underline;'>!주의!(파일 이름은 각각 [01기주간_이영민(or 01기야간_이영민) / 이영민_주민등록등본 / 이영민_기본증명서] 의 양식을 따릅니다!)</b><br>
        전화 상담 시에는 <b style='color: blue; text-decoration: underline;'>전화번호와 이름을 받은 후 문자안내</b>를 진행합니다."""

        task_2_ready = """<br><b style='font-size: 18px;'>개강 일주일 전(개강 D-7)</b>
        <br><b>개강보고 시간표 정리 및 검토</b><br>
        개강보고 시에 사용되는 시간표는 원장님께 받을 수 있습니다. 개강 시간표를 받은 후에 시간, 날짜 등이 맞는지 아래의 기준을 통해 확인합니다.<br>
        <span style='color: blue;'>일반[이론: 80 / 실기: 80], 사회복지사[이론: 32 / 실기: 10]<br>
        간호조무사, 물리치료사, 작업치료사[이론: 31 / 실기: 11], 간호사[이론: 26 / 실기: 6], 경력자[이론: 80 / 실기: 40]</span><br>
        """

        task_3_open = """<br><b style='font-size: 18px;'>개강 전(개강 D-1)</b>
        <br><b>수강료 납부 대장 및 사물함 주기, 커리큘럼 만들기</b><br>
        수강료 납부 대장은 원장님이 볼 서류 1부, 경기도청 자료 1부, 총 2부로 나뉩니다.<br>
        원장님이 보는 서류에는 <b style='color: blue;'>학생별 납입 여부, 납입 방법, 납입 금액, 책 수령 여부 등을 표기</b>합니다.<br>
        ([File - New - 수강료 수납대장])<br>
        <span style='color: blue;'>사물함 주기는 주간반은 노란색, 야간반은 보라색이며 기수대로 출력하여 사물함에 부착 합니다.</span><br>
        ([File - New - 사물함 주기])<br>
        <b style='text-decoration: underline'>이 때 책값을 납부한 사람은 사물함에 책을 넣어주고 납부대장에 표기 합니다.</b><br>
        커리큘럼은 기존에 있는 파일에서 기수, 기간, 시간표만 수정하여 만들 수 있습니다. <span style='color: blue;'>*커리큘럼의 시간표에는 강사님들의 이름을 지워야 합니다!*</span><br>
        <b style='color: greeen;, text-decoration: underline;'>개강보고를 하기 전까지 해당 기수의 입학을 인정할 수 있습니다.(개강보고는 개강 후 3일(D+2)에 진행합니다.</b>"""

        task_4_report = """<br><b style='font-size: 18px;'>개강 보고(개강 D+2)</b>
        <br><b>경기도청 개강 보고와 출석부 출력</b><br>
        <b style='font-size: 15; color: red; text-decoration: underline;'>경기도청에 개강을 보고하는 것은 실제 개강 후 2일 뒤, 16시 입니다.(업무보고 메뉴에 반영된 대로.)</b><br>
        DBMS의 [File - 경기도청 - 개강보고]를 통해 생성된 명단(Excel)을 복사하여 붙여넣기 함으로 번거로운 과정을 줄일 수 있습니다!<br>
        <b style='color: red; text-decoration: underline;'>만약 명단을 만들었는데 프로그램이 비정상적으로 종료된다면, 해당 기수에 모든 데이터가 입력되어 있는지 확인해주세요!</b><br>
        그 후에 <b style='color: red; text-decoration: underline;'>명단 수가 맞는지, 누락자 혹은 추가 인원이 없는지 꼭 확인한 후에 보고 메일을 발송해야 합니다!.<br>
        <b style='color: blue; text-decoration: underline;'>경기도청에 보고가 완료되었다면, 더이상 인원이 추가되거나 삭제될 수 없습니다.</b><br>
        <b>보고가 완료된 후에는 출석부를 만들어 출력합니다.</b>[File - 경기도청 - 출석부]로 <span style='color: blue;'>출석부 명단을 생성할 수 있습니다.</span><br>
        """

        task_5_finish = """<br><b style='font-size: 18px;'>개강 후 & 종강</b><br>
        <br>개강 후에는 <b style='color: red;'>받지 않은 서류, 사진, 혹은 흐릿한 자격증 등이 있는지 파악하고, 빠른 시일 내에 받아 DBMS를 최신화 시키고 스캔한 서류를 보관합니다.</b><br>
        <b style='color: red;'>또, 개강 후에는 수강생들의 수강료 수납 여부, 서류 제출 여부등을 파악하여 빠른 시일 내에 작업을 완료해야 합니다!</b><br>
        종강 시에는 따로 종강보고가 없습니다.<br>
        종강이 다가올 경우 종강하는 기수의 학생들에게 <b>대체실습 기수를 부여</b>해야 합니다.<br>
        각 선생님들의 대체실습 가능 여부(주간 or 야간 or 주말)를 판단하여 적절한 반에 배치하고, <b style='color: red; text-decoration: underline;'>DBMS에 대체실습 란에 대체실습 기수를 작성합니다.</b><br>
        """

        self.text_list = []

        self.text_list.append(main_task_title)
        self.text_list.append(file_lotation)
        self.text_list.append(task_0_summary)
        self.text_list.append(task_1_normal)
        self.text_list.append(task_2_ready)
        self.text_list.append(task_3_open)
        self.text_list.append(task_4_report)
        self.text_list.append(task_5_finish)

        for i in range(len(self.text_list)):
            self.text_list[i] = "<p style='font-size: 15px;'>" + self.text_list[i] + "</p>"
            self.page_list[1] += self.text_list[i]

        
        temp_training_title = "<br><b style='font-size: 20px; font-weight: bolder; color: blue;'>대체실습</b>"

        file_lotation = """<br><b>파일 경로</b><span style='color: blue;'>기본 경로는 'D:\남양노아요양보호사교육원\경기도청\\02_대체실습' 입니다.</span>
        <br>일정 및 명단: 01_1. 대체실습 일정 및 명단 / 실시, 수료 보고: 01. 실시수료보고 / 대체실습 출석부: 평가감상출석서식<br>
        """

        temp_0_summary = """<b><b style='font-size: 18px;'>SUMMARY</b>
        <br>DBMS에 대체실습 기수 입력 ➜ 대체실습 일정 만들기<span style='color: red;'> + 기존 인원 추가!</span><br>
        ➜ 대체실습 실시 보고서 작성[File - New - 경기도청 보고 - 대체실습 실시보고] & 검토<span style='color: red;'> + 기존 인원 추가!</span> ➜ 대체실습 출석부 생성<br>
        ➜ 대체실습 보고서(사진) 만들기 ➜ 대체실습 수료보고[File - New - 경기도청 보고 - 대체실습 수료보고]<span style='color: red;'> + 기존 인원 추가!</span>
        """

        temp_1_report = """<br><b style='font-size: 18px;'>대체실습 실시 보고와 수료보고</b>
        <br><b>경기도청 대체실습 실시 보고</b><br>
        실시보고는 주중 개강반, 주말 개강반에 따라 나뉩니다.<br>
        대체실습 기수가 <b style='text-decoration: underline'>주중에 개강</b>할 시 <b style='color: blue;'>전날에 실시 보고를 올립니다.</b><br>
        대체실습 기수가 <b style='text-decoration: underline'>주말에 개강</b>할 시 <b style='color: blue;'>전주 금요일에 실시 보고를 올립니다.</b><br>
        <br><br><b>경기도청 대체실습 수료 보고</b><br>
        수료보고는 주중 개강반, 주말 개강반에 따라 나뉩니다.<br>
        대체실습 기수가 <b style='text-decoration: underline'>주중에 종강</b>할 시 <b style='color: blue;'>다음날에 실시 보고를 올립니다.</b><br>
        대체실습 기수가 <b style='text-decoration: underline'>주말에 종강</b>할 시 <b style='color: blue;'>다음주 월요일에 실시 보고를 올립니다.</b><br>"""

        temp_2_open = """<br><b style='font-size: 18px;'>대체실습 개강 전</b>
        <br><b>대체실습 시작 전(D-2)</b><br>
        대체실습이 시작되기 전에, 원장님과의 상의를 통하여 대체실습 기수를 만들어야 합니다.<br>
        주간, 야간, 주말반 중에서 각 학생들이 참여할 수 있는 반을 조사하여 대체실습 기수를 생성합니다.<br>
        <b style='color: red; text-decoration: underline;'>생성된 대체실습 기수의 시작날짜, 종료날짜를 DBMS에 기입합니다.</b><br>
        그 후 대체실습 <b>일정</b>을 작성해야 합니다. <b>대체실습 일정은 각 요일별로 참여하는 인원의 명단을 의미합니다.(파일 있음)</b><br>
        <span style='color: blue;'>(* 일정을 작성하는 이유는, 자격증반은 대체실습이 1일이고, 일반반은 그렇지 않기 때문입니다.)<br>
        <br>일정을 작성하고 대체실습 실시 보고서를 작성합니다.([File - New - 경기도청 보고 - 대체실습 실시보고]) 이 때 <b style='color: red;'>교수진의 이름이 정확한지, 명단, 날짜 등은 정확한지 자세한 검토가 필요합니다!</b></span><br>
        <b style='color: red; font-size: 18px;'>!중요: 별도로 기존에 대체실습을 진행하시다가 넘어오신 분들은 일정 및 보고서에 ★따로★ 추가해야 합니다!</b><br>
        보고서가 작성이 완료되면 대체실습이 시작하기 하루 전날(D-1)에 보고합니다.<br>
        대체실습 실시보고가 끝나면 대체실습 출석부를 작성하여 출력합니다.<br>
        """

        temp_3_picture = """<br><b style='font-size: 18px;'>대체실습 사진</b>
        <br><b>날짜 별 대체실습 사진 첨부</b><br>
        대체실습을 진행하는 날마다, 4장의 사진이 필요합니다.<br>
        사진은 <span style='color: blue;'>1. 전체 인원과 교수, 2. 환기 사진, 3. 시험보는 인원</span>이 찍혀있어야 합니다.<br>
        <span style='color: blue;'>* 시험을 보는 인원은 자격증 반인 경우입니다. 해당 학생이 시험보는 사진을 게시합니다.<br>
        * 시험보는 인원이 없다면, 환기 2장, 전체 사진 2장(앞, 뒤)으로 합니다.</span><br>
        그 후 대체실습 사진을 모아 서류를 작성합니다.<br>
        """

        temp_4_end = """<br><b style='font-size: 18px;'>대체실습 수료</b>
        <br><b>대체실습 수료보고</b><br>
        대체실습의 수료보고는 요일에 따라 나뉩니다.<br>
        <b style='text-decoration: underline'>주중에 종강</b>할 시 <b style='color: blue;'>다음날에 실시 보고를 올립니다.</b><br>
        <b style='text-decoration: underline'>주말에 종강</b>할 시 <b style='color: blue;'>다음주 월요일에 실시 보고를 올립니다.</b><br>
        ([File - New - 경기도청 보고 - 대체실습 수료보고])<br>
        <b style='color: red; font-size: 18px;'>!중요: 별도로 기존에 대체실습을 진행하시다가 넘어오신 분들은 일정 및 보고서에 ★따로★ 추가해야 합니다!</b><br>
        """

        self.text_list = []

        self.text_list.append(temp_training_title)
        self.text_list.append(file_lotation)
        self.text_list.append(temp_0_summary)
        self.text_list.append(temp_1_report)
        self.text_list.append(temp_2_open)
        self.text_list.append(temp_3_picture)
        self.text_list.append(temp_4_end)

        for i in range(len(self.text_list)):
            self.text_list[i] = "<p style='font-size: 15px;'>" + self.text_list[i] + "</p>"
            self.page_list[2] += self.text_list[i]

        exam_title = "<br><b style='font-size: 20px; font-weight: bolder; color: blue;'>국시원 - 응시원서 접수 및 응시표 출력</b>"

        file_lotation = """<br><b>파일 경로</b><span style='color: blue;'>: 'D:\남양노아요양보호사교육원\경기도청\\03_시험준비 및 자격증발급관련' 입니다.</span>
        """

        exam_0_summary = """<b><b style='font-size: 18px;'>SUMMARY</b>
        <br>DBMS에 시험 회차 입력(재시험을 봐야하는 사람도 해당 회차로 시험 회차 변경) ➜ 응시접수 명단[File - 국시원 - 응시접수 명단] 생성 ➜ ID, PW 있는 사람 조사해서 미리 기입<br>
        ➜ 국시원 계정이 없는 사람은 회원가입 진행 ➜ 응시 접수 & 가상계좌 응시접수 명단에 기입 ➜ 입금 후 입금확인 여부 기입<br>
        ➜ 응시 수수료 감면 대상자 체크 후 환불 ➜ 응시표 및 코로나 자가문진표 출력 후 배부<br>
        """

        exam_1_acceptance = """<br><b style='font-size: 18px;'>응시원서 접수</b>
        <br><b>응시원서 접수 및 결제</b><br>
        <b style='color: red; text-decoration: underline;'>응시원서 접수기간 전에(D-n), 해당 시험에 시험을 보는 사람들의 시험 회차를 입력해야 합니다.</b><br>
        응시원서 접수기간 전 주에(D-7) <span style='color: blue;'>응시접수 명단[File - 국시원 - 응시접수 명단]을 생성</span>한 후에 해당 인원들 중 <b style='color: red;'>국시원 ID, PW가 있는 사람은 미리 입력해 줍니다.</b><br>
        <b style='color: blue;'>프로그램으로 기입된 사람들은 시험회차가 해당 회차로 입력된 사람들입니다. 재시험인 분들도 모두 시험 회차를 해당 회차로 변경해야 합니다!</b><br>
        국시원 ID, PW 파악이 완료되면, 나머지 사람들의 회원가입을 진행합니다. (id: noacw00000 / pw: noa3564626*, id는 순서대로 입력합니다.(1 -> 2 -> 3 -> ...))<br>
        국시원 계정 파악이 완료된 후 응시원서 접수를 시작합니다. 응시원서 접수 시 <b style='color: red; text-deocration: underline'>생성된 가상계좌번호를 응시접수 명단에 기입해야 합니다!(중요)</b><br>
        파악한 가상계좌로 응시원서비를 모두 입금하고, 비고란에 입금 여부를 작성하여 파악하기 쉽게 합니다.<br>
        <span style='color: red;'><span style='text-decoration: underline;'>(중요!)응시수수료 감면 대상자</span>에 해당하는 사람들을 체크하여 별도로 표기하고, 추후에 환불해줍니다.</span><br>
        """

        exam_2_print = """<br><b style='font-size: 18px;'>응시표 출력</b>
        <br><b>응시표 출력</b><br>
        응시표 출력 기간이 되면 응시표를 출력할 수 있습니다. 처음 하루, 이틀은 사용자가 많아 국시원 홈페이지가 마비될 수 있습니다.<br>
        저장된 응시접수 명단에 있는 ID와 PW로 로그인하여 응시표를 출력합니다. <b style='color: red;'>응시표를 출력할 때는 <span style='text-decoration: underline;'>'첫장'</span>만 출력합니다!(응시 지역까지만 출력)</b><br>
        추가적으로 코로나 자가문진표를 <b style='color: blue;'>인원수대로</b>출력하여 모두에게 전달할 수 있도록 준비합니다.</b><br>
        <span style='color: red;'>응시표를 나누어 드릴 때 시험 잘보시라고 화이팅 한번만 해주세요 ^^</span><br>
        """

        self.text_list = []

        self.text_list.append(exam_title)
        self.text_list.append(file_lotation)
        self.text_list.append(exam_0_summary)
        self.text_list.append(exam_1_acceptance)
        self.text_list.append(exam_2_print)

        for i in range(len(self.text_list)):
            self.text_list[i] = "<p style='font-size: 15px;'>" + self.text_list[i] + "</p>"
            self.page_list[3] += self.text_list[i]

        ending_title = "<br><b style='font-size: 20px; font-weight: bolder; color: blue;'>국시원 - 합격자 명단</b>"

        file_lotation = """<br><b>파일 경로</b><span style='color: blue;'>: 'D:\남양노아요양보호사교육원\경기도청\\03_시험준비 및 자격증발급관련' 입니다.</span>
        """

        ending_0_summary = """<b><b style='font-size: 18px;'>SUMMARY</b>
        <br>DBMS에 시험 회차 입력(재시험을 봐야하는 사람도 해당 회차로 시험 회차 변경) ➜ 학생 서류 작성[File - 국시원 - 서류작성]<br>
        ➜ 합격자 명단 작성 및 옮기기 ➜ 합격자 사진 모으기[File - 국시원 - 사진 모으기] ➜ 학생 서류 출력 [File - 국시원 - 서류출력]<br>
        """

        ending_1_exam = """<br><b style='font-size: 18px;'>개인 서류 작성</b>
        <br><b>교육수료증명서, 대체실습확인서, 요양보호사 자격증 발급, 재발급, 신청서 작성</b><br>
        합격자 발표가 나오면 <span style='color: red;'>합격한 학생들의 </span><span style='text-decoration: underline;'>서류</span>를 작성해야 합니다.<br>
        (합격하지 못한 학생들도 파일은 생성됩니다.)<br>
        <span style='text-decoration: underline;'>(교육수료증명서, 대체실습확인서, 요양보호사 자격증 발급, 재발급, 신청서)</span><br>
        [File - 국시원 - 서류]탭에서 해당 회차 학생들의 서류를 생성할 수 있습니다.<b style='color: red;'>단, <span style='font-size: 20px;'>문서를 생성하는 학생들의 "모든 항목"이 DBMS에 작성돼 있어야 합니다!</span></b><br>
        """

        ending_2_pass = """<br><b style='font-size: 18px;'>시험 합격 및 자격증 발급</b>
        <br><b>합격자 명단 작성 및 옮기기</b><br>
        <br>시험 합격자 발표일이 되면, 응시접수 명단을 이용하여 개인 계정으로 로그인 한 후, 시험 합격 여부를 조사하여 시험에 떨어진 인원이 있는지 파악합니다.<br>
        시험에 불합격한 사람이 있다면 DBMS에 해당 인원의 시험 회차를 NULL로 변경합니다. 나중에 시험에 재응시를 원하면 재응시 시험 회차를 입력할 수 있습니다.<br>
        그 다음 합격자 명단을 생성합니다.[File - 국시원 - 합격자 명단] 합격자 명단 파일을 생성하고, <span style='color: red; text-decoration: underline;'>불합격된 인원은 명단에서 제외시킵니다.(시험회차를 변경했다면 제외되어 있을 것입니다.)</span><br>
        <b style='color: red; font-size: 20px; text-decoration: underline;'>(중요!) 시험을 이전 회차에 응시하고 합격했지만, 건강검진 등의 이유로 자격증 신청이 미루어진 사람도 합격자 명단에 "수기로" 기입합니다!</b><br>
        <span style='color: blue; font-size: 20px;'>추가적으로, 위 기능으로 생성된 파일의 이름 맨 뒤에 "작성용"으로 되어 있습니다.<br>
        이 파일에 작성된 학생들을 복사해서 "제출용"파일에 붙여넣기 합니다. 제출용 파일을 경기도청에 제출합니다.</span><br>
        """

        ending_3_get_picture = """<br><b style='font-size: 18px;'>합격자 사진 모으기</b>
        <br><b>합격자 사진 주민번호로 모으기</b><br>
        합격자의 사진의 이름을 주민등록번호로 변경하여 사진을 모아야 합니다. 해당 작업은 [File - 국시원 - 사진 모으기] 기능을 통해 해결할 수 있습니다.<br>
        <b style='color: red; font-size: 20px;'>(주의!)불합격한 사람은 DBMS에서 시험회차를 변경한 후에 위 기능을 수행해야 합니다! 변경하지 않고 실행하면 나중에 찾아서 지워야 하는데 쉽지 않습니다!</b><br>
        <b style='color: red; font-size: 20px; text-decoration: underline;'>(중요!) 시험을 이전 회차에 응시하고 합격했지만, 건강검진 등의 이유로 자격증 신청이 미루어진 사람은 직접 사진을 옮기고 이름을 주민번호로 변경해야 합니다!</b><br>
        """

        ending_4_print = """<br><b style='font-size: 18px;'>서류 출력</b>
        <br><b>교육수료증명서, 대체실습확인서, 요양보호사 자격증 발급, 재발급, 신청서 출력</b><br>
        <br>다음으로 위에서 작성한 학생의 서류를 출력해야 합니다. [File - 국시원 - 서류 출력] 기능을 통해 모두의 서류를 출력할 수 있습니다.<br>
        <b style='color: blue;'>출력하기 전에, DBMS에서 시험회차를 검색하여 임의로 추가된 사람이 없는지 먼저 확인하고, 프린터기가 이면지가 아닌 A4용지를 사용하는지 필히 확인해주세요!</b><br>
        <b style='color: red; font-size: 20px;'>(주의!)불합격한 사람은 DBMS에서 시험회차를 변경한 후에 위 기능을 수행해야 합니다! 변경하지 않고 실행하면 불합격한 사람도 모두 출력됩니다!</b><br>
        <b style='color: red; font-size: 20px; text-decoration: underline;'>(중요!) 시험을 이전 회차에 응시하고 합격했지만, 건강검진 등의 이유로 자격증 신청이 미루어진 사람은 직접 프린트 해야합니다!</b><br>
        """

        ending_5_final = """<br><b style='font-size: 18px;'>검토</b>
        <br><b>인원수 체크</b><br>
        마지막으로 합격자 명단의 명 수, 사진의 개수, 출력된 서류의 개수를 모두 비교하여 수가 맞는지 확인합니다. 이 작업을 위해 <span style='color: red; font-size: 20px;'>합격자 명단을 정확하게 작성해야 합니다.</span>합격자 명단이 비교의 기준이 됩니다.<br>
        <b style='color: red;'>서류의 개수가 맞고, 누락되거나 강제로 추가된 인원은 없는지 확인</b>하고 경기도청에 제출합니다.<br>
        """

        self.text_list = []

        self.text_list.append(ending_title)
        self.text_list.append(file_lotation)
        self.text_list.append(ending_0_summary)
        self.text_list.append(ending_1_exam)
        self.text_list.append(ending_2_pass)
        self.text_list.append(ending_3_get_picture)
        self.text_list.append(ending_4_print)
        self.text_list.append(ending_5_final)

        for i in range(len(self.text_list)):
            self.text_list[i] = "<p style='font-size: 15px;'>" + self.text_list[i] + "</p>"
            self.page_list[4] += self.text_list[i]

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()


class ToDoList(QWidget):
    global db

    def __init__(self):
        super().__init__()

        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.setWindowTitle("업무 보고")
        self.setGeometry(700, 300, 200, 400)

        self.main_box = QHBoxLayout()
        self.deadline_box = QVBoxLayout()
        self.schedule_box = QVBoxLayout()
        self.line_frame = QFrame()
        self.line_frame.setFrameShape(QFrame.VLine)
        self.line_frame.setFrameShadow(QFrame.Plain)
        self.line_frame.setLineWidth(2)

        self.main_box.addLayout(self.deadline_box)
        self.main_box.addWidget(self.line_frame)        
        self.main_box.addLayout(self.schedule_box)

        # self.main_box.setAlignment(Qt.AlignTop)
        self.deadline_box.setAlignment(Qt.AlignTop)
        self.schedule_box.setAlignment(Qt.AlignTop)
        self.setLayout(self.main_box)

        self.base_path = "D:\\남양노아요양보호사교육원"

        self.deadline_label_list = []
        self.schedule_label_list = []

        self.doc_type_list = ["개강보고", "대체실습 실시보고", "대체실습 수료보고", "응시원서 접수시작", "응시원서 접수마감", "응시표 출력", "시험 합격자 서류"]
        self.todo_dict = {"개강보고":"시간표 받기 & 검토 ➜ 수강료 수납대장 & 사물함 주기 <br>➜ 커리큘럼 만들기 ➜ 개강보고(D-day) ➜ 출석부 만들기", 
        "대체실습 실시보고":"DBMS에 대체실습 기수 입력 ➜ 대체실습 일정(명단) 만들기 + 기존 인원 추가 <br>➜ 대체실습 실시 보고서 작성 & 검토 + 기존 인원 추가! ➜ 대체실습 출석부 생성 + 기존 인원 추가",
        "대체실습 수료보고":"대체실습 보고서(사진) 만들기 ➜ 대체실습 수료보고",
        "응시원서 접수시작":"DBMS에 시험 회차 입력 및 변경 ➜ 응시접수 명단 생성 <br>➜ ID, PW 있는 사람 조사 후 기입 ➜ 회원가입 후 ID, PW 입력",
        "응시원서 접수마감":"응시 접수 & 가상계좌 응시접수 명단에 기입 ➜ 입금 후 입금확인 여부 기입<br> ➜ 응시 수수료 감면 대상자 체크 후 환불",
        "응시표 출력":"응시표 및 코로나 자가문진표 출력 후 배부",
        "시험 합격자 서류":"학생 서류 작성 ➜ 합격자 명단 작성 및 옮기기 + 지난회차 합격자 중 자격증 미신청자 입력 <br>➜ 합격자 사진 모으기 ➜ 학생 서류 출력"}
        self.deadline_dict = {}
        self.deadline_priority_list = []
        self.schedule_dict = {}
        self.schedule_priority_list = []

        self.label_deadline = QLabel("마감 임박 업무", self)
        self.label_deadline.setStyleSheet("font-size: 20px; font-weight: bold; color: green; text-decoration: underline;")
        #  border-style: solid; border-width: 2px;
        self.deadline_box.addWidget(self.label_deadline)
        self.label_schedule = QLabel("남은 일정(~D-100)", self)
        self.label_schedule.setStyleSheet("font-size: 20px; font-weight: bold; color: green; text-decoration: underline;")
        self.schedule_box.addWidget(self.label_schedule)

        self.initUI()
        # self.show()

    def exam_pass_list(self, exam_round):
        db.main.auto.makeDocument("교육수료증명서", exam_round)
        db.main.auto.makeDocument("대체실습확인서", exam_round)
        db.main.auto.makeDocument("요양보호사 자격증 발급,재발급 신청서", exam_round)

    def initUI(self):
        for doc_type in self.doc_type_list:
            res = db.main.dbPrograms.dDayCheck(doc_type, isDeadline=True)

            if not res:
                pass
            else:
                for rs in res:
                    if doc_type == "개강보고":
                        name = str(rs[0]) + str(rs[1]) + " 개강보고"
                        due_date = rs[2] + datetime.timedelta(days=2)
                        self.deadline_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        dday = rs[-1] - 2
                        
                    elif doc_type == "대체실습 실시보고":
                        name = "대체실습 " + rs[0] + " 실시보고"
                        week_day = rs[1].weekday()
                        if week_day >= 5:
                            # 주말에 개강 시, 금요일에 보고
                            day = week_day - 4
                            due_date = rs[1] - datetime.timedelta(days=day)
                            self.deadline_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        else:
                            # 주중에 개강 시, 개강 전날에 보고
                            day = 1
                            due_date = rs[1] - datetime.timedelta(days=day)
                            self.deadline_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}

                        dday = rs[-1] + day

                    elif doc_type == "대체실습 수료보고":
                        name = "대체실습 " + rs[0] + " 수료보고"
                        week_day = rs[2].weekday()
                        if week_day >= 5:
                            # 주말에 종강 시, 다음 주 월요일에 보고
                            day = int((week_day - 7) * -1)
                            due_date = rs[2] + datetime.timedelta(days=day)
                            self.deadline_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        else:
                            # 주중에 종강 시, 종강 다음날에 보고
                            day = 1
                            due_date = rs[2] + datetime.timedelta(days=day)
                            self.deadline_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}

                        dday = rs[-1] - day

                    elif doc_type == "응시원서 접수시작":
                        name = str(rs[0]) + "회 응시원서 접수시작"
                        self.deadline_dict[name] = {"마감일자":rs[1].strftime("%m월 %d일")}
                        dday = rs[-1]

                    elif doc_type == "응시원서 접수마감":
                        name = str(rs[0]) + "회 응시원서 접수마감"
                        self.deadline_dict[name] = {"마감일자":rs[2].strftime("%m월 %d일")}
                        dday = rs[-1]

                    elif doc_type == "응시표 출력":
                        name = str(rs[0]) + "회 응시표 출력(D+2까지 출력 가능!)"
                        self.deadline_dict[name] = {"마감일자":rs[3].strftime("%m월 %d일")}
                        dday = rs[-1]

                    elif doc_type == "시험 합격자 서류":
                        name = str(rs[0]) + "회 합격자 서류"
                        self.deadline_dict[name] = {"마감일자":rs[6].strftime("%m월 %d일")}
                        dday = rs[-1]
                        path = self.base_path + "\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용".format(str(rs[0]))
                        write_file_path = path + "\\화성시-남양노아요양보호사교육원-{}회합격자명단_작성용.xlsx".format(str(rs[0]))
                        submit_file_path = path + "\\화성시-남양노아요양보호사교육원-{}회합격자명단_작성용_제출용.xls".format(str(rs[0]))
                        if not os.path.exists(path):
                            os.makedirs(path)
                        
                        if not os.path.exists(write_file_path):
                            origin_file = "D:\\Master\\files\\화성시-남양노아요양보호사교육원-00회합격자명단_작성용.xlsx"
                            shutil.copy(origin_file, write_file_path)

                        if not os.path.exists(submit_file_path):
                            origin_file = "D:\\Master\\files\\화성시-남양노아요양보호사교육원-00회합격자명단_제출용.xls"
                            shutil.copy(origin_file, submit_file_path)

                    self.deadline_dict[name]["D-day"] = dday
                    self.deadline_dict[name]["documentType"] = doc_type

        if len(self.deadline_dict) != 0:
            for i in range(len(self.deadline_dict) * 4):
                self.deadline_label_list.append(QLabel(self))
                self.deadline_label_list[i].setStyleSheet("font-size: 15px;")

            # 마감이 임박한 업무들을 앞으로 정렬
            self.deadline_priority_list = sorted(self.deadline_dict, key=lambda name: self.deadline_dict[name]["D-day"], reverse=True)

            for idx, name in enumerate(self.deadline_priority_list):
                idx *= 4
                self.deadline_label_list[idx].setText(name + "<b style='color: blue;'> " + self.deadline_dict[name]["마감일자"] + " </b>")

                self.deadline_label_list[idx + 1].setText("D-day : D" + str(self.deadline_dict[name]["D-day"]))

                color = "blue"
                if int(self.deadline_dict[name]["D-day"]) < 0:
                    if int(self.deadline_dict[name]["D-day"]) > -3:
                        color = "red"
                    self.deadline_dict[name]["D-day"] = "D" + str(self.deadline_dict[name]["D-day"])
                elif int(self.deadline_dict[name]["D-day"]) == 0:
                    self.deadline_dict[name]["D-day"] = "D-day!"
                elif int(self.deadline_dict[name]["D-day"]) > 0:
                    color = "purple"
                    self.deadline_dict[name]["D-day"] = "D+" + str(self.deadline_dict[name]["D-day"])

                self.deadline_label_list[idx + 1].setText("D-day : " + self.deadline_dict[name]["D-day"])
                self.deadline_label_list[idx + 1].setStyleSheet("font-weight: bold; font-size: 15px; color: {};".format(color))

                self.deadline_label_list[idx + 2].setText("<b style='font-size: 12px;'>ToDo: " + self.todo_dict[self.deadline_dict[name]["documentType"]] + "</b>")

                # 마지막 QFrame 은 넣지 않는다.
                if idx // 4 == len(self.deadline_priority_list) - 1:
                    continue
                # label list 각 순번의 마지막은 QFrame 으로 변경
                self.deadline_label_list[idx + 3] = QFrame()
                self.deadline_label_list[idx + 3].setFrameShape(QFrame.HLine)
                self.deadline_label_list[idx + 3].setFrameShadow(QFrame.Plain)
                self.deadline_label_list[idx + 3].setLineWidth(1)

            for lbl in self.deadline_label_list:
                self.deadline_box.addWidget(lbl)

        else:
            self.deadline_label_list.append(QLabel(self))
            self.deadline_label_list[0].setStyleSheet("font-size: 15px; font-weight: bold;")
            self.deadline_label_list[0].setText("마감 임박 일정이 존재하지 않습니다.")

            self.deadline_box.addWidget(self.deadline_label_list[0])

        self.schedule_box.addWidget(self.label_schedule)

        for doc_type in self.doc_type_list:
            res = db.main.dbPrograms.dDayCheck(doc_type, isDeadline=False)

            if not res:
                pass
            else:
                for rs in res:
                    if doc_type == "개강보고":
                        name = str(rs[0]) + str(rs[1]) + " 개강보고"
                        due_date = rs[2] + datetime.timedelta(days=2)
                        self.schedule_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        dday = rs[-1] - 2
                        
                    elif doc_type == "대체실습 실시보고":
                        name = "대체실습 " + rs[0] + " 실시보고"
                        week_day = rs[1].weekday()
                        if week_day >= 5:
                            # 주말에 개강 시, 금요일에 보고
                            day = week_day - 4
                            due_date = rs[1] - datetime.timedelta(days=day)
                            self.schedule_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        else:
                            # 주중에 개강 시, 개강 전날에 보고
                            day = 1
                            due_date = rs[1] - datetime.timedelta(days=day)
                            self.schedule_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}

                        dday = rs[-1] + day

                    elif doc_type == "대체실습 수료보고":
                        name = "대체실습 " + rs[0] + " 수료보고"
                        week_day = rs[2].weekday()
                        if week_day >= 5:
                            # 주말에 종강 시, 다음 주 월요일에 보고
                            day = int((week_day - 7) * -1)
                            due_date = rs[2] + datetime.timedelta(days=day)
                            self.schedule_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}
                        else:
                            # 주중에 종강 시, 종강 다음날에 보고
                            day = 1
                            due_date = rs[2] + datetime.timedelta(days=day)
                            self.schedule_dict[name] = {"마감일자":due_date.strftime("%m월 %d일")}

                        dday = rs[-1] - day

                    elif doc_type == "응시원서 접수시작":
                        name = str(rs[0]) + "회 응시원서 접수시작"
                        self.schedule_dict[name] = {"마감일자":rs[1].strftime("%m월 %d일")}
                        dday = rs[-1]

                    elif doc_type == "응시원서 접수마감":
                        name = str(rs[0]) + "회 응시원서 접수마감"
                        self.schedule_dict[name] = {"마감일자":rs[2].strftime("%m월 %d일")}
                        dday = rs[-1]

                    elif doc_type == "응시표 출력":
                        name = str(rs[0]) + "회 응시표 출력"
                        self.schedule_dict[name] = {"마감일자":rs[3].strftime("%m월 %d일")}
                        dday = rs[-1] - 2

                    elif doc_type == "시험 합격자 서류":
                        name = str(rs[0]) + "회 합격자 서류"
                        self.schedule_dict[name] = {"마감일자":rs[6].strftime("%m월 %d일")}
                        dday = rs[-1]

                    self.schedule_dict[name]["D-day"] = dday

        for i in range(len(self.schedule_dict)):
            self.schedule_label_list.append(QLabel(self))
            self.schedule_label_list[i].setStyleSheet("font-size: 15px;")

        self.schedule_priority_list = sorted(self.schedule_dict, key=lambda name: self.schedule_dict[name]["D-day"], reverse=True)

        for idx, name in enumerate(self.schedule_priority_list):
            self.schedule_label_list[idx].setText(name + "<b> " + self.schedule_dict[name]["마감일자"] + " </b><b style='color: blue;'> D-day: D" + str(self.schedule_dict[name]["D-day"]) + " </b>")

        for lbl in self.schedule_label_list:
            self.schedule_box.addWidget(lbl)

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()

    def showEvent(self, QShowEvent):
        pass

class MemberManagement(QWidget):
    def __init__(self):
        super().__init__()
        self.initUI()

    def initUI(self):
        self.table = QTreeView(self)
        self.table.setAlternatingRowColors(True)
        self.table.setRootIsDecorated(False)

        """
        좌측 - 테이블
        우측 - 출력 멤버
        중앙 - 추가, 모두추가
        우측 하단 - 가위표(삭제)(클릭 - index를 가져옴 - 해당 인덱스 멤버 삭제)
        """
        
        # QTreeView set read only
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.readDB = QtGui.QStandardItemModel(0, 1, self)
        # Qt.Horizontal: 수평값. 기본적으로 넣어야 함.
        self.readDB.setHeaderData(0, Qt.Horizontal, "선택")

        self.table.setModel(self.readDB)
        self.table.clicked.connect(self.selected)

        self.layoutInfo = QVBoxLayout()
        self.textInfo = QTextEdit()
        self.textInfo.setReadOnly(True)
        self.textInfo.setFixedWidth(400)
        # self.textInfo.setFontPointSize(12)
        self.textInfo.setCurrentFont(QtGui.QFont("맑은 고딕"))
        self.layoutInfo.addWidget(self.textInfo)

class DocumentChecker(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.setWindowTitle("수강생 파일 체크")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

        self.non_doc_dict = {}
        self.checker_dict = {"사진": False, "주민등록등본": False, "기본증명서":False, "자격증": False}

        self.scroll_area = QScrollArea()
        self.label = QLabel(self)
        self.scroll_area.setWidget(self.label)

    def initUI(self):
        pass

    def checkFiles(self):
        dict_ready_res = db.main.dbPrograms.SELECT("classNumber, classTime", "lecture", orderBy="classNumber *1")
        for rows in dict_ready_res:
            self.non_doc_dict[str(rows[0]) + str(rows[1])] = []

        res = db.main.dbPrograms.SELECT("name, classNumber, classTime, license", "user", orderBy="classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1")
        for rows in res:
            file_path = f"D:\\남양노아요양보호사교육원\\교육생관리\\{rows[1]}\\{rows[1]}{rows[2]}\\{rows[0]}"
            if not os.path.exists(file_path):
                self.non_doc_dict[str(rows[1]) + str(rows[2])].append("{} / 폴더 전체".format(rows[0]))
                continue

            check_list = os.listdir(file_path)
            for file in check_list:
                file_name, file_extension = os.path.splitext(file)

class Calender(QWidget):
    global db
    def __init__(self):
        super().__init__()
        self.setWindowTitle("실습 날짜 설정")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

        self.selectedDate = {}

        self.main_box = QVBoxLayout()
        self.calender = QCalendarWidget()
        self.label = QLabel()
        self.btn = QPushButton("입력", self)

        self.calender.setGridVisible(True)
        self.calender.setVerticalHeaderFormat(False)

        self.main_box.addWidget(self.label)
        self.main_box.addWidget(self.calender)
        self.main_box.addWidget(self.btn)
        self.setLayout(self.main_box)

        self.fm_selected = QTextCharFormat()
        self.fm_selected.setForeground(Qt.red)
        self.fm_selected.setBackground(Qt.yellow)

        self.fm_origin = QTextCharFormat()

        self.initUI()

    def initUI(self):
        self.calender.clicked.connect(self.showDate)
        self.btn.clicked.connect(self.returnDate)

    def showDate(self, date):
        dateList = date.toString().split(" ")
        # date.toString() -> 수 6 22 2022
        selected_year = dateList[3]
        selected_month = dateList[1]
        selected_date = dateList[2]
        selected_days = dateList[0]
        if selected_year not in self.selectedDate.keys():
            self.selectedDate[selected_year] = []

        dateInfo = f"{selected_month.rjust(2, '0')}/{selected_date.rjust(2, '0')}"

        if dateInfo not in self.selectedDate[selected_year]:
            self.calender.setDateTextFormat(date, self.fm_selected)
            self.selectedDate[selected_year].append(dateInfo)
        else:
            self.calender.setDateTextFormat(date, self.fm_origin)
            self.selectedDate[selected_year].remove(dateInfo)

        self.selectedDate = dict(sorted(self.selectedDate.items(), key = lambda item: item[0]))
        for year in self.selectedDate.keys():
            self.selectedDate[year].sort()

        myStr = ""
        for year in self.selectedDate.keys():
            myStr += year + ": "
            myStr += " | ".join(self.selectedDate[year])
            myStr += "\n"
        print(f"date.year: {date.year()} / date.month:{date.month()} / date.getDate(): {date.getDate()} / date.day:{date.day()} / date.dayOfWeek:{date.dayOfWeek()}")
        print(self.selectedDate)
        self.label.setText(myStr)

    def returnDate(self):
        # 이거 작동 안됨
        return self.selectedDate

class Kuksiwon(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.setWindowTitle("국시원")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

        self.doc_type = ""

        self.main_box = QVBoxLayout()
        self.input_box = QHBoxLayout()
        self.btn_box = QHBoxLayout()

        self.main_box.addLayout(self.input_box)
        self.main_box.addLayout(self.btn_box)

        self.setLayout(self.main_box)

        self.initUI()

    def initUI(self):
        self.label_exam = QLabel("시험회차", self)
        self.input_box.addWidget(self.label_exam)
        self.combobox_exam = QComboBox(self)
        self.input_box.addWidget(self.combobox_exam)
        self.btn_create = QPushButton("생성", self)
        self.btn_cancel = QPushButton("취소", self)

        self.btn_create.clicked.connect(self.createFile)
        self.btn_cancel.clicked.connect(self.close)

        self.btn_box.addWidget(self.btn_create)
        self.btn_box.addWidget(self.btn_cancel)

    def createFile(self):
        exam_round = self.combobox_exam.currentText()
        if self.combobox_exam == "선택":
            QMessageBox.about(self, "안내", "옵션이 선택되지 않았습니다.")
            return

        db.logger.info("$UI KUKSIWON Request [REQUEST|{}][EXAM|{}회] 수행 요청".format(self.doc_type, exam_round))
        if "출력" in self.doc_type:
            ans = QMessageBox.question(self, "확인", "{}회 합격자 {}를 출력합니다.".format(exam_round, self.doc_type[3:]), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                non_list = db.main.auto.printDocument(exam_round, self.doc_type[3:])
                if non_list[:9] == "Traceback":
                    QMessageBox.about(self, "Traceback", non_list)
                else:
                    QMessageBox.about(self, "완료", "문서가 출력되었습니다.\n파일 에러: {}\n\n*파일에러: 파일이 존재하지 않거나 문서 이름이 정확하지 않습니다.".format(non_list))

        elif self.doc_type == "합격자 사진":
            ans = QMessageBox.question(self, "확인", "{}회 합격자 사진을 수집합니다.".format(exam_round), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                path, non_list = db.main.auto.gatherPictures(exam_round)
                if path[:9] == "Traceback" and non_list == "ERROR":
                    QMessageBox.about(self, "Traceback", path)
                else:
                    QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n경로: {}\n파일 에러: {}\n\n*파일에러: 파일이 존재하지 않거나 문서 이름이 정확하지 않습니다.".format(path, non_list))

        else:
            ans = QMessageBox.question(self, "확인", "{}회 {} 파일을 생성합니다.".format(exam_round, self.doc_type), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                if self.doc_type == "응시접수 명단":
                    path = db.main.auto.accountList(exam_round)
                    if path[:9] == "Traceback":
                        QMessageBox.about(self, "Traceback", path)
                    else:
                        QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n경로: {}".format(path))

                elif self.doc_type == "합격자 명단":
                    non_input_list = db.main.auto.examPassList(exam_round)
                    if non_input_list[:9] == "Traceback":
                        QMessageBox.about(self, "Traceback", non_input_list)
                    else:
                        QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n{}".format(non_input_list))

                else:
                    non_input_list = db.main.auto.makeDocument(exam_round, self.doc_type)
                    if non_input_list[:9] == "Traceback":
                        QMessageBox.about(self, "Traceback", non_input_list)
                    else:
                        QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n{}".format(non_input_list))
                
        self.close()
            

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.createFile()

    def showEvent(self, QShowEvent):
        source = self.sender()
        self.doc_type = source.text()
        self.combobox_exam.clear()
        self.exam_list = []

        self.combobox_exam.addItem("선택")
        rs = db.main.dbPrograms.SELECT("round", "exam", where="TIMESTAMPDIFF(DAY, passDate, CURDATE()) < 90", orderBy="round *1")

        if rs == "error":
            QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
        else:
            for row in rs:
                if not row[0] in self.exam_list:
                    self.exam_list.append(str(row[0]))
            
            self.combobox_exam.addItems(self.exam_list)


class ClassOpening(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.setWindowTitle("요양보호사 기수 opening")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

        self.doc_type = ""

        self.main_box = QVBoxLayout()
        self.input_box = QHBoxLayout()
        self.btn_box = QHBoxLayout()

        self.main_box.addLayout(self.input_box)
        self.main_box.addLayout(self.btn_box)

        self.setLayout(self.main_box)

        self.initUI()

    def initUI(self):
        self.label_number = QLabel("기수", self)
        self.input_box.addWidget(self.label_number)
        self.combobox_N = QComboBox(self)
        self.combobox_N.setFixedWidth(100)
        self.input_box.addWidget(self.combobox_N)

        self.label_time = QLabel("반", self)
        self.input_box.addWidget(self.label_time)
        self.combobox_T = QComboBox(self)
        self.combobox_T.setFixedWidth(100)
        self.input_box.addWidget(self.combobox_T)

        self.btn_create = QPushButton("생성", self)
        self.btn_box.addWidget(self.btn_create)
        self.btn_cancel = QPushButton("취소", self)
        self.btn_box.addWidget(self.btn_cancel)

        self.btn_create.clicked.connect(self.createFile)
        self.btn_cancel.clicked.connect(self.close)

    def createFile(self):
        class_number = self.combobox_N.currentText()
        class_time = self.combobox_T.currentText()

        if class_number == "선택" or class_time == "선택":
            QMessageBox.about(self, "안내", "옵션이 선택되지 않았습니다.")
            return

        if self.doc_type == "수강료 수납대장":
            ans = QMessageBox.question(self, "확인", "{}기{} 수강료 수납대장 파일을 생성합니다.".format(class_number, class_time), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                path = db.main.auto.paymentList(class_number, class_time)
                if path[:9] == "Traceback":
                    QMessageBox.about(self, "Traceback", path)
                else:
                    QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n경로: {}".format(path))

        elif self.doc_type == "사물함 주기":
            ans = QMessageBox.question(self, "확인", "{}기{} 사물함 주기 파일을 생성합니다.".format(class_number, class_time), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                path = db.main.auto.locker(class_number, class_time)
                if path[:9] == "Traceback":
                    QMessageBox.about(self, "Traceback", path)
                else:
                    QMessageBox.about(self, "완료", "파일이 생성되었습니다.\n경로: {}".format(path))

        self.close()
        


    def showEvent(self, QShowEvent):
        source = self.sender()
        self.doc_type = source.text()
        self.combobox_N.clear()
        self.combobox_T.clear()
        self.class_num_list = []

        self.combobox_N.addItem("선택")
        self.combobox_T.addItem("선택")
        self.combobox_T.addItem("주간")
        self.combobox_T.addItem("야간")
        rs = db.main.dbPrograms.SELECT("classNumber", "lecture", where="TIMESTAMPDIFF(DAY, startDate, CURDATE()) < 60", orderBy="classNumber *1")

        if rs == "error":
            QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
        else:
            for row in rs:
                if not row[0] in self.class_num_list:
                    self.class_num_list.append(row[0])
            
            self.combobox_N.addItems(self.class_num_list)

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.createFile()



class Report(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.setWindowTitle("경기도청 보고 데이터")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.doc_type = ""

    def initUI(self):
        self.box = QVBoxLayout()
        self.setLayout(self.box)
        self.hbox1 = QHBoxLayout()
        self.hbox2 = QHBoxLayout()
        self.box.addLayout(self.hbox1)
        self.box.addLayout(self.hbox2)

        self.label_number = QLabel("기수", self)
        self.hbox1.addWidget(self.label_number)
        self.combobox_N = QComboBox(self)
        self.combobox_N.setFixedWidth(100)
        self.hbox1.addWidget(self.combobox_N)

        self.label_time = QLabel("반", self)
        self.label_time.setFixedWidth(48)
        self.label_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.hbox1.addWidget(self.label_time)
        self.combobox_T = QComboBox(self)
        self.combobox_T.setFixedWidth(100)
        self.hbox1.addWidget(self.combobox_T)

        self.hbox2.addStretch(1)
        self.btn_create = QPushButton("생성", self)
        self.hbox2.addWidget(self.btn_create)
        self.btn_cancel = QPushButton("취소", self)
        self.hbox2.addWidget(self.btn_cancel)

        self.btn_create.clicked.connect(self.getData)
        self.btn_cancel.clicked.connect(self.close)

    def getData(self):
        class_number = self.combobox_N.currentText()
        class_time = self.combobox_T.currentText()

        if class_number == "선택" or class_time == "선택":
            QMessageBox.about(self, "안내", "옵션이 선택되지 않았습니다.")
            return

        ans = QMessageBox.question(self, "확인", "{}기{} {} 데이터를 생성합니다.".format(class_number, class_time, self.doc_type), QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            QMessageBox.about(self, "안내", "OK버튼을 눌러 작업을 진행해 주세요.\n생성이 완료되면 엑셀 파일이 열립니다.")
            res = db.main.auto.report(self.doc_type, class_number, class_time)
            if res[:9] == "Traceback":
                QMessageBox.about(self, "Traceback", res)

        self.close()


    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.getData()

    def showEvent(self, QShowEvent):
        self.setWindowTitle("경기도청 " + self.doc_type)
        self.combobox_N.clear()
        self.combobox_T.clear()
        self.class_num_list = []

        """
        개강보고
        대체실습 실시보고
        대체실습 수료보고
        """

        if self.doc_type == "개강보고":
            self.label_time.setText("반")
            self.combobox_T.setEnabled(True)
            self.combobox_N.addItem("선택")
            self.combobox_T.addItem("선택")
            self.combobox_T.addItem("주간")
            self.combobox_T.addItem("야간")
            rs = db.main.dbPrograms.SELECT("classNumber", "lecture", where="TIMESTAMPDIFF(DAY, startDate, CURDATE()) < 60", orderBy="classNumber *1")

        elif self.doc_type == "출석부":
            self.label_time.setText("반")
            self.combobox_T.setEnabled(True)
            self.combobox_N.addItem("선택")
            self.combobox_T.addItem("선택")
            self.combobox_T.addItem("주간")
            self.combobox_T.addItem("야간")
            rs = db.main.dbPrograms.SELECT("classNumber", "lecture", where="TIMESTAMPDIFF(DAY, startDate, CURDATE()) < 60", orderBy="classNumber *1")

        else:
            self.label_time.setText("대체실습")
            self.combobox_T.setEnabled(False)
            rs = db.main.dbPrograms.SELECT("classNumber", "temptraining", where="TIMESTAMPDIFF(DAY, endDate, CURDATE()) < 60", orderBy="classNumber *1")
        
        if rs == "error":
            QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
        else:
            for row in rs:
                if not row[0] in self.class_num_list:
                    self.class_num_list.append(row[0])
            
            self.combobox_N.addItems(self.class_num_list)
        
        


class scanFile(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.setWindowTitle("파일 스캔")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.file_list = []
        self.file_index = 0
        
    def initUI(self):
        self.grid = QGridLayout()
        self.setLayout(self.grid)
        self.target_table = "user"
        cnt_row = 5
        cnt_col = 7
        self.resize(600, 400)

        self.labelImg = QLabel(self)
        self.labelImg.setFixedSize(500, 600)
        self.grid.addWidget(self.labelImg, 0, 0, cnt_row, 1)
        self.label_id_user = QLabel("ID", self)
        self.label_id_user.setFixedWidth(90)
        self.label_id_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_id_user, 0, 1)
        self.text_id_user = QLineEdit()
        self.grid.addWidget(self.text_id_user, 0, 2)
        self.label_name_user = QLabel("이름", self)
        self.label_name_user.setFixedWidth(90)
        self.label_name_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_name_user, 0, 3)
        self.text_name_user = QLineEdit()
        self.grid.addWidget(self.text_name_user, 0, 4)
        self.label_licen_user = QLabel("자격증", self)
        self.label_licen_user.setFixedWidth(90)
        self.label_licen_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_licen_user, 0, 5)
        self.text_licen_user = QLineEdit()
        self.grid.addWidget(self.text_licen_user, 0, 6)

        self.label_clsN_user = QLabel("기수", self)
        self.label_clsN_user.setFixedWidth(90)
        self.label_clsN_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_clsN_user, 1, 1)
        self.text_clsN_user = QLineEdit()
        self.grid.addWidget(self.text_clsN_user, 1, 2)
        self.label_clsT_user = QLabel("반", self)
        self.label_clsT_user.setFixedWidth(90)
        self.label_clsT_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_clsT_user, 1, 3)
        self.text_clsT_user = QLineEdit()
        self.grid.addWidget(self.text_clsT_user, 1, 4)
        self.label_temp = QLabel("대체실습", self)
        self.label_temp.setFixedWidth(90)
        self.label_temp.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_temp, 1, 5)
        self.text_temp = QLineEdit()
        self.grid.addWidget(self.text_temp, 1, 6)

        self.label_RRN = QLabel("주민등록번호", self)
        self.label_RRN.setFixedWidth(90)
        self.label_RRN.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_RRN, 2, 1)
        self.text_RRN = QLineEdit()
        self.grid.addWidget(self.text_RRN, 2, 2, 1, 2)
        self.label_phone = QLabel("전화번호", self)
        self.label_phone.setFixedWidth(90)
        self.label_phone.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_phone, 2, 4)
        self.text_phone = QLineEdit()
        self.grid.addWidget(self.text_phone, 2, 5, 1, 2)
        
        
        self.label_adr = QLabel("도로명주소", self)
        self.label_adr.setFixedWidth(90)
        self.label_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_adr, 3, 1)
        self.text_adr = QLineEdit()
        self.grid.addWidget(self.text_adr, 3, 2, 1, 5)
        self.label_origin_adr = QLabel("본적주소", self)
        self.label_origin_adr.setFixedWidth(90)
        self.label_origin_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
        self.grid.addWidget(self.label_origin_adr, 4, 1)
        self.text_origin_adr = QLineEdit()
        self.grid.addWidget(self.text_origin_adr, 4, 2, 1, 5)

        self.btn_insert = QPushButton("Insert", self)
        self.btn_insert.clicked.connect(self.scanner)
        self.btn_cancel = QPushButton("Close", self)
        self.btn_cancel.clicked.connect(self.close)

        self.grid.addWidget(self.btn_insert, cnt_row, cnt_col - 2)
        self.grid.addWidget(self.btn_cancel, cnt_row, cnt_col - 1)

    def refreshUI(self):
        print("self.file_list")
        print(self.file_list)

        file_name = self.file_list[self.file_index]
        pixmap = QPixmap(file_name)
        pixmap = pixmap.scaledToWidth(500)

        self.labelImg.setPixmap(QPixmap(pixmap))
        self.text_id_user.clear()
        self.text_name_user.clear()
        self.text_licen_user.clear()
        self.text_clsN_user.clear()
        self.text_clsT_user.clear()
        self.text_temp.clear()
        self.text_RRN.clear()
        self.text_phone.clear()
        self.text_adr.clear()
        self.text_origin_adr.clear()

        doc_type = "주민등록등본"
        name = "name"
        adr = "주소"
        origin_adr = "본적주소"

        self.text_name_user.setText(name)
        self.text_adr.setText(adr)
        self.text_origin_adr.setText(origin_adr)

        print("self.text_id_user.text()")
        print("\"" + self.text_id_user.text() + "\"")
        print(type(self.text_id_user.text()))
        print("self.text_name_user.text()")
        print("\"" + self.text_name_user.text() + "\"")
        print(type(self.text_name_user.text()))
        print("self.text_licen_user.text()")
        print("\"" + self.text_licen_user.text() + "\"")
        print(type(self.text_licen_user.text()))
        print("self.text_clsN_user.text()")
        print("\"" + self.text_clsN_user.text() + "\"")
        print(type(self.text_clsN_user.text()))
        print("self.text_clsT_user.text()")
        print("\"" + self.text_clsT_user.text() + "\"")
        print(type(self.text_clsT_user.text()))
        print("self.text_temp.text()")
        print("\"" + self.text_temp.text() + "\"")
        print(type(self.text_temp.text()))
        print("self.text_RRN.text()")
        print("\"" + self.text_RRN.text() + "\"")
        print(type(self.text_RRN.text()))
        print("self.text_phone.text()")
        print("\"" + self.text_phone.text() + "\"")
        print(type(self.text_phone.text()))
        print("self.text_adr.text()")
        print("\"" + self.text_adr.text() + "\"")
        print(type(self.text_adr.text()))
        print("self.text_origin_adr.text()")
        print("\"" + self.text_origin_adr.text() + "\"")
        print(type(self.text_origin_adr.text()))

    def scanner(self):
        rs = db.main.dbPrograms.SELECT("*", "user", "name='' and RRN=''")
        print("rs")
        print("\"" + str(rs) + "\"")

        self.file_index += 1
        self.refreshUI()

    def showEvent(self, QShowEvent):
        self.file_index = 0
        self.refreshUI()
        return
        """
        doc_name, name, adr, origin_adr 받아서 text.setText 하기!
        """
        self.file = file_name
        img_origin = Image.open(r"C:\Users\David\Desktop\SKM_C364e21121319310_0002.jpg")
        # img_origin.show() file open
        print(img_origin.size)

        #((left, up, right, bottom))
        img_cropped = img_origin.crop((0, 580, 800, 700))
        # img_cropped.show()

        import pytesseract

        pytesseract.pytesseract.tesseract_cmd = r"C:\Program Files\Tesseract-OCR\tesseract"

        image = img_cropped
        gibon = pytesseract.image_to_string(image, lang="kor", config="preserve_interword_spaces=1 --psm 4")
        print(gibon)
        str_arr = gibon.split(' ')
        print("print str_arr before")
        print(str_arr)
        res = []

        for string in str_arr:
            if '\n' in string:
                imsi = string.split('\n')
                for substr in imsi:
                    if substr != '':
                        res.append(substr)

            else:
                res.append(string)

        del res[0]
        del res[-1]
        print("res")
        print(res)

        

        

        

"""
* 일괄 변경 창 추가하기!
일괄 변경 창 클릭 -> 기수, 반 선택 후 바꿀 column 선택 -> 값 입렵 ==> 일괄변경 처리 !

"""
class BatchUpdate(QWidget):
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

        self.mode = "시험회차"

    def initUI(self):
        self.box = QVBoxLayout()
        self.setLayout(self.box)
        box_top = QHBoxLayout()
        box_middle = QHBoxLayout()
        box_bottom = QHBoxLayout()

        self.box.addLayout(box_top)
        self.box.addLayout(box_middle)
        self.box.addLayout(box_bottom)
        
        self.label_N = QLabel("기수", self)
        self.label_T = QLabel("반", self)

        self.combobox_N = QComboBox(self)
        self.combobox_N.setFixedWidth(100)
        self.combobox_T = QComboBox(self)
        self.combobox_T.setFixedWidth(100)

        box_top.addWidget(self.label_N)
        box_top.addWidget(self.combobox_N)
        box_top.addWidget(self.label_T)
        box_top.addWidget(self.combobox_T)

        self.label_exam_or_temp = QLabel("시험 회차")
        box_middle.addWidget(self.label_exam_or_temp)
        self.combobox_exam_or_temp = QComboBox(self)
        box_middle.addWidget(self.combobox_exam_or_temp)

        self.btn_update = QPushButton("일괄 변경", self)
        self.btn_update.clicked.connect(self.batch)
        self.btn_cancel = QPushButton("취소", self)
        self.btn_cancel.clicked.connect(self.close)
        box_bottom.addStretch(1)
        box_bottom.addWidget(self.btn_update)
        box_bottom.addWidget(self.btn_cancel)

    def batch(self):
        db.logger.info("$UI BATCH UPDATE Request [REQUEST|{}] 일괄 수정 요청".format(self.mode))
        if self.mode == "이수시간 일괄 변경":
            if self.combobox_N.currentText() == "선택" or self.combobox_T.currentText() == "선택":
                QMessageBox.warning(self, "오류", "입력값 오류")
        
        else:
            if self.combobox_N.currentText() == "선택" or self.combobox_T.currentText() == "선택" or self.combobox_exam_or_temp.currentText() == "선택":
                QMessageBox.warning(self, "오류", "입력값 오류")
                return

        ans = QMessageBox.question(self, "데이터 수정 확인", "이 기능은 \"데이터가 없는 사람(NULL)만\" 값을 변경해 줍니다. 변경하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if ans == QMessageBox.Yes:
            pass
        else:
            return

        if self.mode == "시험회차 일괄 변경":
            exam = self.combobox_exam_or_temp.currentText()
            exam = exam[:-1]
            query = "exam={}".format(exam)
            where = "classNumber='{}' and classTime='{}' and exam is NULL".format(self.combobox_N.currentText(), self.combobox_T.currentText())

            res = db.main.dbPrograms.UPDATE("user", query, where)
            if res == "error":
                QMessageBox.warning(self, "오류", "데이터 수정에 오류가 발생했습니다!")
            else:
                QMessageBox.about(self, "완료", "데이터를 성공적으로 수정했습니다.")
            db.main.showTable(Refresh=True)
            db.main.textInfo.clear()
            
        elif self.mode == "대체실습 일괄 변경":
            temp_class_number = self.combobox_exam_or_temp.currentText()
            query = "temporaryClassNumber='{}'".format(temp_class_number)
            where = "classNumber='{}' and classTime='{}' and temporaryClassNumber is NULL".format(self.combobox_N.currentText(), self.combobox_T.currentText())

            res = db.main.dbPrograms.UPDATE("user", query, where)
            if res == "error":
                QMessageBox.warning(self, "오류", "데이터 수정에 오류가 발생했습니다!")
            else:
                QMessageBox.about(self, "완료", "데이터를 성공적으로 수정했습니다.")
            db.main.showTable(Refresh=True)
            db.main.textInfo.clear()

        elif self.mode == "이수시간 일괄 변경":
            checker = "(totalCreditHour is NULL and theoryCreditHour is NULL and practicalCreditHour is NULL and trainingCreditHour is NULL)"
            query = "totalCreditHour=80, theoryCreditHour=80, practicalCreditHour=80, trainingCreditHour=80"
            where = "classNumber='{}' and classTime='{}'".format(self.combobox_N.currentText(), self.combobox_T.currentText())

            member = db.main.dbPrograms.SELECT("id, name, RRN, license", "user", where)
            for rows in member:
                if rows[3] == "일반":
                    query = "totalCreditHour=240, theoryCreditHour=80, practicalCreditHour=80, trainingCreditHour=80"
                elif rows[3] == "간호사":
                    query = "totalCreditHour=40, theoryCreditHour=26, practicalCreditHour=6, trainingCreditHour=8"
                elif rows[3] == "사회복지사":
                    query = "totalCreditHour=50, theoryCreditHour=32, practicalCreditHour=10, trainingCreditHour=8"
                elif rows[3] == "간호조무사" or rows[3] == "물리치료사" or rows[3] == "작업치료사":
                    query = "totalCreditHour=50, theoryCreditHour=31, practicalCreditHour=11, trainingCreditHour=8"
                else:
                    query = "totalCreditHour=NULL, theoryCreditHour=NULL, practicalCreditHour=NULL, trainingCreditHour=NULL"

                where = "id={} and name='{}' and RRN='{}' and {}".format(rows[0], rows[1], rows[2], checker)
                db.main.dbPrograms.UPDATE("user", query, where)

            QMessageBox.about(self, "완료", "데이터를 성공적으로 수정했습니다.")
            db.main.showTable(Refresh=True)
            db.main.textInfo.clear()

        self.close()


    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.batch()

    def showEvent(self, QShowEvent):
        self.setWindowTitle(self.mode)
        self.combobox_N.clear()
        self.combobox_T.clear()
        self.combobox_exam_or_temp.clear()
        self.class_num_list = []
        
        self.combobox_N.addItem("선택")
        self.combobox_T.addItem("선택")
        self.combobox_T.addItem("주간")
        self.combobox_T.addItem("야간")
        self.combobox_exam_or_temp.addItem("선택")

        rs = db.main.dbPrograms.SELECT("classNumber, classTime", "lecture", where="TIMESTAMPDIFF(DAY, startDate, CURDATE()) < 120", orderBy="classNumber *1")
        if rs == "error":
            QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
            self.close()
        else:
            for row in rs:
                if not row[0] in self.class_num_list:
                    self.class_num_list.append(row[0])
            
            self.combobox_N.addItems(self.class_num_list)

        if self.mode == "시험회차 일괄 변경":
            self.label_exam_or_temp.setText("시험 회차")
            rs = db.main.dbPrograms.SELECT("round", "exam", "TIMESTAMPDIFF(DAY, startAcceptance, CURDATE()) < 90", orderBy="round *1")
            if rs == "error":
                QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
                self.close()
            else:
                for row in rs:
                    self.combobox_exam_or_temp.addItem(str(row[0]) + "회")

        elif self.mode == "대체실습 일괄 변경":
            self.label_exam_or_temp.setText("대체실습")
            rs = db.main.dbPrograms.SELECT("classNumber", "temptraining", "TIMESTAMPDIFF(DAY, startDate, CURDATE()) < 30", orderBy="classNumber *1")
            if rs == "error":
                QMessageBox.information(self, "ERROR", "class batchUpdate returns error", QMessageBox.Yes, QMessageBox.Yes)
                self.close()
            else:
                for row in rs:
                    self.combobox_exam_or_temp.addItem(str(row[0]))

        elif self.mode == "이수시간 일괄 변경":
            self.label_exam_or_temp.setText("")
            self.combobox_exam_or_temp.setEnabled(False)

        # self.combobox_exam_or_temp.addItem("Default")

class UPDATE(QWidget):
    # 새 창을 띄우기 위해 서로 global로 연결
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.target_table = ""
        self.base_path = "D:\\남양노아요양보호사교육원\\교육생관리"
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

    def initUI(self):
        self.setWindowTitle("데이터 수정")

        self.grid = QGridLayout()
        self.setLayout(self.grid)

    def generateDirectory(self, number, time, name):
        path = self.base_path + "\\{}\\{}{}".format(number, number, time)
        if not os.path.exists(path):
            os.makedirs(path)
        
        path = path + "\\{}".format(name)

        if not os.path.exists(path):
            os.makedirs(path)

    def changeName(self, number, time, b_name, a_name):
        """폴더 이름, 사진, 기본증명서, 주민등록등본, 외국인등록증 이름 변경"""
        before_name = self.base_path + "\\{}\\{}{}\\{}".format(number, number, time, b_name)
        after_name = self.base_path + "\\{}\\{}{}\\{}".format(number, number, time, a_name)
        os.rename(before_name, after_name)

        before_files = {"사진": after_name + "\\{}{}_{}.jpg".format(number, time, b_name), "기본증명서": after_name + "\\{}_기본증명서.jpg".format(b_name), "주민등록등본": after_name + "\\{}주민등록등본.jpg".format(b_name), "외국인등록증": after_name + "\\{}_외국인등록증.jpg".format(b_name)}
        after_files = {"사진": None, "기본증명서": None, "주민등록등본": None, "외국인등록증": None}

        # 사진 존재 여부 확인
        if not os.path.exists(before_files["사진"]):
            before_files["사진"] = after_name + "\\{}{}_{}.JPG".format(number, time, b_name)
            if not os.path.exists(before_files["사진"]):
                before_files["사진"] = None
            else:
                after_files["사진"] = after_name + "\\{}{}_{}.JPG".format(number, time, a_name)
        else:
            after_files["사진"] = after_name + "\\{}{}_{}.jpg".format(number, time, a_name)

        # 서류 확인
        for document in ["기본증명서", "주민등록등본", "외국인등록증"]:
            if not os.path.exists(before_files[document]):
                before_files[document] = after_name + "\\{}_{}.JPG".format(b_name, document)
                if not os.path.exists(before_files[document]):
                    before_files[document] = None
                else:
                    after_files[document] = after_name + "\\{}_{}.JPG".format(a_name, document)
            else:
                after_files[document] = after_name + "\\{}_{}.jpg".format(a_name, document)

        for file in before_files:
            if before_files[file] == None:
                continue
            os.rename(before_files[file], after_files[file])

    def changeClass(self, name, b_number, b_time, a_number, a_time):
        before_class = self.base_path + "\\{}\\{}{}\\{}".format(b_number, b_number, b_time, name)
        after_class = self.base_path + "\\{}\\{}{}".format(a_number, a_number, a_time)
        shutil.move(before_class, after_class)

        # 사진 존재 여부 확인
        before_picture = after_class + "\\{}\\{}{}_{}.jpg".format(name, b_number, b_time, name)
        if not os.path.exists(before_picture):
            before_picture = after_class + "\\{}\\{}{}_{}.JPG".format(name, b_number, b_time, name)
            if not os.path.exists(before_picture):
                return
            else:
                after_picture = after_class + "\\{}\\{}{}_{}.JPG".format(name, a_number, a_time, name)
        else:
            after_picture = after_class + "\\{}\\{}{}_{}.jpg".format(name, a_number, a_time, name)

        os.rename(before_picture, after_picture)

    def dataUpdate(self):
        if self.target_table == "user":
            if self.text_id_user.text().strip() == "" or self.text_name_user.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID, 이름값을 입력해야 합니다!")
                return
            
            user_list = []
            user_list.append(self.text_id_user.text().strip())
            user_list.append(self.text_name_user.text().strip())
            user_list.append(self.text_RRN.text().strip())
            user_list.append(self.text_phone.text().strip())
            user_list.append(self.text_licen_user.text().strip())
            user_list.append(self.text_adr.text().strip())
            user_list.append(self.text_origin_adr.text().strip())
            user_list.append(self.text_clsN_user.text().strip())
            user_list.append(self.text_clsT_user.text().strip())
            # 총 이수, 이론, 실기, 실습 시간으로 NULL값 추가
            try:
                total_Hour = str(int(self.text_theory_time.text().strip()) + int(self.text_practice_time.text().strip()) + int(self.text_training_time.text().strip()))
            except:
                total_Hour = NULL
            user_list.append(total_Hour)
            user_list.append(self.text_theory_time.text().strip())
            user_list.append(self.text_practice_time.text().strip())
            user_list.append(self.text_training_time.text().strip())
            user_list.append(self.text_temp.text().strip())
            user_list.append(self.text_exam.text().strip())

            query_list = ["id", "name", "RRN", "phoneNumber", "license", "address", "originAddress", "classNumber", "classTime", \
                "totalCreditHour", "theoryCreditHour", "practicalCreditHour", "trainingCreditHour", "temporaryClassNumber", "exam"]

            where = "id = '{}' and name = '{}'".format(self.key_dict["ID"], self.key_dict["name"])
            query = ""
            for i in range(len(user_list)):
                query += query_list[i] + "="

                # 값이 없거나 NULL값인 경우는 그냥('없이) query문에 들어가고, 아닌 경우는 '를 붙혀서 query문에 넣는다!
                if user_list[i] == "" or user_list[i] == NULL:
                    user_list[i] = NULL
                    query += user_list[i]

                else:
                    query += "'" + user_list[i] + "'"

                if i != len(user_list) - 1:
                    query += ", "
            
            ask = "ID: {}\t이름: {}\t주민등록번호: {}\n전화번호: {}\t자격증: {}\n주소: {}\n본적주소: {}\n기수: {}\t반: {}\t 대체실습: {}\n총 이수시간: {}\t이론이수: {}\t실습이수: {}\t실기이수: {}\n시험회차: {}회"\
                .format(user_list[0], user_list[1], user_list[2], user_list[3], user_list[4], user_list[5], user_list[6], user_list[7], user_list[8], user_list[13], user_list[9], user_list[10], user_list[11], user_list[12], user_list[14])
            ask += "\n해당 정보로 업데이트합니다."

            table = "수강생"
            classification = "{}{} {}".format(user_list[1], user_list[7], user_list[8])
                
        elif self.target_table == "lecture":
            if self.text_clsN_lecture.text().strip() == "" or self.text_clsT_lecture.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수, 반을 입력해야 합니다!")
                return

            lect_list = []
            lect_list.append(self.text_clsN_lecture.text().strip())
            lect_list.append(self.text_clsT_lecture.text().strip())
            lect_list.append(self.text_startD_lecture.text().strip())
            lect_list.append(self.text_endD_lecture.text().strip())

            query_list = ["classNumber", "classTime", "startDate", "endDate"]

            where = "classNumber = '{}' and classTime = '{}'".format(self.key_dict["기수"], self.key_dict["반"])

            query = ""
            for i in range(len(lect_list)):
                query += query_list[i] + "="

                if lect_list[i] == "" or lect_list[i] == NULL:
                    lect_list[i] = NULL
                    query += lect_list[i]

                else:
                    query += "'" + lect_list[i] + "'"

                if i != len(lect_list) - 1:
                    query += ", "

            ask = "기수: {}\t반: {}\n시작일: {}\n종료일: {}\n".format(lect_list[0], lect_list[1], lect_list[2], lect_list[3])
            ask += "\n해당 정보로 업데이트합니다."

            table = "기수"
            classification = "{}{}".format(lect_list[0], lect_list[1])

        elif self.target_table == "teacher":
            if self.text_id_teacher.text().strip() == "" or self.text_name_teacher.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID, 이름값을 입력해야 합니다!")
                return
            
            teach_list = []
            teach_list.append(self.text_id_teacher.text().strip())
            teach_list.append(self.text_category.text().strip())
            teach_list.append(self.text_name_teacher.text().strip())
            teach_list.append(self.text_DOB.text().strip())
            teach_list.append(self.text_licen_teacher.text().strip())
            teach_list.append(self.text_min_career.text().strip())
            teach_list.append(self.text_ACK.text().strip())

            query_list = ["id", "category", "name", "dateOfBirth", "license", "minCareer", "ACKDate"]

            where = "id = '{}' and name = '{}'".format(self.key_dict["ID"], self.key_dict["name"])

            query = ""
            for i in range(len(teach_list)):
                query += query_list[i] + "="

                if teach_list[i] == "" or teach_list[i] == NULL:
                    teach_list[i] = NULL
                    query += teach_list[i]

                else:
                    query += "'" + teach_list[i] + "'"

                if i != len(teach_list) - 1:
                    query += ", "

            ask = "ID: {}\t이름: {}\t자격증: {}\n생년월일: {}\t구분: {}\n최소경력: {}\n도 승인일자: {}\n"\
                .format(teach_list[0], teach_list[1], teach_list[2], teach_list[3], teach_list[4], teach_list[5], teach_list[6])
            ask += "\n해당 정보로 업데이트합니다."

            table = "강사"
            classification = "{} / {}".format(teach_list[1], teach_list[3])

        elif self.target_table == "facility":
            if self.text_id_facility.text().strip() == "" or self.text_name_facility.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID, 이름값을 입력해야 합니다!")
                return
            
            facility_list = []
            facility_list.append(self.text_id_facility.text().strip())
            facility_list.append(self.text_name_facility.text().strip())
            facility_list.append(self.text_catg_facility.text().strip())
            facility_list.append(self.text_start_contract.text().strip())
            facility_list.append(self.text_end_contract.text().strip())
            facility_list.append(self.text_person_facility.text().strip())

            query_list = ["id", "name", "category", "contractTermStart", "contractTermEnd", "personnel"]

            where = "id = '{}' and name = '{}'".format(self.key_dict["ID"], self.key_dict["name"])

            query = ""
            for i in range(len(facility_list)):
                query += query_list[i] + "="

                if facility_list[i] == "" or facility_list[i] == NULL:
                    facility_list[i] = NULL
                    query += facility_list[i]

                else:
                    query += "'" + facility_list[i] + "'"

                if i != len(facility_list) - 1:
                    query += ", "

            ask = "ID: {}\t기관명: {}\t구분: {}\t1일 실습인원: {}\n계약 시작일: {}\t\t계약 종료일: {}\n"\
                .format(facility_list[0], facility_list[1], facility_list[2], facility_list[5], facility_list[3], facility_list[4])
            ask += "\n해당 정보로 업데이트합니다."

            table = "기관"
            classification = "{} / {}".format(facility_list[0], facility_list[1])

        elif self.target_table == "temptraining":
            if self.text_clsN_temp_training.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수를 입력해야 합니다!")
                return

            temp_list = []
            temp_list.append(self.text_clsN_temp_training.text().strip())
            temp_list.append(self.text_startD_temp_training.text().strip())
            temp_list.append(self.text_endD_temp_training.text().strip())
            temp_list.append(self.text_awardD.text().strip())

            query_list = ["classNumber", "startDate", "endDate", "awardDate"]

            where = "classNumber = '{}'".format(self.key_dict["기수"])

            query = ""
            for i in range(len(temp_list)):
                query += query_list[i] + "="

                if temp_list[i] == "" or temp_list[i] == NULL:
                    temp_list[i] = NULL
                    query += temp_list[i]

                else:
                    query += "'" + temp_list[i] + "'"

                if i != len(temp_list) - 1:
                    query += ", "

            ask = "기수: {}\n시작일: {}\n종료일: {}\n수여일: {}\n".format(temp_list[0], temp_list[1], temp_list[2], temp_list[3])
            ask += "\n해당 정보로 업데이트합니다."

            table = "대체실습"
            classification = "{}".format(temp_list[0])

        elif self.target_table == "temptrainingteacher":
            if self.text_clsN_temp_training_teacher.text().strip() == "" or self.text_teacher.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수, 담당강사를 입력해야 합니다!")
                return

            temp_training_teacher_list = []
            temp_training_teacher_list.append(self.text_clsN_temp_training_teacher.text().strip())
            temp_training_teacher_list.append(self.text_teacher.text().strip())

            query_list = ["classNumber", "teacherName"]

            where = "classNumber = '{}' and teacherName = '{}'".format(self.key_dict["기수"], self.key_dict["강사"])

            query = ""
            for i in range(len(temp_training_teacher_list)):
                query += query_list[i] + "="
                if temp_training_teacher_list[i] == "" or temp_training_teacher_list[i] == NULL:
                    temp_training_teacher_list[i] = NULL
                    query += temp_training_teacher_list[i]

                else:
                    query += "'" + temp_training_teacher_list[i] + "'"

                if i != len(temp_training_teacher_list) - 1:
                    query += ", "

            ask = "기수: {}\n강사: {}\n".format(temp_training_teacher_list[0], temp_training_teacher_list[1])
            ask += "\n해당 정보로 업데이트합니다."

            table = "대체실습 강사"
            classification = "{} / {}".format(temp_training_teacher_list[0], temp_training_teacher_list[1])

        elif self.target_table == "exam":
            if self.text_exam_round.text().strip() == "":
                QMessageBox.warning(self, "오류", "시험회차를 입력해야 합니다!")
                return

            exam_list = []
            exam_list.append(self.text_exam_round.text().strip())
            exam_list.append(self.text_exam_startAcceptance.text().strip())
            exam_list.append(self.text_exam_endAcceptance.text().strip())
            exam_list.append(self.text_exam_announceDate.text().strip())
            exam_list.append(self.text_exam_examDate.text().strip())
            exam_list.append(self.text_exam_passDate.text().strip())
            exam_list.append(self.text_exam_submitDate.text().strip())

            query_list = ["round", "startAcceptance", "endAcceptance", "announceDate", "examDate", "passDate", "submitDate"]
            where = "round = {}".format(self.key_dict["시험회차"])

            query = ""
            for i in range(len(exam_list)):
                query += query_list[i] + "="
                if exam_list[i] == "" or exam_list[i] == NULL:
                    exam_list[i] = NULL
                    query += exam_list[i]

                else:
                    query += "'" + exam_list[i] + "'"

                if i != len(exam_list) - 1:
                    query += ", "

            ask = "시험회차: {}\n응시원서 접수 시작일: {}\t응시원서 접수 종료일: {}\n응시표 출력: {}\t시험일: {}\n합격자 발표 예정일: {}\t서류 준비 날짜: {}\n\n해당 정보로 업데이트합니다.".format(exam_list[0], exam_list[1], exam_list[2], exam_list[3], exam_list[4], exam_list[5], exam_list[6])

            table = "시험"
            classification = "{}회".format(exam_list[0])


        ans = QMessageBox.question(self, "데이터 수정 확인", ask, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            db.logger.info("$UI UPDATE Request [TABLE|{}][{}] {} 수정 요청".format(self.target_table, table, classification))
            db.main.dbPrograms.UPDATE(self.target_table, query, where)
            QMessageBox.about(self, "완료", "데이터를 성공적으로 수정했습니다.")
            db.main.showTable(Refresh=True)
            db.main.textInfo.clear()

            if self.target_table == "user":
                name = self.text_name_user.text().strip()
                number = self.text_clsN_user.text().strip()
                time = self.text_clsT_user.text().strip()

                if name != '':
                    if name != self.key_dict["name"]:
                        self.changeName(self.key_dict["기수"], self.key_dict["반"], self.key_dict["name"], name)
                        self.key_dict["name"] = name

                if number != '' and time != '':
                    if number != self.key_dict["기수"] or time != self.key_dict["반"]:
                        self.changeClass(self.key_dict["name"], self.key_dict["기수"], self.key_dict["반"], number, time)
                        self.key_dict["기수"] = number
                        self.key_dict["반"] = time

            self.close()
        else:
            pass

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.dataUpdate()

    def showEvent(self, QShowEvent):
        self.key_dict = {}
        cnt_row = 0
        cnt_col = 0
        # 1. 기존에 있던 label과 Line Edit 삭제
        for i in reversed(range(self.grid.count())):
                self.grid.itemAt(i).widget().deleteLater()
                # self.grid.itemAt(i).widget().hide()

        # 2. 공통으로 들어갈 insert 버튼과 close 버튼 생성
        self.btn_update = QPushButton("Update", self)
        self.btn_update.clicked.connect(self.dataUpdate)
        self.btn_cancel = QPushButton("Close", self)
        self.btn_cancel.clicked.connect(self.close)
        
        # 3. 현재 테이블에 맞는 label 및 Line Edit 생성 및 추가
        if db.main.current_table == "user":
            self.target_table = "user"
            self.setWindowTitle("데이터 수정 - 수강생")
            self.setFixedSize(600, 400)
            cnt_row = 7
            cnt_col = 6

            self.label_id_user = QLabel("ID", self)
            self.label_id_user.setFixedWidth(90)
            self.label_id_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_user, 0, 0)
            self.text_id_user = QLineEdit()
            self.grid.addWidget(self.text_id_user, 0, 1)
            self.label_name_user = QLabel("이름", self)
            self.label_name_user.setFixedWidth(90)
            self.label_name_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_user, 0, 2)
            self.text_name_user = QLineEdit()
            self.grid.addWidget(self.text_name_user, 0, 3)
            self.label_licen_user = QLabel("자격증", self)
            self.label_licen_user.setFixedWidth(90)
            self.label_licen_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_licen_user, 0, 4)
            self.text_licen_user = QLineEdit()
            self.grid.addWidget(self.text_licen_user, 0, 5)

            self.label_clsN_user = QLabel("기수", self)
            self.label_clsN_user.setFixedWidth(90)
            self.label_clsN_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_user, 1, 0)
            self.text_clsN_user = QLineEdit()
            self.grid.addWidget(self.text_clsN_user, 1, 1)
            self.label_clsT_user = QLabel("반", self)
            self.label_clsT_user.setFixedWidth(90)
            self.label_clsT_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsT_user, 1, 2)
            self.text_clsT_user = QLineEdit()
            self.grid.addWidget(self.text_clsT_user, 1, 3)
            self.label_temp = QLabel("대체실습", self)
            self.label_temp.setFixedWidth(90)
            self.label_temp.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_temp, 1, 4)
            self.text_temp = QLineEdit()
            self.grid.addWidget(self.text_temp, 1, 5)

            self.label_RRN = QLabel("주민등록번호", self)
            self.label_RRN.setFixedWidth(90)
            self.label_RRN.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_RRN, 2, 0)
            self.text_RRN = QLineEdit()
            self.grid.addWidget(self.text_RRN, 2, 1, 1, 2)
            self.label_phone = QLabel("전화번호", self)
            self.label_phone.setFixedWidth(90)
            self.label_phone.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_phone, 2, 3)
            self.text_phone = QLineEdit()
            self.grid.addWidget(self.text_phone, 2, 4, 1, 2)
            
            
            self.label_adr = QLabel("도로명주소", self)
            self.label_adr.setFixedWidth(90)
            self.label_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_adr, 3, 0)
            self.text_adr = QLineEdit()
            self.grid.addWidget(self.text_adr, 3, 1, 1, 5)
            self.label_origin_adr = QLabel("본적주소", self)
            self.label_origin_adr.setFixedWidth(90)
            self.label_origin_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_origin_adr, 4, 0)
            self.text_origin_adr = QLineEdit()
            self.grid.addWidget(self.text_origin_adr, 4, 1, 1, 5)

            self.label_total_time = QLabel("총 이수시간은 이론 + 실기 + 실습 이수시간으로 입력됩니다.")
            self.grid.addWidget(self.label_total_time, 5, 0, 1, 6)

            self.label_theory_time = QLabel("이론이수")
            self.label_theory_time.setFixedWidth(90)
            self.label_theory_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_theory_time, 6, 0)
            self.text_theory_time = QLineEdit()
            self.grid.addWidget(self.text_theory_time, 6, 1)

            self.label_practice_time = QLabel("실기이수")
            self.label_practice_time.setFixedWidth(90)
            self.label_practice_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_practice_time, 6, 2)
            self.text_practice_time = QLineEdit()
            self.grid.addWidget(self.text_practice_time, 6, 3)

            self.label_training_time = QLabel("실습이수")
            self.label_training_time.setFixedWidth(90)
            self.label_training_time.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_training_time, 6, 4)
            self.text_training_time = QLineEdit()
            self.grid.addWidget(self.text_training_time, 6, 5)

            self.label_exam = QLabel("시험회차")
            self.label_exam.setFixedWidth(90)
            self.label_exam.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam, 7, 0)
            self.text_exam = QLineEdit()
            self.grid.addWidget(self.text_exam, 7, 1)

            input_user = []
            for i in range(15):
                input_user.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())

                if input_user[i] == "NULL":
                    input_user[i] = ""

            self.text_id_user.setText(str(input_user[0]))
            self.text_name_user.setText(str(input_user[1]))
            self.text_licen_user.setText(str(input_user[4]))
            self.text_clsN_user.setText(str(input_user[7]))
            self.text_clsT_user.setText(str(input_user[8]))
            self.text_temp.setText(str(input_user[13]))
            self.text_RRN.setText(str(input_user[2]))
            self.text_phone.setText(str(input_user[3]))
            self.text_adr.setText(str(input_user[5]))
            self.text_origin_adr.setText(str(input_user[6]))
            self.text_theory_time.setText(str(input_user[10]))
            self.text_practice_time.setText(str(input_user[11]))
            self.text_training_time.setText(str(input_user[12]))
            self.text_exam.setText(str(input_user[14]))

            self.key_dict["ID"] = str(input_user[0])
            self.key_dict["name"] = str(input_user[1])
            self.key_dict["기수"] = str(input_user[7])
            self.key_dict["반"] = str(input_user[8])
            self.key_dict["자격증"] = str(input_user[4])

        elif db.main.current_table == "lecture":
            self.target_table = "lecture"
            self.setWindowTitle("데이터 수정 - 기수")
            self.setFixedSize(460, 200)
            cnt_row = 2
            cnt_col = 4
            self.resize(300, 200)
            self.label_clsN_lecture = QLabel("기수", self)
            self.label_clsN_lecture.setFixedWidth(90)
            self.label_clsN_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_lecture, 0, 0)
            self.text_clsN_lecture = QLineEdit()
            self.text_clsN_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_clsN_lecture, 0, 1)
            self.label_clsT_lecture = QLabel("반", self)
            self.label_clsT_lecture.setFixedWidth(90)
            self.label_clsT_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsT_lecture, 0, 2)
            self.text_clsT_lecture = QLineEdit()
            self.text_clsT_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_clsT_lecture, 0, 3)
            self.label_startD_lecture = QLabel("시작일", self)
            self.label_startD_lecture.setFixedWidth(90)
            self.label_startD_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_startD_lecture, 1, 0)
            self.text_startD_lecture = QLineEdit()
            self.text_startD_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_startD_lecture, 1, 1)
            self.label_endD_lecture = QLabel("종료일", self)
            self.label_endD_lecture.setFixedWidth(90)
            self.label_endD_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_endD_lecture, 1, 2)
            self.text_endD_lecture = QLineEdit()
            self.text_endD_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_endD_lecture, 1, 3)

            input_lecture = []
            for i in range(4):
                input_lecture.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())

                if input_lecture[i] == "NULL":
                    input_lecture[i] = ""

            self.text_clsN_lecture.setText(str(input_lecture[0]))
            self.text_clsT_lecture.setText(str(input_lecture[1]))
            self.text_startD_lecture.setText(str(input_lecture[2]))
            self.text_endD_lecture.setText(str(input_lecture[3]))

            self.key_dict["기수"] = str(input_lecture[0])
            self.key_dict["반"] = str(input_lecture[1])

        elif db.main.current_table == "teacher":
            self.target_table = "teacher"
            self.setWindowTitle("데이터 수정 - 강사")
            self.setFixedSize(530, 200)
            cnt_row = 3
            cnt_col = 6
            self.resize(400, 200)
            self.label_id_teacher = QLabel("ID", self)
            self.label_id_teacher.setFixedWidth(90)
            self.label_id_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_teacher, 0, 0)
            self.text_id_teacher = QLineEdit()
            self.grid.addWidget(self.text_id_teacher, 0, 1)
            self.label_name_teacher = QLabel("이름", self)
            self.label_name_teacher.setFixedWidth(90)
            self.label_name_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_teacher, 0, 2)
            self.text_name_teacher = QLineEdit()
            self.grid.addWidget(self.text_name_teacher, 0, 3)
            self.label_licen_teacher = QLabel("자격증", self)
            self.label_licen_teacher.setFixedWidth(90)
            self.label_licen_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_licen_teacher, 0, 4)
            self.text_licen_teacher = QLineEdit()
            self.grid.addWidget(self.text_licen_teacher, 0, 5)
            self.label_DOB = QLabel("생년월일", self)
            self.label_DOB.setFixedWidth(90)
            self.label_DOB.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_DOB, 1, 0)
            self.text_DOB = QLineEdit()
            self.grid.addWidget(self.text_DOB, 1, 1, 1, 2)
            self.label_category = QLabel("전임/외래", self)
            self.label_category.setFixedWidth(90)
            self.label_category.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_category, 1, 3)
            self.text_category = QLineEdit()
            self.grid.addWidget(self.text_category, 1, 4, 1, 2)
            self.label_min_career = QLabel("최소경력", self)
            self.label_min_career.setFixedWidth(90)
            self.label_min_career.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_min_career, 2, 0)
            self.text_min_career = QLineEdit()
            self.grid.addWidget(self.text_min_career, 2, 1, 1, 2)
            self.label_ACK = QLabel("도 승인일자")
            self.label_ACK.setFixedWidth(90)
            self.label_ACK.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_ACK, 2, 3)
            self.text_ACK = QLineEdit()
            self.grid.addWidget(self.text_ACK, 2, 4, 1, 2)

            input_teacher = []
            for i in range(7):
                input_teacher.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())

                if input_teacher[i] == "NULL":
                    input_teacher[i] = ""

            self.text_id_teacher.setText(str(input_teacher[0]))
            self.text_name_teacher.setText(str(input_teacher[2]))
            self.text_licen_teacher.setText(str(input_teacher[4]))
            self.text_DOB.setText(str(input_teacher[3]))
            self.text_category.setText(str(input_teacher[1]))
            self.text_min_career.setText(str(input_teacher[5]))
            self.text_ACK.setText(str(input_teacher[6]))

            self.key_dict["ID"] = str(input_teacher[0])
            self.key_dict["name"] = str(input_teacher[2])

        elif db.main.current_table == "facility":
            self.target_table = "facility"
            self.setWindowTitle("데이터 수정 - 기관")
            self.setFixedSize(530, 200)
            cnt_row = 3
            cnt_col = 6
            self.resize(400, 200)
            self.label_id_facility = QLabel("ID", self)
            self.label_id_facility.setFixedWidth(90)
            self.label_id_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_facility, 0, 0)
            self.text_id_facility = QLineEdit()
            next_id = str(int(db.main.dbPrograms.SELECT("id", "facility", orderBy="id desc limit 1")[0][0]) + 1)
            self.text_id_facility.setText(next_id)
            self.grid.addWidget(self.text_id_facility, 0, 1, 1, 2)
            self.label_catg_facility = QLabel("구분", self)
            self.label_catg_facility.setFixedWidth(90)
            self.label_catg_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_catg_facility, 0, 3)
            self.text_catg_facility = QLineEdit()
            self.grid.addWidget(self.text_catg_facility, 0, 4, 1, 2)
            self.label_name_facility = QLabel("기관명", self)
            self.label_name_facility.setFixedWidth(90)
            self.label_name_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_facility, 1, 0)
            self.text_name_facility = QLineEdit()
            self.grid.addWidget(self.text_name_facility, 1, 1, 1, 2)
            self.label_person_facility = QLabel("1일 실습인원", self)
            self.label_person_facility.setFixedWidth(90)
            self.label_person_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_person_facility, 1, 3)
            self.text_person_facility = QLineEdit()
            self.grid.addWidget(self.text_person_facility, 1, 4, 1, 2)
            self.label_start_contract = QLabel("계약 시작일", self)
            self.label_start_contract.setFixedWidth(90)
            self.label_start_contract.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_start_contract, 2, 0)
            self.text_start_contract = QLineEdit()
            self.grid.addWidget(self.text_start_contract, 2, 1, 1, 2)
            self.label_end_contract = QLabel("계약 종료일")
            self.label_end_contract.setFixedWidth(90)
            self.label_end_contract.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_end_contract, 2, 3)
            self.text_end_contract = QLineEdit()
            self.grid.addWidget(self.text_end_contract, 2, 4, 1, 2)

            input_facility = []
            for i in range(7):
                input_facility.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())

                if input_facility[i] == "NULL":
                    input_facility[i] = ""

            self.text_id_facility.setText(str(input_facility[0]))
            self.text_name_facility.setText(str(input_facility[1]))
            self.text_catg_facility.setText(str(input_facility[2]))
            self.text_start_contract.setText(str(input_facility[3]))
            self.text_end_contract.setText(str(input_facility[4]))
            self.text_person_facility.setText(str(input_facility[5]))

            self.key_dict["ID"] = str(input_facility[0])
            self.key_dict["name"] = str(input_facility[1])

        elif db.main.current_table == "temptraining":
            self.target_table = "temptraining"
            self.setWindowTitle("데이터 수정 - 대체실습")
            self.setFixedSize(460, 200)
            cnt_row = 2
            cnt_col = 4
            self.resize(300, 200)
            self.label_clsN_temp_training = QLabel("기수", self)
            self.label_clsN_temp_training.setFixedWidth(90)
            self.label_clsN_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_temp_training, 0, 0)
            self.text_clsN_temp_training = QLineEdit()
            self.text_clsN_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_clsN_temp_training, 0, 1)
            self.label_startD_temp_training = QLabel("시작일", self)
            self.label_startD_temp_training.setFixedWidth(90)
            self.label_startD_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_startD_temp_training, 0, 2)
            self.text_startD_temp_training = QLineEdit()
            self.text_startD_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_startD_temp_training, 0, 3)
            self.label_endD_temp_training = QLabel("종료일", self)
            self.label_endD_temp_training.setFixedWidth(90)
            self.label_endD_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_endD_temp_training, 1, 0)
            self.text_endD_temp_training = QLineEdit()
            self.text_endD_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_endD_temp_training, 1, 1)
            self.label_awardD = QLabel("수여일", self)
            self.label_awardD.setFixedWidth(90)
            self.label_awardD.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_awardD, 1, 2)
            self.text_awardD = QLineEdit()
            self.text_awardD.setFixedWidth(120)
            self.grid.addWidget(self.text_awardD, 1, 3)

            input_temp_training = []
            for i in range(4):
                input_temp_training.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())

                if input_temp_training[i] == "NULL":
                    input_temp_training[i] = ""

            self.text_clsN_temp_training.setText(str(input_temp_training[0]))
            self.text_startD_temp_training.setText(str(input_temp_training[1]))
            self.text_endD_temp_training.setText(str(input_temp_training[2]))
            self.text_awardD.setText(str(input_temp_training[3]))

            self.key_dict["기수"] = str(input_temp_training[0])

        elif db.main.current_table == "temptrainingteacher":
            self.target_table = "temptrainingteacher"
            self.setWindowTitle("데이터 수정 - 대체실습 강사")
            self.setFixedSize(300, 200)
            cnt_row = 2
            cnt_col = 2
            self.resize(300, 200)
            self.label_clsN_temp_training_teacher = QLabel("기수", self)
            self.label_clsN_temp_training_teacher.setFixedWidth(90)
            self.label_clsN_temp_training_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_temp_training_teacher, 0, 0)
            self.text_clsN_temp_training_teacher = QLineEdit()
            self.text_clsN_temp_training_teacher.setFixedWidth(90)
            self.grid.addWidget(self.text_clsN_temp_training_teacher, 0, 1)
            self.label_teacher = QLabel("강사", self)
            self.label_teacher.setFixedWidth(90)
            self.label_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_teacher, 1, 0)
            self.text_teacher = QLineEdit()
            self.text_teacher.setFixedWidth(90)
            self.grid.addWidget(self.text_teacher, 1, 1)

            input_temp_training_teacher = []
            for i in range(2):
                input_temp_training_teacher.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())
                
                if input_temp_training_teacher[i] == "NULL":
                    input_temp_training_teacher[i] = ""

            self.text_clsN_temp_training_teacher.setText(str(input_temp_training_teacher[0]))
            self.text_teacher.setText(str(input_temp_training_teacher[1]))

            self.key_dict["기수"] = str(input_temp_training_teacher[0])
            self.key_dict["강사"] = str(input_temp_training_teacher[1])

        elif db.main.current_table == "exam":
            self.target_table = "exam"
            self.setWindowTitle("데이터 수정 - 국시원 시험 정보")
            self.setFixedSize(938, 89)
            cnt_row = 2
            cnt_col = 7
            self.label_exam_round = QLabel("시험회차", self)
            self.label_exam_round.setFixedWidth(100)
            self.label_exam_round.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_round, 0, 0)
            self.text_exam_round = QLineEdit()
            self.text_exam_round.setFixedWidth(100)
            self.grid.addWidget(self.text_exam_round, 1, 0)

            self.label_exam_startAcceptance = QLabel("응시원서 접수 시작일")
            self.label_exam_startAcceptance.setFixedWidth(120)
            self.label_exam_startAcceptance.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_startAcceptance, 0, 1)
            self.text_exam_startAcceptance = QLineEdit()
            self.text_exam_startAcceptance.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_startAcceptance, 1, 1)

            self.label_exam_endAcceptance = QLabel("응시원서 접수 종료일")
            self.label_exam_endAcceptance.setFixedWidth(120)
            self.label_exam_endAcceptance.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_endAcceptance, 0, 2)
            self.text_exam_endAcceptance = QLineEdit()
            self.text_exam_endAcceptance.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_endAcceptance, 1, 2)

            self.label_exam_announceDate = QLabel("시험장소 공고일(응시표 출력)")
            self.label_exam_announceDate.setFixedWidth(180)
            self.label_exam_announceDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_announceDate, 0, 3)
            self.text_exam_announceDate = QLineEdit()
            self.text_exam_announceDate.setFixedWidth(180)
            self.grid.addWidget(self.text_exam_announceDate, 1, 3)

            self.label_exam_examDate = QLabel("시험일")
            self.label_exam_examDate.setFixedWidth(120)
            self.label_exam_examDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_examDate, 0, 4)
            self.text_exam_examDate = QLineEdit()
            self.text_exam_examDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_examDate, 1, 4)

            self.label_exam_passDate = QLabel("시험 합격일")
            self.label_exam_passDate.setFixedWidth(120)
            self.label_exam_passDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_passDate, 0, 5)
            self.text_exam_passDate = QLineEdit()
            self.text_exam_passDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_passDate, 1, 5)

            self.label_exam_submitDate = QLabel("서류 준비 일자")
            self.label_exam_submitDate.setFixedWidth(120)
            self.label_exam_submitDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_submitDate, 0, 6)
            self.text_exam_submitDate = QLineEdit()
            self.text_exam_submitDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_submitDate, 1, 6)

            input_exam = []
            for i in range(7):
                input_exam.append(db.main.readDB.index(db.main.table.currentIndex().row(), i).data())
                
                if input_exam[i] == "NULL":
                    input_exam[i] = ""

            self.text_exam_round.setText(str(input_exam[0]))
            self.text_exam_startAcceptance.setText(str(input_exam[1]))
            self.text_exam_endAcceptance.setText(str(input_exam[2]))
            self.text_exam_announceDate.setText(str(input_exam[3]))
            self.text_exam_examDate.setText(str(input_exam[4]))
            self.text_exam_passDate.setText(str(input_exam[5]))
            self.text_exam_submitDate.setText(str(input_exam[6]))

            self.key_dict["시험회차"] = str(input_exam[0])
        
        # 4. 재생성된 버튼 추가
        self.grid.addWidget(self.btn_update, cnt_row, cnt_col - 2)
        self.grid.addWidget(self.btn_cancel, cnt_row, cnt_col - 1)


class INSERT(QWidget):
    # 새 창을 띄우기 위해 서로 global로 연결
    global db

    def __init__(self):
        super().__init__()
        self.initUI()
        self.target_table = ""
        self.base_path = "D:\\남양노아요양보호사교육원\\교육생관리"
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))

    def initUI(self):
        self.setWindowTitle("데이터 삽입")

        # 총 이수, 이론 이수, 실기, 실습 이수시간은 삽입에서 넣지 않고, NULL값을 일단 너놓자! Update에서 구현하기
        # self.labelTotalH = QLabel("총 이수시간", self)
        # grid.addWidget(self.labelTotalH, 9, 0)
        # self.textTotalH = QLineEdit()
        # grid.addWidget(self.textTotalH, 9, 1)

        # btn = QPushButton("수강생 관리", self)
        # btn.resize(btn.sizeHint())
        # btn.move(50, 50)
        # btn.clicked.connect(QCoreApplication.instance().quit)

        self.grid = QGridLayout()
        self.setLayout(self.grid)

    def generateDirectory(self, number, time, name):
        path = self.base_path + "\\{}\\{}{}".format(number, number, time)
        if not os.path.exists(path):
            os.makedirs(path)
        
        path = path + "\\{}".format(name)

        if not os.path.exists(path):
            os.makedirs(path)

    def dataInsert(self):
        if self.target_table == "user":
            if self.text_id_user.text().strip() == "" or self.text_name_user.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID, 이름값을 입력해야 합니다!")
                return

            user_list = []
            user_list.append(self.text_id_user.text().strip())
            user_list.append(self.text_name_user.text().strip())
            user_list.append(self.text_RRN.text().strip())
            user_list.append(self.text_phone.text().strip())
            user_list.append(self.text_licen_user.text().strip())
            user_list.append(self.text_adr.text().strip())
            user_list.append(self.text_origin_adr.text().strip())
            user_list.append(self.text_clsN_user.text().strip())
            user_list.append(self.text_clsT_user.text().strip())
            # 총 이수, 이론, 실기, 실습 시간으로 NULL값 추가
            user_list.append(NULL)
            user_list.append(NULL)
            user_list.append(NULL)
            user_list.append(NULL)
            user_list.append(self.text_temp.text().strip())
            # 시험 회차 정보는 데이터 수정에서 입력
            user_list.append(NULL)

            query = ""
            for i in range(len(user_list)):
                # 값이 없거나 NULL값인 경우는 그냥('없이) query문에 들어가고, 아닌 경우는 '를 붙혀서 query문에 넣는다!
                if user_list[i] == "" or user_list[i] == NULL:
                    user_list[i] = NULL
                    query += user_list[i]

                else:
                    query += "'" + user_list[i] + "'"

                if i != len(user_list) - 1:
                    query += ", "
            
            ask = "ID: {}\t이름: {}\t주민등록번호: {}\n전화번호: {}\t자격증: {}\n주소: {}\n본적주소: {}\n기수: {}\t반: {}\t 대체실습: {}\n총 이수시간: {}\t이론이수: {}\t실습이수: {}\t 실기이수: {}\n"\
                .format(user_list[0], user_list[1], user_list[2], user_list[3], user_list[4], user_list[5], user_list[6], user_list[7], user_list[8], user_list[13], user_list[9], user_list[10], user_list[11], user_list[12])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            more_check = "name = '{}' and classNumber = '{}' and classTime = '{}'".format(user_list[1], user_list[7], user_list[8])
            check = db.main.dbPrograms.SELECT("*", "user", more_check)
            if check:
                QMessageBox.warning(self, "오류", "{} {}반 {}님이 이미 존재합니다!".format(user_list[7], user_list[8], user_list[1]))
                return

            table = "수강생"
            classification = "{}{} {}".format(user_list[1], user_list[7], user_list[8])

                
        elif self.target_table == "lecture":
            if self.text_clsN_lecture.text().strip() == "" or self.text_clsT_lecture.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수, 반을 입력해야 합니다!")
                return

            lect_list = []
            lect_list.append(self.text_clsN_lecture.text().strip())
            lect_list.append(self.text_clsT_lecture.text().strip())
            lect_list.append(self.text_startD_lecture.text().strip())
            lect_list.append(self.text_endD_lecture.text().strip())

            query = ""
            for i in range(len(lect_list)):
                if lect_list[i] == "" or lect_list[i] == NULL:
                    lect_list[i] = NULL
                    query += lect_list[i]

                else:
                    query += "'" + lect_list[i] + "'"

                if i != len(lect_list) - 1:
                    query += ", "

            ask = "기수: {}\t반: {}\n시작일: {}\n종료일: {}\n".format(lect_list[0], lect_list[1], lect_list[2], lect_list[3])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            table = "기수"
            classification = "{}{}".format(lect_list[0], lect_list[1])

        elif self.target_table == "teacher":
            if self.text_id_teacher.text().strip() == "" or self.text_name_teacher.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID, 이름값을 입력해야 합니다!")
                return
            
            teach_list = []
            teach_list.append(self.text_id_teacher.text().strip())
            teach_list.append(self.text_category.text().strip())
            teach_list.append(self.text_name_teacher.text().strip())
            teach_list.append(self.text_DOB.text().strip())
            teach_list.append(self.text_licen_teacher.text().strip())
            teach_list.append(self.text_min_career.text().strip())
            teach_list.append(self.text_ACK.text().strip())

            query = ""
            for i in range(len(teach_list)):
                if teach_list[i] == "" or teach_list[i] == NULL:
                    teach_list[i] = NULL
                    query += teach_list[i]

                else:
                    query += "'" + teach_list[i] + "'"

                if i != len(teach_list) - 1:
                    query += ", "

            ask = "ID: {}\t이름: {}\t자격증: {}\n생년월일: {}\t구분: {}\n최소경력: {}\n도 승인일자: {}\n"\
                .format(teach_list[0], teach_list[1], teach_list[2], teach_list[3], teach_list[4], teach_list[5], teach_list[6])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            table = "강사"
            classification = "{} / {}".format(teach_list[1], teach_list[3])

        elif self.target_table == "facility":
            if self.text_id_facility.text().strip() == "" or self.text_name_facility.text().strip() == "":
                QMessageBox.warning(self, "오류", "ID와 기관명을 입력해야 합니다!")
                return

            facility_list = []
            facility_list.append(self.text_id_facility.text().strip())
            facility_list.append(self.text_name_facility.text().strip())
            facility_list.append(self.text_catg_facility.text().strip())
            facility_list.append(self.text_start_contract.text().strip())
            facility_list.append(self.text_end_contract.text().strip())
            facility_list.append(self.text_person_facility.text().strip())

            query = ""
            for i in range(len(facility_list)):
                if facility_list[i] == "" or facility_list[i] == NULL:
                    facility_list[i] = NULL
                    query += facility_list[i]

                else:
                    query += "'" + facility_list[i] + "'"

                if i != len(facility_list) - 1:
                    query += ", "

            ask = "ID: {}\t기관명: {}\t구분: {}\t1일 실습인원: {}\n계약 시작일: {}\t\t계약 종료일: {}\n"\
                .format(facility_list[0], facility_list[1], facility_list[2], facility_list[5], facility_list[3], facility_list[4])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            table = "기관"
            classification = "{} / {}".format(facility_list[0], facility_list[1])

        elif self.target_table == "temptraining":
            if self.text_clsN_temp_training.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수를 입력해야 합니다!")
                return

            temp_list = []
            temp_list.append(self.text_clsN_temp_training.text().strip())
            temp_list.append(self.text_startD_temp_training.text().strip())
            temp_list.append(self.text_endD_temp_training.text().strip())
            temp_list.append(self.text_awardD.text().strip())

            query = ""
            for i in range(len(temp_list)):
                if temp_list[i] == "" or temp_list[i] == NULL:
                    temp_list[i] = NULL
                    query += temp_list[i]

                else:
                    query += "'" + temp_list[i] + "'"

                if i != len(temp_list) - 1:
                    query += ", "

            ask = "기수: {}\n시작일: {}\n종료일: {}\n수여일: {}\n".format(temp_list[0], temp_list[1], temp_list[2], temp_list[3])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            table = "대체실습"
            classification = "{}".format(temp_list[0])

        elif self.target_table == "temptrainingteacher":
            if self.text_clsN_temp_training_teacher.text().strip() == "" or self.text_teacher.text().strip() == "":
                QMessageBox.warning(self, "오류", "기수, 담당강사를 입력해야 합니다!")
                return

            temp_training_teacher_list = []
            temp_training_teacher_list.append(self.text_clsN_temp_training_teacher.text().strip())
            temp_training_teacher_list.append(self.text_teacher.text().strip())

            query = ""
            for i in range(len(temp_training_teacher_list)):
                if temp_training_teacher_list[i] == "" or temp_training_teacher_list[i] == NULL:
                    temp_training_teacher_list[i] = NULL
                    query += temp_training_teacher_list[i]

                else:
                    query += "'" + temp_training_teacher_list[i] + "'"

                if i != len(temp_training_teacher_list) - 1:
                    query += ", "

            ask = "기수: {}\n강사: {}\n".format(temp_training_teacher_list[0], temp_training_teacher_list[1])
            ask += "\n해당 정보를 데이터베이스에 추가합니다."

            table = "대체실습 강사"
            classification = "{} / {}".format(temp_training_teacher_list[0], temp_training_teacher_list[1])

        elif self.target_table == "exam":
            if self.text_exam_round.text().strip() == "":
                QMessageBox.warning(self, "오류", "시험회차를 입력해야 합니다!")
                return

            exam_list = []
            exam_list.append(self.text_exam_round.text().strip())
            exam_list.append(self.text_exam_startAcceptance.text().strip())
            exam_list.append(self.text_exam_endAcceptance.text().strip())
            exam_list.append(self.text_exam_announceDate.text().strip())
            exam_list.append(self.text_exam_examDate.text().strip())
            exam_list.append(self.text_exam_passDate.text().strip())
            exam_list.append(self.text_exam_submitDate.text().strip())

            query = ""
            for i in range(len(exam_list)):
                if exam_list[i] == "" or exam_list[i] == NULL:
                    exam_list[i] = NULL
                    query += exam_list[i]

                else:
                    query += "'" + exam_list[i] + "'"

                if i != len(exam_list) - 1:
                    query += ", "

            ask = "시험회차: {}\n응시원서 접수 시작일: {}\t응시원서 접수 종료일: {}\n응시표 출력: {}\t시험일: {}\n합격자 발표 예정일: {}\t서류 준비 날짜: {}\n\n해당 정보를 데이터베이스에 추가합니다.".format(exam_list[0], exam_list[1], exam_list[2], exam_list[3], exam_list[4], exam_list[5], exam_list[6])

            table = "시험"
            classification = "{}회".format(exam_list[0])

        ans = QMessageBox.question(self, "데이터 삽입 확인", ask, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            db.logger.info("$UI INSERT Request [TABLE|{}][{}] {} 삭제 요청".format(self.target_table, table, classification))
            db.main.dbPrograms.INSERT(self.target_table, query)
            QMessageBox.about(self, "완료", "데이터를 성공적으로 추가했습니다.")
            if self.target_table == "user":
                name = self.text_name_user.text().strip()
                number = self.text_clsN_user.text().strip()
                time = self.text_clsT_user.text().strip()

                if not (number == "" or time == ""):
                    self.generateDirectory(number, time, name)

            db.main.showTable(Refresh=True)
            db.main.textInfo.clear()
            self.close()
        else:
            pass

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            self.close()
        
        elif e.key() == Qt.Key_Enter or e.key() == Qt.Key_Return:
            self.dataInsert()

    def showEvent(self, QShowEvent):

        cnt_row = 0
        cnt_col = 0
        # 1. 기존에 있던 label과 Line Edit 삭제
        for i in reversed(range(self.grid.count())):
                self.grid.itemAt(i).widget().deleteLater()
                # self.grid.itemAt(i).widget().hide()

        # 2. 공통으로 들어갈 insert 버튼과 close 버튼 생성
        self.btn_insert = QPushButton("Insert", self)
        self.btn_insert.clicked.connect(self.dataInsert)
        self.btn_cancel = QPushButton("Close", self)
        self.btn_cancel.clicked.connect(self.close)
        
        # 3. 현재 테이블에 맞는 label 및 Line Edit 생성 및 추가
        if db.main.current_table == "user":
            self.target_table = "user"
            self.setWindowTitle("데이터 삽입 - 수강생")
            cnt_row = 5
            cnt_col = 6
            self.setFixedSize(600, 400)

            self.label_id_user = QLabel("ID", self)
            self.label_id_user.setFixedWidth(90)
            self.label_id_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_user, 0, 0)
            self.text_id_user = QLineEdit()
            next_id = str(int(db.main.dbPrograms.SELECT("id", "user", orderBy="id desc limit 1")[0][0]) + 1)
            self.text_id_user.setText(next_id)
            self.grid.addWidget(self.text_id_user, 0, 1)
            self.label_name_user = QLabel("이름", self)
            self.label_name_user.setFixedWidth(90)
            self.label_name_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_user, 0, 2)
            self.text_name_user = QLineEdit()
            self.grid.addWidget(self.text_name_user, 0, 3)
            self.label_licen_user = QLabel("자격증", self)
            self.label_licen_user.setFixedWidth(90)
            self.label_licen_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_licen_user, 0, 4)
            self.text_licen_user = QLineEdit()
            self.grid.addWidget(self.text_licen_user, 0, 5)

            self.label_clsN_user = QLabel("기수", self)
            self.label_clsN_user.setFixedWidth(90)
            self.label_clsN_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_user, 1, 0)
            self.text_clsN_user = QLineEdit()
            self.grid.addWidget(self.text_clsN_user, 1, 1)
            self.label_clsT_user = QLabel("반", self)
            self.label_clsT_user.setFixedWidth(90)
            self.label_clsT_user.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsT_user, 1, 2)
            self.text_clsT_user = QLineEdit()
            self.grid.addWidget(self.text_clsT_user, 1, 3)
            self.label_temp = QLabel("대체실습", self)
            self.label_temp.setFixedWidth(90)
            self.label_temp.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_temp, 1, 4)
            self.text_temp = QLineEdit()
            self.grid.addWidget(self.text_temp, 1, 5)

            self.label_RRN = QLabel("주민등록번호", self)
            self.label_RRN.setFixedWidth(90)
            self.label_RRN.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_RRN, 2, 0)
            self.text_RRN = QLineEdit()
            self.grid.addWidget(self.text_RRN, 2, 1, 1, 2)
            self.label_phone = QLabel("전화번호", self)
            self.label_phone.setFixedWidth(90)
            self.label_phone.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_phone, 2, 3)
            self.text_phone = QLineEdit()
            self.grid.addWidget(self.text_phone, 2, 4, 1, 2)
            
            
            self.label_adr = QLabel("도로명주소", self)
            self.label_adr.setFixedWidth(90)
            self.label_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_adr, 3, 0)
            self.text_adr = QLineEdit()
            self.grid.addWidget(self.text_adr, 3, 1, 1, 5)
            self.label_origin_adr = QLabel("본적주소", self)
            self.label_origin_adr.setFixedWidth(90)
            self.label_origin_adr.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_origin_adr, 4, 0)
            self.text_origin_adr = QLineEdit()
            self.grid.addWidget(self.text_origin_adr, 4, 1, 1, 5)

            self.text_id_user.setFocus()

        elif db.main.current_table == "lecture":
            self.target_table = "lecture"
            self.setWindowTitle("데이터 삽입 - 기수")
            self.setFixedSize(460, 200)
            cnt_row = 2
            cnt_col = 4
            self.resize(300, 200)
            self.label_clsN_lecture = QLabel("기수", self)
            self.label_clsN_lecture.setFixedWidth(90)
            self.label_clsN_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_lecture, 0, 0)
            self.text_clsN_lecture = QLineEdit()
            self.text_clsN_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_clsN_lecture, 0, 1)
            self.label_clsT_lecture = QLabel("반", self)
            self.label_clsT_lecture.setFixedWidth(90)
            self.label_clsT_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsT_lecture, 0, 2)
            self.text_clsT_lecture = QLineEdit()
            self.text_clsT_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_clsT_lecture, 0, 3)
            self.label_startD_lecture = QLabel("시작일", self)
            self.label_startD_lecture.setFixedWidth(90)
            self.label_startD_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_startD_lecture, 1, 0)
            self.text_startD_lecture = QLineEdit()
            self.text_startD_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_startD_lecture, 1, 1)
            self.label_endD_lecture = QLabel("종료일", self)
            self.label_endD_lecture.setFixedWidth(90)
            self.label_endD_lecture.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_endD_lecture, 1, 2)
            self.text_endD_lecture = QLineEdit()
            self.text_endD_lecture.setFixedWidth(120)
            self.grid.addWidget(self.text_endD_lecture, 1, 3)

        elif db.main.current_table == "teacher":
            self.target_table = "teacher"
            self.setWindowTitle("데이터 삽입 - 강사")
            self.setFixedSize(530, 200)
            cnt_row = 3
            cnt_col = 6
            self.resize(400, 200)
            self.label_id_teacher = QLabel("ID", self)
            self.label_id_teacher.setFixedWidth(90)
            self.label_id_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_teacher, 0, 0)
            self.text_id_teacher = QLineEdit()
            next_id = str(int(db.main.dbPrograms.SELECT("id", "teacher", orderBy="id desc limit 1")[0][0]) + 1)
            self.text_id_teacher.setText(next_id)
            self.grid.addWidget(self.text_id_teacher, 0, 1)
            self.label_name_teacher = QLabel("이름", self)
            self.label_name_teacher.setFixedWidth(90)
            self.label_name_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_teacher, 0, 2)
            self.text_name_teacher = QLineEdit()
            self.grid.addWidget(self.text_name_teacher, 0, 3)
            self.label_licen_teacher = QLabel("자격증", self)
            self.label_licen_teacher.setFixedWidth(90)
            self.label_licen_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_licen_teacher, 0, 4)
            self.text_licen_teacher = QLineEdit()
            self.grid.addWidget(self.text_licen_teacher, 0, 5)
            self.label_DOB = QLabel("생년월일", self)
            self.label_DOB.setFixedWidth(90)
            self.label_DOB.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_DOB, 1, 0)
            self.text_DOB = QLineEdit()
            self.grid.addWidget(self.text_DOB, 1, 1, 1, 2)
            self.label_category = QLabel("전임/외래", self)
            self.label_category.setFixedWidth(90)
            self.label_category.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_category, 1, 3)
            self.text_category = QLineEdit()
            self.grid.addWidget(self.text_category, 1, 4, 1, 2)
            self.label_min_career = QLabel("최소경력", self)
            self.label_min_career.setFixedWidth(90)
            self.label_min_career.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_min_career, 2, 0)
            self.text_min_career = QLineEdit()
            self.grid.addWidget(self.text_min_career, 2, 1, 1, 2)
            self.label_ACK = QLabel("도 승인일자")
            self.label_ACK.setFixedWidth(90)
            self.label_ACK.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_ACK, 2, 3)
            self.text_ACK = QLineEdit()
            self.grid.addWidget(self.text_ACK, 2, 4, 1, 2)

        elif db.main.current_table == "facility":
            self.target_table = "facility"
            self.setWindowTitle("데이터 삽입 - 기관")
            self.setFixedSize(530, 200)
            cnt_row = 3
            cnt_col = 6
            self.resize(400, 200)
            self.label_id_facility = QLabel("ID", self)
            self.label_id_facility.setFixedWidth(90)
            self.label_id_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_id_facility, 0, 0)
            self.text_id_facility = QLineEdit()
            next_id = str(int(db.main.dbPrograms.SELECT("id", "facility", orderBy="id desc limit 1")[0][0]) + 1)
            self.text_id_facility.setText(next_id)
            self.grid.addWidget(self.text_id_facility, 0, 1, 1, 2)
            self.label_catg_facility = QLabel("구분", self)
            self.label_catg_facility.setFixedWidth(90)
            self.label_catg_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_catg_facility, 0, 3)
            self.text_catg_facility = QLineEdit()
            self.grid.addWidget(self.text_catg_facility, 0, 4, 1, 2)
            self.label_name_facility = QLabel("기관명", self)
            self.label_name_facility.setFixedWidth(90)
            self.label_name_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_name_facility, 1, 0)
            self.text_name_facility = QLineEdit()
            self.grid.addWidget(self.text_name_facility, 1, 1, 1, 2)
            self.label_person_facility = QLabel("1일 실습인원", self)
            self.label_person_facility.setFixedWidth(90)
            self.label_person_facility.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_person_facility, 1, 3)
            self.text_person_facility = QLineEdit()
            self.grid.addWidget(self.text_person_facility, 1, 4, 1, 2)
            self.label_start_contract = QLabel("계약 시작일", self)
            self.label_start_contract.setFixedWidth(90)
            self.label_start_contract.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_start_contract, 2, 0)
            self.text_start_contract = QLineEdit()
            self.grid.addWidget(self.text_start_contract, 2, 1, 1, 2)
            self.label_end_contract = QLabel("계약 종료일")
            self.label_end_contract.setFixedWidth(90)
            self.label_end_contract.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_end_contract, 2, 3)
            self.text_end_contract = QLineEdit()
            self.grid.addWidget(self.text_end_contract, 2, 4, 1, 2)

        elif db.main.current_table == "temptraining":
            self.target_table = "temptraining"
            self.setWindowTitle("데이터 삽입 - 대체실습")
            self.setFixedSize(460, 200)
            cnt_row = 2
            cnt_col = 4
            self.resize(300, 200)
            self.label_clsN_temp_training = QLabel("기수", self)
            self.label_clsN_temp_training.setFixedWidth(90)
            self.label_clsN_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_temp_training, 0, 0)
            self.text_clsN_temp_training = QLineEdit()
            self.text_clsN_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_clsN_temp_training, 0, 1)
            self.label_startD_temp_training = QLabel("시작일", self)
            self.label_startD_temp_training.setFixedWidth(90)
            self.label_startD_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_startD_temp_training, 0, 2)
            self.text_startD_temp_training = QLineEdit()
            self.text_startD_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_startD_temp_training, 0, 3)
            self.label_endD_temp_training = QLabel("종료일", self)
            self.label_endD_temp_training.setFixedWidth(90)
            self.label_endD_temp_training.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_endD_temp_training, 1, 0)
            self.text_endD_temp_training = QLineEdit()
            self.text_endD_temp_training.setFixedWidth(120)
            self.grid.addWidget(self.text_endD_temp_training, 1, 1)
            self.label_awardD = QLabel("수여일", self)
            self.label_awardD.setFixedWidth(90)
            self.label_awardD.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_awardD, 1, 2)
            self.text_awardD = QLineEdit()
            self.text_awardD.setFixedWidth(120)
            self.grid.addWidget(self.text_awardD, 1, 3)
            
        elif db.main.current_table == "temptrainingteacher":
            self.target_table = "temptrainingteacher"
            self.setWindowTitle("데이터 삽입 - 대체실습 강사")
            self.setFixedSize(300, 200)
            cnt_row = 2
            cnt_col = 2
            self.resize(300, 200)
            self.label_clsN_temp_training_teacher = QLabel("기수", self)
            self.label_clsN_temp_training_teacher.setFixedWidth(90)
            self.label_clsN_temp_training_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_clsN_temp_training_teacher, 0, 0)
            self.text_clsN_temp_training_teacher = QLineEdit()
            self.text_clsN_temp_training_teacher.setFixedWidth(90)
            self.grid.addWidget(self.text_clsN_temp_training_teacher, 0, 1)
            self.label_teacher = QLabel("강사", self)
            self.label_teacher.setFixedWidth(90)
            self.label_teacher.setAlignment(Qt.AlignRight | Qt.AlignVCenter)
            self.grid.addWidget(self.label_teacher, 1, 0)
            self.text_teacher = QLineEdit()
            self.text_teacher.setFixedWidth(90)
            self.grid.addWidget(self.text_teacher, 1, 1)

        elif db.main.current_table == "exam":
            self.target_table = "exam"
            self.setWindowTitle("데이터 삽입 - 국시원 시험 정보")
            self.setFixedSize(938, 89)
            cnt_row = 2
            cnt_col = 7
            self.label_exam_round = QLabel("시험회차", self)
            self.label_exam_round.setFixedWidth(100)
            self.label_exam_round.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_round, 0, 0)
            self.text_exam_round = QLineEdit()
            self.text_exam_round.setFixedWidth(100)
            self.grid.addWidget(self.text_exam_round, 1, 0)

            self.label_exam_startAcceptance = QLabel("응시원서 접수 시작일")
            self.label_exam_startAcceptance.setFixedWidth(120)
            self.label_exam_startAcceptance.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_startAcceptance, 0, 1)
            self.text_exam_startAcceptance = QLineEdit()
            self.text_exam_startAcceptance.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_startAcceptance, 1, 1)

            self.label_exam_endAcceptance = QLabel("응시원서 접수 종료일")
            self.label_exam_endAcceptance.setFixedWidth(120)
            self.label_exam_endAcceptance.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_endAcceptance, 0, 2)
            self.text_exam_endAcceptance = QLineEdit()
            self.text_exam_endAcceptance.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_endAcceptance, 1, 2)

            self.label_exam_announceDate = QLabel("시험장소 공고일(응시표 출력)")
            self.label_exam_announceDate.setFixedWidth(180)
            self.label_exam_announceDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_announceDate, 0, 3)
            self.text_exam_announceDate = QLineEdit()
            self.text_exam_announceDate.setFixedWidth(180)
            self.grid.addWidget(self.text_exam_announceDate, 1, 3)

            self.label_exam_examDate = QLabel("시험일")
            self.label_exam_examDate.setFixedWidth(120)
            self.label_exam_examDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_examDate, 0, 4)
            self.text_exam_examDate = QLineEdit()
            self.text_exam_examDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_examDate, 1, 4)

            self.label_exam_passDate = QLabel("시험 합격일")
            self.label_exam_passDate.setFixedWidth(120)
            self.label_exam_passDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_passDate, 0, 5)
            self.text_exam_passDate = QLineEdit()
            self.text_exam_passDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_passDate, 1, 5)

            self.label_exam_submitDate = QLabel("서류 준비 일자")
            self.label_exam_submitDate.setFixedWidth(120)
            self.label_exam_submitDate.setAlignment(Qt.AlignCenter | Qt.AlignVCenter)
            self.grid.addWidget(self.label_exam_submitDate, 0, 6)
            self.text_exam_submitDate = QLineEdit()
            self.text_exam_submitDate.setFixedWidth(120)
            self.grid.addWidget(self.text_exam_submitDate, 1, 6)


        # 4. 버튼을 추가하기 위한 row와 column check
        # 이거 왜 자꾸 증가하냐;
        # totalRow = self.grid.rowCount()
        # totalCol = self.grid.columnCount()
        
        # 5. 재생성된 버튼 추가
        self.grid.addWidget(self.btn_insert, cnt_row, cnt_col - 2)
        self.grid.addWidget(self.btn_cancel, cnt_row, cnt_col - 1)

    # def closeEvent(self, QCloseEvent):
    #     ans = QMessageBox.question(self, "삽입 취소", "데이터 삽입을 취소하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
    #     if ans == QMessageBox.Yes:
    #         QCloseEvent.accept()
    #     else:
    #         QCloseEvent.ignore()



class mainLayout(QWidget, DB):
    def __init__(self):
        super(mainLayout, self).__init__()
        self.dbPrograms = DB()

        self.select_list_user = ["ID", "이름", "주민등록번호", "전화번호", "자격증", "주소", "본적주소", "기수", "반", "총 이수시간", "이론", "실기", "실습", "대체실습", "시험회차"]
        self.select_list_lecture = ["기수", "반", "시작일", "종료일"]
        self.select_list_teacher = ["ID", "분류", "이름", "생년월일", "자격증", "경력", "도 승인날짜"]
        self.select_list_facility = ["ID", "기관명", "구분", "계약 시작일", "계약 종료일", "1일 실습인원"]
        self.select_list_temptraining = ["기수", "시작일", "종료일", "수여일"]
        self.select_list_temptrainingteacher = ["기수", "강사"]
        self.select_list_exam = ["시험 회차", "접수 시작일", "접수 종료일", "응시표 출력", "시험일자", "합격자 발표(예정)", "서류 준비 기한"]

        self.auto = Automation()

        self.today = datetime.date.today()

        vbox = QVBoxLayout()
        first_hbox = QHBoxLayout()

        self.R_category = QComboBox(self)
        self.R_category.setFixedWidth(100)
        self.R_category.addItem("선택")

        self.R_searchBox = QLineEdit(self)
        self.R_searchBox.returnPressed.connect(self.search)

        self.R_searchBtn = QPushButton("검색", self)
        self.R_searchBtn.clicked.connect(self.search)

        first_hbox.addWidget(self.R_category)
        first_hbox.addWidget(self.R_searchBox)
        first_hbox.addWidget(self.R_searchBtn)

        second_hbox = QHBoxLayout()

        self.current_table = ""
        self.searching = False
        self.searchInfo = {"TABLE": '', "WHERE": '', "LIKE": '', "ORDER BY": ''}

        self.table = QTreeView(self)
        self.table.setAlternatingRowColors(True)
        self.table.setRootIsDecorated(False)
        # self.table.setFixedSize(800, 700)
        # QTreeView set read only
        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        self.readDB = QtGui.QStandardItemModel(0, 1, self)
        # Qt.Horizontal: 수평값. 기본적으로 넣어야 함.
        self.readDB.setHeaderData(0, Qt.Horizontal, "선택")

        self.table.setModel(self.readDB)
        self.table.clicked.connect(self.selected)

        self.layoutInfo = QVBoxLayout()
        self.labelInfo = QLabel("선택된 객체의 정보가 표시됩니다.")
        self.textInfo = QTextEdit()
        self.textInfo.setReadOnly(True)
        self.textInfo.setFixedWidth(400)
        # self.textInfo.setFontPointSize(12)
        self.textInfo.setCurrentFont(QtGui.QFont("맑은 고딕"))
        self.layoutInfo.addWidget(self.labelInfo)
        self.layoutInfo.addWidget(self.textInfo)

        # self.gridInfo = QGridLayout()
        
        # 마지막 인자 1 왜넘겨주는지 모르겠음.
        # self.readDB.insertRows(self.readDB.rowCount(), 1)
        # # 삽입된 정보는 read only로 바꿔줘야 함.
        # self.readDB.setData(self.readDB.index(0, 0), self.readDB.rowCount())
        # self.readDB.setData(self.readDB.index(0, 1), "이영민")
        # self.readDB.setData(self.readDB.index(0, 2), "990728-1234567")
        # self.readDB.setData(self.readDB.index(0, 3), "010-1234-5678")
        # self.readDB.setData(self.readDB.index(0, 4), "일반")
        # self.readDB.insertRows(self.readDB.rowCount(), 1)
        # self.readDB.insertRows(self.readDB.rowCount(), 1)
        # self.readDB.insertRows(self.readDB.rowCount(), 1)
        # self.readDB.insertRows(self.readDB.rowCount(), 1)

        second_hbox.addWidget(self.table)
        second_hbox.addLayout(self.layoutInfo)

        # second_hbox.addLayout(self.gridInfo)

        vbox.addLayout(first_hbox)
        vbox.addLayout(second_hbox)

        self.setLayout(vbox)        

    def selected(self):
        if self.current_table == "user":
            self.textInfo.clear()

            ID = "ID: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            name = "이름: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            RRN = "주민등록번호: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            phone = "전화번호: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            licen = "자격증: " + str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            adr = "주소: " + str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            oAdr = "본적주소: " + str(self.readDB.index(self.table.currentIndex().row(), 6).data())
            clsN = "기수: " + str(self.readDB.index(self.table.currentIndex().row(), 7).data())
            clsT = "반: " + str(self.readDB.index(self.table.currentIndex().row(), 8).data())
            totalH = "총 이수시간: " + str(self.readDB.index(self.table.currentIndex().row(), 9).data())
            theH = "이론: " + str(self.readDB.index(self.table.currentIndex().row(), 10).data())
            pracH = "실기: " + str(self.readDB.index(self.table.currentIndex().row(), 11).data())
            trainH = "실습: " + str(self.readDB.index(self.table.currentIndex().row(), 12).data())
            tempC = "대체실습: " + str(self.readDB.index(self.table.currentIndex().row(), 13).data())
            exam = "시험회차: " + str(self.readDB.index(self.table.currentIndex().row(), 14).data())

            send_string = ID + "\n\n" + name + "\n\n" + RRN + "\n\n" + phone + "\n\n" + licen + "\n\n" + adr + "\n\n" + oAdr + "\n\n" +\
                clsN + "\n\n" + clsT + "\n\n" + totalH + "\n\n" + theH + "\n\n" + pracH + "\n\n" + trainH + "\n\n" + tempC + "\n\n" + exam

        elif self.current_table == "lecture":
            self.textInfo.clear()

            clsN = "기수: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            clsT = "반: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            startD = "시작일: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            endD = "종료일: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data())

            send_string = clsN + "\n\n" + clsT + "\n\n" + startD + "\n\n" + endD
            
        elif self.current_table == "teacher":
            self.textInfo.clear()
            
            ID = "ID: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            categ = "분류: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            name = "이름: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            DOB = "생년월일: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            licen = "자격증: " + str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            career = "경력: " + str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            ACKDate = "도 승인일자: " + str(self.readDB.index(self.table.currentIndex().row(), 6).data())

            send_string = ID + "\n\n" + categ + "\n\n" + DOB + "\n\n" + licen + "\n\n" + career + "\n\n" + ACKDate

        elif self.current_table == "facility":
            self.textInfo.clear()

            ID = "ID: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            name = "기관명: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            categ = "구분: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            conStart = "계약 시작일: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            conEnd = "계약 종료일: " + str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            personnel = "1일 실습인원: " + str(self.readDB.index(self.table.currentIndex().row(), 5).data())

            send_string = ID + "\n\n" + name + "\n\n" + categ + "\n\n" + conStart + "\n\n" + conEnd + "\n\n" + personnel

        elif self.current_table == "temptraining":
            self.textInfo.clear()

            clsN = "기수: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            startD = "시작일: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            endD = "종료일: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            awardD = "수여일: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data())

            send_string = clsN + "\n\n" + startD + "\n\n" + endD + "\n\n" + awardD
            
        elif self.current_table == "temptrainingteacher":
            self.textInfo.clear()

            clsN = "기수: " + str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            teacher = "반: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())

            send_string = clsN + "\n\n" + teacher

        elif self.current_table == "exam":
            self.textInfo.clear()

            examRound = "시험회차: 제" + str(self.readDB.index(self.table.currentIndex().row(), 0).data()) + "회"
            examDueStart = "응시원서 접수 시작일: " + str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            examDueEnd = "응시원서 접수 종료일: " + str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            examTicket = "응시표 출력: " + str(self.readDB.index(self.table.currentIndex().row(), 3).data()) + " 부터"
            examDay = "시험일: " + str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            examPass = "합격자 발표 예정일: " + str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            examSubmit = "서류 준비 일자: " + str(self.readDB.index(self.table.currentIndex().row(), 6).data())

            send_string = examRound + "\n\n" + examDueStart + "\n\n" + examDueEnd + "\n\n" + examTicket + "\n\n" + examDay + "\n\n" + examPass + "\n\n" + examSubmit

        self.textInfo.setText(send_string)

    def selectTable(self):
        if self.current_table == "user":
            self.readDB.setColumnCount(15)
            self.readDB.setHorizontalHeaderLabels(self.select_list_user)

        elif self.current_table == "lecture":
            self.readDB.setColumnCount(4)
            self.readDB.setHorizontalHeaderLabels(self.select_list_lecture)

        elif self.current_table == "teacher":
            self.readDB.setColumnCount(7)
            self.readDB.setHorizontalHeaderLabels(self.select_list_teacher)

        elif self.current_table == "facility":
            self.readDB.setColumnCount(6)
            self.readDB.setHorizontalHeaderLabels(self.select_list_facility)

        elif self.current_table == "temptraining":
            self.readDB.setColumnCount(4)
            self.readDB.setHorizontalHeaderLabels(self.select_list_temptraining)

        elif self.current_table == "temptrainingteacher":
            self.readDB.setColumnCount(2)
            self.readDB.setHorizontalHeaderLabels(self.select_list_temptrainingteacher)

        elif self.current_table == "exam":
            self.readDB.setColumnCount(7)
            self.readDB.setHorizontalHeaderLabels(self.select_list_exam)

    def showTable(self, Refresh=False):
        source = self.sender()
        self.changeCategory(Refresh=Refresh)
        self.readDB.clear()

        if Refresh == False:
            self.R_searchBox.clear()
            self.searching = False

            if source.text() == "수강생 관리":
                self.current_table = "user"
                order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"

                self.readDB.setColumnCount(15)
                self.readDB.setHorizontalHeaderLabels(self.select_list_user)

            elif source.text() == "기수 관리":
                self.current_table = "lecture"
                order = "classNumber *1, FIELD(classTime, '주간', '야간')"

                self.readDB.setColumnCount(4)
                self.readDB.setHorizontalHeaderLabels(self.select_list_lecture)

            elif source.text() == "강사 관리":
                self.current_table = "teacher"
                order = "id *1"

                self.readDB.setColumnCount(7)
                self.readDB.setHorizontalHeaderLabels(self.select_list_teacher)

            elif source.text() == "실습기관 관리":
                self.current_table = "facility"
                order = "id *1, FIELD(category, '시설', '재가')"

                self.readDB.setColumnCount(6)
                self.readDB.setHorizontalHeaderLabels(self.select_list_facility)

            elif source.text() == "대체실습":
                self.current_table = "temptraining"
                order = "classNumber *1"

                self.readDB.setColumnCount(4)
                self.readDB.setHorizontalHeaderLabels(self.select_list_temptraining)

            elif source.text() == "대체실습 담당강사":
                self.current_table = "temptrainingteacher"
                order = "classNumber *1"

                self.readDB.setColumnCount(2)
                self.readDB.setHorizontalHeaderLabels(self.select_list_temptrainingteacher)

            elif source.text() == "국시원 시험":
                self.current_table = "exam"
                order = "round *1"

                self.readDB.setColumnCount(7)
                self.readDB.setHorizontalHeaderLabels(self.select_list_exam)

        elif Refresh == True:
            if self.current_table == "user":
                order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"

                self.readDB.setColumnCount(15)
                self.readDB.setHorizontalHeaderLabels(self.select_list_user)

            elif self.current_table == "lecture":
                order = "classNumber *1, FIELD(classTime, '주간', '야간')"

                self.readDB.setColumnCount(4)
                self.readDB.setHorizontalHeaderLabels(self.select_list_lecture)

            elif self.current_table == "teacher":
                order = "id *1"

                self.readDB.setColumnCount(7)
                self.readDB.setHorizontalHeaderLabels(self.select_list_teacher)

            elif self.current_table == "facility":
                order = "id *1, FIELD(category, '시설', '재가')"

                self.readDB.setColumnCount(6)
                self.readDB.setHorizontalHeaderLabels(self.select_list_facility)

            elif self.current_table == "temptraining":
                order = "classNumber *1"

                self.readDB.setColumnCount(4)
                self.readDB.setHorizontalHeaderLabels(self.select_list_temptraining)

            elif self.current_table == "temptrainingteacher":
                order = "classNumber *1"

                self.readDB.setColumnCount(2)
                self.readDB.setHorizontalHeaderLabels(self.select_list_temptrainingteacher)

            elif self.current_table == "exam":
                order = "round *1"
                
                self.readDB.setColumnCount(7)
                self.readDB.setHorizontalHeaderLabels(self.select_list_exam)

        if not self.searching:
            rs = self.dbPrograms.SELECT("*", self.current_table, orderBy=order)
        else:
            rs = self.dbPrograms.SELECT("*", self.searchInfo['TABLE'], where=f"{self.searchInfo['WHERE']} LIKE {self.searchInfo['LIKE']}", orderBy=self.searchInfo['ORDER BY'])

        if rs == "error":
            QMessageBox.information(self, "SQL query Error", "SQL query returns error!", QMessageBox.Yes, QMessageBox.Yes)
        else:
            cols = self.readDB.columnCount()
            for i in range(len(rs)):
                self.readDB.insertRows(self.readDB.rowCount(), 1)
                for j in range(cols):
                    string = str(rs[i][j])
                    if string == "None":
                        string = NULL
                    self.readDB.setData(self.readDB.index(i, j), string)

    def changeCategory(self, Refresh=False):
        source = self.sender()

        if Refresh == False:
            self.R_category.clear()
            if source.text() == "수강생 관리":
                self.R_category.addItem("이름")
                self.R_category.addItem("기수/반")
                self.R_category.addItem("대체실습")
                self.R_category.addItem("전화번호")
                self.R_category.addItem("생년월일")
                self.R_category.addItem("ID")
                self.R_category.addItem("자격증")
                self.R_category.addItem("시험회차")
                self.R_category.addItem("SQL")

            elif source.text() == "기수 관리":
                self.R_category.addItem("기수/반")
                self.R_category.addItem("SQL")

            elif source.text() == "강사 관리":
                self.R_category.addItem("ID")
                self.R_category.addItem("이름")
                self.R_category.addItem("자격증")
                self.R_category.addItem("SQL")

            elif source.text() == "실습기관 관리":
                self.R_category.addItem("기관명")
                self.R_category.addItem("구분")

            elif source.text() == "대체실습":
                self.R_category.addItem("기수")
                # 시작일은 검색어 "이후"의 날짜들 모두, 종료일은 검색어 "이전"의 날짜들 모두
                # (시작일을 2022-01-01로 검색할 경우 1월 1일 이후에 시작하는 기수 검색)
                # (종료일을 2022-01-01로 검색할 경우 1월 1일 이전에 종료된 기수 검색)
                self.R_category.addItem("시작일")
                self.R_category.addItem("종료일")
                self.R_category.addItem("수여일")
                self.R_category.addItem("SQL")

            elif source.text() == "대체실습 담당강사":
                self.R_category.addItem("기수")
                self.R_category.addItem("강사")
                self.R_category.addItem("SQL")

            elif source.text() == "국시원 시험":
                self.R_category.addItem("시험회차")

        elif Refresh == True:
            pass
            # if self.current_table == "user":
            #     self.R_category.addItem("이름")
            #     self.R_category.addItem("기수/반")
            #     self.R_category.addItem("대체실습")
            #     self.R_category.addItem("전화번호")
            #     self.R_category.addItem("생년월일")
            #     self.R_category.addItem("ID")
            #     self.R_category.addItem("자격증")
            #     self.R_category.addItem("시험회차")
            #     self.R_category.addItem("SQL")

            # elif self.current_table == "lecture":
            #     self.R_category.addItem("기수/반")
            #     self.R_category.addItem("SQL")

            # elif self.current_table == "teacher":
            #     self.R_category.addItem("ID")
            #     self.R_category.addItem("이름")
            #     self.R_category.addItem("자격증")
            #     self.R_category.addItem("SQL")

            # elif self.current_table == "facility":
            #     self.R_category.addItem("기관명")
            #     self.R_category.addItem("구분")

            # elif self.current_table == "temptraining":
            #     self.R_category.addItem("기수")
            #     # 시작일은 검색어 "이후"의 날짜들 모두, 종료일은 검색어 "이전"의 날짜들 모두
            #     # (시작일을 2022-01-01로 검색할 경우 1월 1일 이후에 시작하는 기수 검색)
            #     # (종료일을 2022-01-01로 검색할 경우 1월 1일 이전에 종료된 기수 검색)
            #     self.R_category.addItem("시작일")
            #     self.R_category.addItem("종료일")
            #     self.R_category.addItem("수여일")
            #     self.R_category.addItem("SQL")

            # elif self.current_table == "temptrainingteacher":
            #     self.R_category.addItem("기수")
            #     self.R_category.addItem("강사")
            #     self.R_category.addItem("SQL")

            # elif self.current_table == "exam":
            #     self.R_category.addItem("시험회차")

    def search(self):
        order = None
        keyWord = self.R_searchBox.text()
        if keyWord == "":
            QMessageBox.information(self, "검색어 오류", "검색어가 존재하지 않습니다!", QMessageBox.Yes, QMessageBox.Yes)
            return

        current_table = self.current_table
        current_category = self.R_category.currentText()

        if current_category == "ID":
            current_category = "id"

        elif current_category == "이름":
            current_category = "name"

        elif current_category == "자격증":
            current_category = "license"

        elif current_category == "기수/반":
            words = keyWord.split(" ")
            if len(words) == 1:
                if keyWord[-1] == "간":
                    current_category = "classTime"

                else:
                    current_category = "classNumber"
            
            elif len(words) == 2:
                if words[0][-1] == "간":
                    current_category = "classTime LIKE '%{}%' and classNumber".format(words[0])
                    keyWord = words[1]

                else:
                    current_category = "classNumber LIKE '%{}%' and classTime".format(words[0])
                    keyWord = words[1]

            if current_table == "user":
                order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"
            else:
                order = "classNumber *1, FIELD(classTime, '주간', '야간')"

        elif current_category == "전화번호":
            current_category = "phoneNumber"
            order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"

        elif current_category == "생년월일":
            current_category = "RRN"
            order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"

        elif current_category == "대체실습":
            current_category = "temporaryClassNumber"
            order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"

        elif current_category == "시험회차":
            if current_table == "user":
                current_category = "exam"
                order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"
            elif current_table == "exam":
                current_category = "round"
                order = "round *1"

        elif current_category == "기관명":
            current_category = "name"

        elif current_category == "구분":
            current_category = "category"
            order = "id *1"
            
        elif current_category == "시작일":
            current_category = "startDate"
            
        elif current_category == "종료일":
            current_category = "endDate"
            
        elif current_category == "수여일":
            current_category = "awardDate"
            
        elif current_category == "강사":
            current_category = "teacherName"
            order = "classNumber *1"

        elif current_category == "기수":
            current_category = "classNumber"
            order = "classNumber *1"

        elif current_category == "SQL":
            ans = QMessageBox.question(self, "SQL query문 전달", "SQL query문을 전달합니다.", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            if ans == QMessageBox.Yes:
                res = self.dbPrograms.SQL(keyWord)
                if res != "error":
                    QMessageBox.about(self, "완료", "query를 성공적으로 전달했습니다.")
                else:
                    QMessageBox.information(self, "SQL query Error", "SQL query returns error!", QMessageBox.Yes, QMessageBox.Yes)

                self.showTable(Refresh=True)
                self.textInfo.clear()
                return
            else:
                return
            
        else:
            QMessageBox.information(self, "Category Error", "error!", QMessageBox.Yes, QMessageBox.Yes)
            return

        try:
            self.readDB.clear()
            self.selectTable()
            if current_category == "name" or current_category == "teacherName":
                like = f"'%{keyWord}%'"

            elif current_category == "RRN":
                like = f"'{keyWord}%'"
            else:
                like = f"'%{keyWord}%'"
                
            rs = self.dbPrograms.SELECT("*", current_table, where=f"{current_category} LIKE {like}", orderBy=order)

            if rs == "error":
                QMessageBox.information(self, "SQL query Error", "SQL query returns error!", QMessageBox.Yes, QMessageBox.Yes)
            else:
                search_result = "{}, \"{}\" 검색 결과\n{}개의 검색 결과가 존재합니다.".format(self.R_category.currentText(), self.R_searchBox.text(), len(rs))
                self.textInfo.setText(search_result)
                cols = self.readDB.columnCount()
                for i in range(len(rs)):
                    self.readDB.insertRows(self.readDB.rowCount(), 1)
                    for j in range(cols):
                        string = str(rs[i][j])
                        if string == "None":
                            string = NULL
                        self.readDB.setData(self.readDB.index(i, j), string)
                        self.searching = True
                        self.searchInfo = {"TABLE": current_table, "WHERE": current_category, "LIKE": like, "ORDER BY": order}
        except:
            QMessageBox.information(self, "검색 오류", "잘못된 검색입니다.", QMessageBox.Yes, QMessageBox.Yes)
            return

    def DELETE(self):
        target_table = ""
        check = ""
        path = None
        if self.current_table == "user":
            target_table = "user"
            ID = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            name = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            RRN = str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            phone = str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            licen = str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            adr = str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            oAdr = str(self.readDB.index(self.table.currentIndex().row(), 6).data())
            clsN = str(self.readDB.index(self.table.currentIndex().row(), 7).data())
            clsT =  str(self.readDB.index(self.table.currentIndex().row(), 8).data())
            totalH = str(self.readDB.index(self.table.currentIndex().row(), 9).data())
            theH = str(self.readDB.index(self.table.currentIndex().row(), 10).data())
            pracH = str(self.readDB.index(self.table.currentIndex().row(), 11).data())
            trainH = str(self.readDB.index(self.table.currentIndex().row(), 12).data())
            tempC = str(self.readDB.index(self.table.currentIndex().row(), 13).data())

            query = "id = '{}' and name = '{}'".format(ID, name)

            check = "ID: {}\t이름: {}\t주민등록번호: {}\n전화번호: {}\t자격증: {}\n주소: {}\n본적주소: {}\n기수: {}\t반: {}\t 대체실습: {}\n총 이수시간: {}\t이론이수: {}\t실습이수: {}\t 실기이수: {}\n"\
                .format(ID, name, RRN, phone, licen, adr, oAdr, clsN, clsT, tempC, totalH, theH, pracH, trainH)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            path = "D:\\남양노아요양보호사교육원\\교육생관리\\{}\\{}{}\\{}".format(clsN, clsN, clsT, name)

            table = "수강생"
            classification = "{}{} {}".format(clsN, clsT, name)

        elif self.current_table == "lecture":
            target_table = "lecture"
            clsN = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            clsT = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            startD = str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            endD = str(self.readDB.index(self.table.currentIndex().row(), 3).data())

            query = "classNumber = '{}' and classTime = '{}'".format(clsN, clsT)

            check = "기수: {}\t반: {}\n시작일: {}\n종료일: {}\n".format(clsN, clsT, startD, endD)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            table = "기수"
            classification = "{}{}".format(clsN, clsT)
            
        elif self.current_table == "teacher":
            target_table = "teacher"
            ID = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            categ = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            name = str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            DOB = str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            licen = str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            career = str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            ACKDate = str(self.readDB.index(self.table.currentIndex().row(), 6).data())

            query = "name = '{}'".format(name)

            check = "ID: {}\t이름: {}\t자격증: {}\n생년월일: {}\t구분: {}\n최소경력: {}\n도 승인일자: {}\n"\
                .format(ID, name, licen, DOB, categ, career, ACKDate)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            table = "강사"
            classification = "{} / {}".format(name, DOB)

        elif self.current_table == "facility":
            target_table = "facility"
            ID = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            name = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            category = str(self.readDB.index(self.table.currentIndex().row(), 2).data())

            query = "id = '{}'".format(ID)
            check = "ID: {}\t기관명: {}\t구분: {}\n".format(ID, name, category)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            table = "기관"
            classification = "{} / {}".format(name, category)

        elif self.current_table == "temptraining":
            target_table = "temptraining"
            clsN = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            startD = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            endD = str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            awardD = str(self.readDB.index(self.table.currentIndex().row(), 3).data())

            check = "기수: {}\n시작일: {}\n종료일: {}\n수여일: {}\n".format(clsN, startD, endD, awardD)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            query = "classNumber = '{}'".format(clsN)

            table = "대체실습"
            classification = "{}".format(clsN)
            
        elif self.current_table == "temptrainingteacher":
            target_table = "temptrainingteacher"
            clsN = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            teacher = str(self.readDB.index(self.table.currentIndex().row(), 1).data())

            query = "classNumber = '{}' and teacherName = '{}'".format(clsN, teacher)

            check = "기수: {}\n강사: {}\n".format(clsN, teacher)
            check += "\n해당 정보를 데이터베이스에서 삭제합니다."

            table = "대체실습 강사"
            classification = "{} / {}".format(clsN, teacher)

        elif self.current_table == "exam":
            target_table = "exam"
            examRound = str(self.readDB.index(self.table.currentIndex().row(), 0).data())
            examDueStart = str(self.readDB.index(self.table.currentIndex().row(), 1).data())
            examDueEnd = str(self.readDB.index(self.table.currentIndex().row(), 2).data())
            examTicket = str(self.readDB.index(self.table.currentIndex().row(), 3).data())
            examDay = str(self.readDB.index(self.table.currentIndex().row(), 4).data())
            examPass = str(self.readDB.index(self.table.currentIndex().row(), 5).data())
            examSubmit = str(self.readDB.index(self.table.currentIndex().row(), 6).data())

            query = "round = {}".format(examRound)

            check = "시험회차: {}\n응시원서 접수 시작일: {}\t응시원서 접수 종료일: {}\n응시표 출력: {}\t시험일: {}\n합격자 발표 예정일: {}\t서류 준비 날짜: {}\n\n해당 정보를 데이터베이스에서 삭제합니다.".format(examRound, examDueStart, examDueEnd, examTicket, examDay, examPass, examSubmit)

            table = "시험"
            classification = "{}회".format(examRound)

        ans = QMessageBox.question(self, "데이터 삭제 확인", check, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            db.logger.info("$UI DELETE Request [TABLE|{}][{}] {} 삭제 요청".format(self.current_table, table, classification))
            db.main.dbPrograms.DELETE(target_table, query)
            if path != None and os.path.exists(path):
                ans_dir_delete = QMessageBox.question(self, "폴더 삭제", "해당 데이터의 폴더도 함께 삭제하시겠습니까?\n(해당 작업은 신중해야 합니다.)", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if ans_dir_delete == QMessageBox.Yes:
                    shutil.rmtree(path)

            QMessageBox.about(self, "완료", "데이터를 성공적으로 삭제했습니다.")
            self.showTable(Refresh=True)
            self.textInfo.clear()
        else:
            pass

    def isNULL(self):
        self.changeCategory(Refresh=True)
        self.readDB.clear()
        self.R_searchBox.clear()
        self.selectTable()

        if self.current_table == "":
            QMessageBox.information(self, "객체 오류", "테이블을 먼저 선택해주세요.", QMessageBox.Yes, QMessageBox.Yes)
            return

        elif self.current_table == "user":
            query = "id IS null or name IS null or RRN IS null or phoneNumber IS null or license IS null or address IS null or originAddress IS null or classNumber IS null or classTime IS null or totalCreditHour IS null or theoryCreditHour IS null or practicalCreditHour IS null or trainingCreditHour IS null or temporaryClassNumber IS null or exam IS null"
            order = "classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"
            
        elif self.current_table == "lecture":
            query = "classNumber IS null or classTime IS null or startDate IS null or endDate IS null"
            order = "classNumber *1, FIELD(classTime, '주간', '야간')"

        elif self.current_table == "teacher":
            query = "id IS null or category IS null or name IS null or dateOfBirth IS null or license IS null or minCareer IS null or ACKDate IS null"
            order = "id"

        elif self.current_table == "facility":
            query = "id IS null or name IS null or category IS null or contractTermStart IS null or contractTermEnd IS null or personnel IS null"
            order = "id"

        elif self.current_table == "temptraining":
            query = "classNumber IS null or startDate IS null or endDate IS null or awardDate IS null"
            order = "classNumber"

        elif self.current_table == "temptrainingteacher":
            query = "classNumber IS null or teacherName IS null"
            order = "classNumber"

        elif self.current_table == "exam":
            query = "round IS null or startAcceptance IS null or endAcceptance IS null or announceDate IS null or examDate IS null or passDate IS null or submitDate IS null"
            order = "round *1"


        order += " *1"
        rs = self.dbPrograms.SELECT("*", self.current_table, where=query, orderBy=order)
        if rs == "error":
            QMessageBox.information(self, "SQL query Error", "SQL query returns error!", QMessageBox.Yes, QMessageBox.Yes)
        else:
            cols = self.readDB.columnCount()
            for i in range(len(rs)):
                self.readDB.insertRows(self.readDB.rowCount(), 1)
                for j in range(cols):
                    string = str(rs[i][j])
                    if string == "None":
                        string = NULL
                    self.readDB.setData(self.readDB.index(i, j), string)

    def readFiles(self):
        """
        type. 1
        1. pre_file, next_file 같이 읽어서, 이름을 찾는다.
        2. 2-1. 이름이 같으면 등본에서 이름, 현주소, 주민등록번호를, 기본증명서에서 등록기준지를 가져온다.
        2. 2-2. 이름이 다르다 -> 
        이름이 같은지 확인하는 방법? 알고리즘 여러개 따보자!
        1) image -> string -> list 변환 후 check each value's length -> 3 <= len(list[i]) <=5(?) ==> 이름으로 판단한다!

        type. 2
        1. read_file -> 이미지를 읽어서 우측에 작은 화면으로 띄운다.(win+-> or 방법 있으면 사용)
        2. read_file의 제목을 읽어 주민등록 등본인지, 기본증명서인지 check
        3. insert 창을 active 시키고, 제목에 따라 [이름, 주민번호 == 공통] and ([현주소 == 등본] or [본적주소 == 기본증명서])
            를 읽어서 LineEdit에 먼저 값을 넣는다.
        4. insert 창에는 '데이터 입력', '건너뛰기', '삽입', '취소' 버튼이 존재한다.
        5. '데이터 입력' 클릭 시, read_file의 제목에 따라 [이름, 주민번호 == 공통] and ([현주소 == 등본] or [본적주소 == 기본증명서])
            를 읽어서 LineEdit에 값을 추가한다.(등본 입력 후 기본증명서 입력 시 사용. LineEdit value != ""일 경우에 추가입력.)
        6. '건너뛰기' 클릭 시 데이터 추가를 하지 않고, 다음 파일로 건너 뛴다. 이미지 창을 닫고 다음 이미지를 같은 크기로 active.
        7. '삽입' 클릭 시 입력된 데이터를 기반으로 Database에 추가하고, 파일 명을 변경하여 해당 수강생의 폴더로 옮긴다.
        8. '취소' 입력 시 함수를 return하고 함수를 종료한다.
        """
        file_path = ""  # scan file path
        file_type = ""  # 주민등록등본 or 기본증명서
        name = ""
        RRN = ""
        adr = ""
        oAdr = ""

        pass

    def backupToExcel(self):
        # print(self.readDB.rowCount()) row 개수. 1개면 1
        # IDF = str(self.readDB.index(1, 0).data()) row 접근 인덱스는 0 ~ rowCount() - 1
        # IDL = str(self.readDB.index(242, 0).data())
        # print("IDF:", IDF, "\nIDL:", IDL)

        today = datetime.datetime.today().strftime("%Y%m%d")
        today = today[2:]
        file_name = ""
        self.backUpWbook = Workbook()
        ws = self.backUpWbook.active
        col = []
        dimension_lst = []

        if self.current_table == "":
            QMessageBox.information(self, "오류", "테이블을 먼저 선택해주세요.", QMessageBox.Yes, QMessageBox.Yes)
            return

        if self.current_table == "user":
            col = self.select_list_user
            dimension_lst = [9, 10, 20, 20, 15, 73, 63, 10, 10, 10, 5, 5, 5, 10, 9]
            file_name = "수강생DB"

        elif self.current_table == "lecture":
            col = self.select_list_lecture
            dimension_lst = [10, 10, 15, 15]
            file_name = "기수,반DB"

        elif self.current_table == "teacher":
            col = self.select_list_teacher
            dimension_lst = [9, 12, 10, 15, 15, 15, 15]
            file_name = "강사DB"

        elif self.current_table == "facility":
            col = self.select_list_facility
            dimension_lst = [9, 20, 10, 15, 15, 9]
            file_name = "기관DB"

        elif self.current_table == "temptraining":
            col = self.select_list_temptraining
            dimension_lst = [10, 15, 15, 15]
            file_name = "대체실습DB"

        elif self.current_table == "temptrainingteacher":
            col = self.select_list_temptrainingteacher
            dimension_lst = [10, 10]
            file_name = "대체실습 강사DB"

        elif self.current_table == "exam":
            col = self.select_list_exam
            dimension_lst = [15, 15, 15, 15, 15, 15, 15]
            file_name = "시험DB"

        for i, val in enumerate(col, start=1):
            ws.cell(row=1, column=i).value = val
            ws.column_dimensions[get_column_letter(i)].width = dimension_lst[i - 1]

        if self.current_table == "user":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())

        elif self.current_table == "lecture":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())
            
        elif self.current_table == "teacher":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())

        elif self.current_table == "facility":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())

        elif self.current_table == "temptraining":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())
            
        elif self.current_table == "temptrainingteacher":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())

        elif self.current_table == "exam":
            for i in range(self.readDB.rowCount()):
                for j in range(len(col)):
                    if str(self.readDB.index(i, j).data()) == "NULL":
                        continue
                    else:
                        ws.cell(row=i + 2, column=j + 1).value = str(self.readDB.index(i, j).data())

        save_path = f"D:\\남양노아요양보호사교육원\\데이터베이스 백업\\{today}_{file_name}.xlsx"

        ans = QMessageBox.question(self, "Back up", f"{file_name}를 Excel 파일로 생성합니다.", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if ans == QMessageBox.Yes:
            self.backUpWbook.save(save_path)
            QMessageBox.about(self, "완료", "데이터를 성공적으로 수정했습니다.\n경로: {}".format(save_path))
            db.logger.info("$UI [Backup to Excel] {}.xlsx파일 생성 완료.".format(file_name))
        else:
            QMessageBox.about(self, "취소", "데티어 백업을 취소했습니다.")


class DBMS(QMainWindow):
    # 새 창을 띄우기 위해 서로 global로 연결
    global log_in_window
    global insert
    global update
    global batch
    global report_gov
    global opening
    global kuksiwon
    global todo_list
    global wi

    global scanner

    def __init__(self):
        super().__init__()
        self.logger = logging.getLogger("UI log")
        fileHandler = logging.FileHandler("D:\\Master\\log\\Program log.log")

        formatter = logging.Formatter('[%(asctime)s][%(levelname)s|%(filename)s:%(lineno)s] in <%(funcName)s> %(name)s >> %(message)s')
        fileHandler.setFormatter(formatter)

        self.logger.addHandler(fileHandler)
        self.logger.setLevel(level=logging.DEBUG)

        self.initUI()

    def initUI(self):
        self.setWindowTitle("NYNOA DBMS")
        self.setWindowIcon(QIcon("D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\남양노아요양보호사-배경제거.png"))
        self.setGeometry(350, 150, 1200, 800)

        self.icon_path = "D:\\Master\\PythonWorkspace\\NYNOA\\Icons\\"

        # status Bar
        # self.statusBar()
        self.statusBar().showMessage("상태바")
        self.main = mainLayout()

        self.setCentralWidget(self.main)

        self.menuOpt()
        self.logger.debug("DBMS Program is running")

        # self.show()

    # 상단 menu
    def menuOpt(self):
        # menu Bar
        menuBar = self.menuBar()    # menu 생성
        menu_file = menuBar.addMenu("File") # group 생성
        menu_edit = menuBar.addMenu("Edit")
        menu_view = menuBar.addMenu("View")

        menu_stu = menuBar.addAction("수강생 관리")
        menu_lecture = menuBar.addAction("기수 관리")
        menu_teach = menuBar.addAction("강사 관리")
        menu_facility = menuBar.addAction("실습기관 관리")
        menu_temp = menuBar.addAction("대체실습")
        menu_tempTeach = menuBar.addAction("대체실습 담당강사")
        menu_exam = menuBar.addAction("국시원 시험")
        menu_isNULL = menuBar.addAction("미입력 데이터")
        menu_todo = menuBar.addAction("To Do List")

        menu_isNULL.setStatusTip("데이터가 입력되지 않은 컬럼들을 찾습니다.")
        menu_isNULL.setShortcut("Ctrl+Shift+Q")

        file_exit = QAction(QIcon(self.icon_path + "cross.png"), 'Exit', self)   # menu 객체 생성
        file_exit.setShortcut("Ctrl+Q")
        file_exit.setStatusTip("나가기")
        # file_exit.triggered.connect(QCoreApplication.instance().quit)   # 종료 기능 추가 / self.close()로도 종료 가능
        file_exit.triggered.connect(self.close)      # 위와 같은 기능. 메서드를 전달하는 것이기 때문에 ()없이!

        file_db = QMenu("DataBase", self)
        db_dump = QAction(QIcon(self.icon_path + "database.png"), "DB 내보내기", self)
        db_apply = QAction(QIcon(self.icon_path + "database.png"), "DB 가져오기", self)
        db_sort = QAction(QIcon(self.icon_path + "sort.png"), "id 정렬", self)

        file_db.addAction(db_sort)
        file_db.addAction(db_dump)
        file_db.addAction(db_apply)

        db_sort.triggered.connect(self.idSort)
        db_dump.triggered.connect(self.databaseManagement)
        db_apply.triggered.connect(self.databaseManagement)

        file_wi = QAction(QIcon(self.icon_path + "ribbon.png"), "인수인계사항", self)
        file_wi.setStatusTip("기관장이 해야할 일 리스트 입니다.")
        file_wi.triggered.connect(self.workingInformation_show)

        file_new = QMenu("New", self)   # sub menu 객체 생성
        new_data = QAction(QIcon(self.icon_path + "new.png"), "데이터 추가", self)
        new_data.setShortcut("Ctrl+N")
        new_data.setStatusTip("선택된 테이블에 새로운 데이터를 추가합니다.")

        new_payment = QAction(QIcon(self.icon_path + "documents.png"), "수강료 수납대장", self)
        new_payment.setStatusTip("[수강료 수납대장: 원장님 서류]를 작성합니다.")

        new_locker = QAction(QIcon(self.icon_path + "list.png"), "사물함 주기", self)
        new_locker.setStatusTip("사물함 주기(명찰)파일을 생성합니다.")
        
        file_new.addAction(new_data)
        file_new.addAction(new_payment)
        file_new.addAction(new_locker)

        file_scan = QMenu("Scan", self)
        read_data = QAction("파일 스캔", self)

        file_scan.addAction(read_data)

        file_backUp = QAction(QIcon(self.icon_path + "backUp.png"), "Back up", self)
        file_backUp.setStatusTip("현재 선택된 데이터베이스 테이블을 엑셀 파일로 생성해 백업합니다.")
        file_backUp.triggered.connect(self.main.backupToExcel)

        file_report = QMenu("경기도청", self)
        beginning_lecture = QAction(QIcon(self.icon_path + "government.png"), "개강보고", self)
        attendence = QAction(QIcon(self.icon_path + "government.png"), "출석부", self)
        implement_temp_class = QAction(QIcon(self.icon_path + "government.png"), "대체실습 실시보고", self)
        complete_temp_class = QAction(QIcon(self.icon_path + "government.png"), "대체실습 수료보고", self)

        file_report.addAction(beginning_lecture)
        file_report.addAction(attendence)
        file_report.addAction(implement_temp_class)
        file_report.addAction(complete_temp_class)

        beginning_lecture.triggered.connect(self.report_gov_show)
        attendence.triggered.connect(self.report_gov_show)
        implement_temp_class.triggered.connect(self.report_gov_show)
        complete_temp_class.triggered.connect(self.report_gov_show)
        
        # menu에 addAction 할 경우, 이렇게 하면 안되고, 함수를 따로 생성해서 넘겨주어야 한다. 이유는 모름.
        # new_data.triggered.connect(insert.show())
        new_data.triggered.connect(self.INSERT_show)
        new_payment.triggered.connect(self.opening_show)
        new_locker.triggered.connect(self.opening_show)

        file_kuksiwon = QMenu("국시원", self)
        members = QAction(QIcon(self.icon_path + "kuksiwon.png"), "응시접수 명단", self)
        pass_list = QAction(QIcon(self.icon_path + "kuksiwon.png"), "합격자 명단", self)
        document_1 = QAction(QIcon(self.icon_path + "document.png"), "교육수료증명서", self)
        document_2 = QAction(QIcon(self.icon_path + "document.png"), "대체실습확인서", self)
        document_3 = QAction(QIcon(self.icon_path + "document.png"), "요양보호사 자격증 발급,재발급 신청서", self)
        print_document_1 = QAction(QIcon(self.icon_path + "printer.png"), "출력|교육수료증명서", self)
        print_document_2 = QAction(QIcon(self.icon_path + "printer.png"), "출력|대체실습확인서", self)
        print_document_3 = QAction(QIcon(self.icon_path + "printer.png"), "출력|요양보호사 자격증 발급,재발급 신청서", self)
        gather_pictures = QAction(QIcon(self.icon_path + "picture.png"), "합격자 사진", self)
        

        file_kuksiwon.addAction(members)
        file_kuksiwon.addAction(pass_list)
        file_kuksiwon.addAction(document_1)
        file_kuksiwon.addAction(document_2)
        file_kuksiwon.addAction(document_3)
        file_kuksiwon.addAction(print_document_1)
        file_kuksiwon.addAction(print_document_2)
        file_kuksiwon.addAction(print_document_3)
        file_kuksiwon.addAction(gather_pictures)

        members.triggered.connect(self.kuksiwon_show)
        pass_list.triggered.connect(self.kuksiwon_show)
        document_1.triggered.connect(self.kuksiwon_show)
        document_2.triggered.connect(self.kuksiwon_show)
        document_3.triggered.connect(self.kuksiwon_show)
        print_document_1.triggered.connect(self.kuksiwon_show)
        print_document_2.triggered.connect(self.kuksiwon_show)
        print_document_3.triggered.connect(self.kuksiwon_show)
        gather_pictures.triggered.connect(self.kuksiwon_show)

        read_data.setShortcut("Ctrl+F")
        read_data.setStatusTip("폴더를 스켄하여 데이터베이스에 데이터를 삽입합니다.")
        read_data.triggered.connect(self.scan_show)
        
        edit_batch = QMenu("일괄 변경", self)
        batch_data_exam = QAction(QIcon(self.icon_path + "batchEdit.png"), "시험회차 일괄 변경", self)
        batch_data_exam.setStatusTip("특정 기수, 반의 시험 회차를 일괄적으로 설정합니다.")

        batch_data_temp = QAction(QIcon(self.icon_path + "batchEdit.png"), "대체실습 일괄 변경", self)
        batch_data_temp.setStatusTip("특정 기수, 반의 대체실습 기수를 일괄적으로 변경합니다.")

        batch_data_time = QAction(QIcon(self.icon_path + "time.png"), "이수시간 일괄 변경", self)
        batch_data_time.setStatusTip("특정 기수, 반의 이론, 실기, 실습 이수시간을 일괄적으로 변경합니다.")

        batch_data_exam.triggered.connect(self.batch_show)
        batch_data_temp.triggered.connect(self.batch_show)
        batch_data_time.triggered.connect(self.batch_show)

        edit_batch.addAction(batch_data_exam)
        edit_batch.addAction(batch_data_temp)
        edit_batch.addAction(batch_data_time)

        mod_data = QAction(QIcon(self.icon_path + "edit.png"), "데이터 수정", self)
        mod_data.setShortcut("F2")
        mod_data.setStatusTip("테이블에서 선택된 데이터를 수정합니다.")
        mod_data.triggered.connect(self.UPDATE_show)

        del_data = QAction(QIcon(self.icon_path + "delete.png"), "데이터 삭제", self)
        del_data.setShortcut(Qt.Key_Delete)
        del_data.setStatusTip("테이블에서 선택된 데이터를 삭제합니다.")
        del_data.triggered.connect(self.main.DELETE)

        view_stat = QAction("가이드 표시", self, checkable=True)
        view_stat.setChecked(True)
        view_stat.triggered.connect(self.triState)

        menu_file.addMenu(file_db)
        menu_file.addAction(file_wi)
        menu_file.addAction(file_backUp)
        menu_file.addMenu(file_kuksiwon)
        menu_file.addMenu(file_report)
        menu_file.addMenu(file_scan)
        menu_file.addMenu(file_new)     # sub menu 등록
        menu_file.addAction(file_exit)  # menu 등록(액션 추가)
        menu_view.addAction(view_stat)

        menu_edit.addMenu(edit_batch)
        menu_edit.addAction(mod_data)
        menu_edit.addAction(del_data)

        menu_stu.triggered.connect(self.main.showTable)
        menu_lecture.triggered.connect(self.main.showTable)
        menu_teach.triggered.connect(self.main.showTable)
        menu_facility.triggered.connect(self.main.showTable)
        menu_temp.triggered.connect(self.main.showTable)
        menu_tempTeach.triggered.connect(self.main.showTable)
        menu_exam.triggered.connect(self.main.showTable)

        menu_isNULL.triggered.connect(self.main.isNULL)
        menu_todo.triggered.connect(self.todo_show)

    def scan_show(self):
        QMessageBox.information(self, "알림", "기능이 준비되지 않았습니다!", QMessageBox.Yes, QMessageBox.Yes)
        return
        path = "D:\\scan"
        file_list = []

        # print(os.listdir(path))

        for f in os.listdir(path):
            file_name, ext = os.path.splitext(f)
            if ext == ".jpg":
                file_list.append(str(path) + "\\" + str(f))
        
        scanner.file_list = file_list
        scanner.show()

    def idSort(self):
        cnt = 0
        ans = QMessageBox.question(self, "Database", "id 정렬을 시작합니다.", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            res = self.main.dbPrograms.SELECT("name, RRN", "user")

        for idx, rows in enumerate(res, start=1):
            name = rows[0]
            RRN = rows[1]

            self.main.dbPrograms.UPDATE("user", "id={}".format(idx), "name='{}' and RRN = '{}'".format(name, RRN))
            print(name, RRN, "완료")
            cnt = idx

        QMessageBox.information(self, "완료", "{}개의 데이터를 수정하였습니다.".format(cnt), QMessageBox.Yes, QMessageBox.Yes)

    def databaseManagement(self):
        source = self.sender()
        msg = ""
        if source.text() == "DB 내보내기":
            msg = "DB를 내보냅니다."
        elif source.text() == "DB 가져오기":
            msg = "DB를 가져옵니다."

        ans = QMessageBox.question(self, "Database", msg, QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        if ans == QMessageBox.Yes:
            if source.text() == "DB 내보내기":
                fname = QFileDialog.getSaveFileName(self, "Save file", r"C:\Bitnami\wampstack-8.1.1-0\mariadb\bin\database_dump\*.sql", "SQL File(*.sql)")
                if fname[-1] == "":
                    QMessageBox.information(self, "취소", "취소되었습니다.")
                    return
                self.main.dbPrograms.dumpDatabase(file_path=fname[0])

            elif source.text() == "DB 가져오기":
                fname = QFileDialog.getOpenFileName(self, "Open file", r"C:\Bitnami\wampstack-8.1.1-0\mariadb\bin\database_dump\*.sql", "SQL File(*.sql)")
                # fname: ('C:/Bitnami/wampstack-8.1.1-0/mariadb/bin/database_dump/ac_bak_2022-02-09.sql', 'All Files (*)')
                if fname[-1] == "":
                    QMessageBox.information(self, "취소", "취소되었습니다.")
                    return
                db_change = QMessageBox.question(self, "DB 변경", "데이터베이스와 연결을 끊고, DB를 선택된 DB로 변경합니다.", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
                if db_change == QMessageBox.Yes:
                    self.main.dbPrograms.applyDatabase(fname[0])

                else:
                    QMessageBox.information(self, "취소", "취소되었습니다.")
                    return
                
                QMessageBox.information(self, "완료", "데이터베이스가 변경되었습니다. 프로그램을 다시 시작해주세요.", QMessageBox.Yes, QMessageBox.Yes)
                sys.exit()

        QMessageBox.information(self, "완료", "완료되었습니다.", QMessageBox.Yes, QMessageBox.Yes)
            

    def workingInformation_show(self):
        wi.show()

    def todo_show(self):
        todo_list.show()

    def kuksiwon_show(self):
        kuksiwon.show()

    def opening_show(self):
        opening.show()

    def report_gov_show(self):
        source = self.sender()
        report_gov.doc_type = source.text()
        report_gov.show()

    def batch_show(self):
        source = self.sender()
        batch.mode = source.text()
        batch.show()

    def UPDATE_show(self):
        if self.main.textInfo.toPlainText() == "":
            QMessageBox.information(self, "객체 오류", "객체를 먼저 선택해주세요.",
            QMessageBox.Yes, QMessageBox.Yes)

        else:
            update.show()


    def INSERT_show(self):
        if self.main.current_table == "":
            QMessageBox.information(self, "테이블 오류", "테이블을 먼저 선택해주세요.\n상단 메뉴에 테이블이 존재합니다.",
            QMessageBox.Yes, QMessageBox.Yes)

        else:
            insert.show()

    def triState(self, state):
        if state:
            self.statusBar().show()
        else:
            self.statusBar().hide()

    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Escape:
            print("ESC is pressed!")

    # context menu. 우클릭 메뉴
    def contextMenuEvent(self, QContextMenuEvent):
        cm = QMenu(self)
        
        quit = cm.addAction("Quit")

        # action: cm의 실행정보를 저장. 전체적인 map의 위치를 넘겨서 우클릭 하는 위치에 따라 다른 이벤트를 적용하도록
        action = cm.exec_(self.mapToGlobal(QContextMenuEvent.pos()))

        if action == quit:
            self.close()


    def closeEvent(self, QCloseEvent):
        # x(창닫기) 버튼을 눌렀을 경우!
        # QMessageBox.question(인자, title, message, 버튼 추가(여러개 가능(|사용)), 버튼 기본값)
        ans = QMessageBox.question(self, "종료", "DBMS를 종료하시겠습니까?", QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
        
        if ans == QMessageBox.Yes:
            ####################### 여기다가 conn.close()추가하기 !!!!!!!!!!!!!!!!!!!!!
            try:
                self.main.dbPrograms.conn.close()
            except:
                print("DBGUI Exception: Database is already closed!")
            QCloseEvent.accept()
        else:
            QCloseEvent.ignore()


if __name__ == '__main__':
    app = QApplication(sys.argv)
    db = DBMS()
    log_in_window = LogIn()
    todo_list = ToDoList()
    insert = INSERT()
    update = UPDATE()
    batch = BatchUpdate()
    report_gov = Report()
    opening = ClassOpening()
    kuksiwon = Kuksiwon()
    wi = WorkingInformation()



    # calender = Calender()
    # calender.show()


    # scanner = scanFile()
    sys.exit(app.exec_())