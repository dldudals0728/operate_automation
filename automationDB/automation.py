from inspect import istraceback
import logging
import os
import random
import shutil
import traceback
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image
from database import DB

class Automation:
    def __init__(self):
        self.logFile = "D:\\Master\\log\\"
        self.makeFilePath = "D:\\Master\\mkfile\\"
        self.docFilePath = "D:\\Master\\files\\"
        self.basePath = "D:\\남양노아요양보호사교육원\\교육생관리\\"
        self.imsi_workbook_path = "D:\\Master\\PythonWorkspace\\imsi.xlsx"
        self.wb = None
        self.ws = None
        self.DB = DB()
        self.wb_imsi = None
        self.ws_imsi = None

        self.logger = logging.getLogger("AUTOMATION log")
        fileHandler = logging.FileHandler("D:\\Master\\log\\Program log.log")

        formatter = logging.Formatter('[%(asctime)s][%(levelname)s|%(filename)s:%(lineno)s] in <%(funcName)s> %(name)s >> %(message)s')
        fileHandler.setFormatter(formatter)

        self.logger.addHandler(fileHandler)
        self.logger.setLevel(level=logging.DEBUG)

    """
    교육수료증명서를 어떻게 일괄적으로 출력 할 것인가!
    1. 기수, 반 별로 생성 및 출력한다.  X
    2. 시험 회차에 따라 일괄 생성 및 출력한다.  O
    3. 개인적으로 생성 및 출력한다. (일일이!!! ) ==> 이건 좀 필요할 듯. 누군가 누락됐을 때 생성할 필요 있음!    O
    """
    def inputChecker(self, res):
        """DB 검색을 통해 반환된 튜플의 입력 여부를 검사합니다."""
        input_list = []
        for rows in res:
            imsi_list = []
            for rs in rows:
                if rs == None:
                    imsi_list.append("")
                else:
                    imsi_list.append(rs)

            input_list.append(tuple(imsi_list))

        return tuple(input_list)

    def nullValueChecker(self, res, checker_index):
        null_list = []
        for rows in res:
            null_check_switch = False
            for rs in rows:
                if rs == None or rs == "":
                    null_check_switch = True
            if null_check_switch == True:
                null_list.append(rows[checker_index])

        return null_list

    def makeDocument(self, exam, doc_type):
        file_path = self.makeFilePath + "\\{}.xlsx".format(doc_type)
        print(file_path)
        self.logger.info("$Automation [Document|{}][Exam|{}회]작성".format(doc_type, exam))
        if doc_type == "교육수료증명서":
            try:
                where = "exam={}".format(exam)
                user_rs = self.inputChecker(self.DB.SELECT("*", "user", where))
                valueErrorList = self.nullValueChecker(user_rs, 1)

                user_query_list = ["id", "name", "RRN", "phoneNumber", "license", "address", "originAddress", "classNumber", "classTime", \
                    "totalCreditHour", "theoryCreditHour", "practicalCreditHour", "trainingCreditHour", "temporaryClassNumber", "exam"]
                item_dict = {}

                for rows in user_rs:
                    item_dict.clear()
                    for index in range(len(rows)):
                        item_dict[user_query_list[index]] = rows[index]

                    save_path = self.basePath + "{}\\{}{}\\{}".format(item_dict["classNumber"], item_dict["classNumber"], item_dict["classTime"], item_dict["name"])

                    self.wb = load_workbook(file_path)
                    self.ws = self.wb.active

                    where = "classNumber = '{}' and classTime = '{}'".format(item_dict["classNumber"], item_dict["classTime"])
                    classInfo_rs = self.DB.SELECT("*", "lecture", where, fetchone=True)

                    item_dict["startDate"] = classInfo_rs[2].strftime("%Y 년  %m 월  %d 일")
                    item_dict["endDate"] = classInfo_rs[3].strftime("%Y 년  %m 월  %d 일")

                    where = "classNumber='{}'".format(item_dict["temporaryClassNumber"])
                    tempInfo_rs = self.DB.SELECT("*", "temptraining", where, fetchone=True)

                    item_dict["startDate_temp"] = tempInfo_rs[1].strftime("%Y 년 %m 월 %d 일")
                    item_dict["endDate_temp"] = tempInfo_rs[2].strftime("%Y 년 %m 월 %d 일")
                    item_dict["awardDate"] = tempInfo_rs[3].strftime("%Y 년    %m 월     %d 일")

                    # 교육수료증명서 호수
                    string = "    {}  년  제  {} 호".format(item_dict["awardDate"][:4], item_dict["id"])
                    self.ws.cell(row=1, column=1).value = string

                    # 이름
                    string = " {}".format(" ".join(item_dict["name"]))
                    self.ws.cell(row=4, column=3).value = string

                    # 주소
                    string = " {}".format(item_dict["address"])
                    self.ws.cell(row=5, column=3).value = string

                    # 주민등록번호
                    string = " {}".format(str(item_dict["RRN"]))
                    self.ws.cell(row=6, column=3).value = string

                    # 전화번호
                    string = item_dict["phoneNumber"]
                    self.ws.cell(row=6, column=6).value = string

                    # 교육과정명
                    string = " 요양보호사 {}".format(item_dict["classNumber"])
                    self.ws.cell(row=7, column=3).value = string

                    # 이론실기 이수기간 / 각 기수별로 기간 선정, 2020 년  11 월  16 일 ∼  21 년  01 월 15 일 형식으로, 끝기간은 년도수 두자리만 표시
                    string = "{} ~ {}".format(item_dict["startDate"], item_dict["endDate"])
                    self.ws.cell(row=9, column=3).value = string

                    # 이론실기 이수시간
                    string = "        {}  시간".format(int(item_dict["theoryCreditHour"]) + int(item_dict["practicalCreditHour"]))
                    self.ws.cell(row=9, column=7).value = string

                    # 실습기간 / 대체실습 각 기수별 or 실습기간 따로 만들기,  21년 01월 18일 ∼ 21년 03월 13일 형식으로, 년도수 두자리만 표시
                    string = "{} \n~ {}".format(item_dict["startDate_temp"], item_dict["endDate_temp"])
                    self.ws.cell(row=12, column=4).value = string

                    # 대체실습이 종료되면, 각 사람마다 실습시간(각 기관) 이 달라짐. 업데이트 필요
                    # 실습시간
                    string = "        {}  시간".format(item_dict["trainingCreditHour"])
                    self.ws.cell(row=12, column=7).value = string

                    # 총 실습시간
                    string = "         {}  시간".format(item_dict["trainingCreditHour"])
                    self.ws.cell(row=18, column=7).value = string

                    # 총 이수시간
                    string = "       {}  시간".format(item_dict["totalCreditHour"])
                    self.ws.cell(row=19, column=7).value = string

                    # 수여일 / 각 인원 대체실습 기준 종료일 바로 다음 월요일 날짜로 지정
                    string = "                                      {}".format(item_dict["awardDate"])
                    self.ws.cell(row=23, column=1).value = string

                    self.wb.save(save_path + "\\{}_{}.xlsx".format(item_dict["name"], doc_type))
                    self.logger.info("$Automation [Document|교육수료증명서][{}{} {}]작성".format(item_dict["classNumber"], item_dict["classTime"], item_dict["name"]))
                    self.wb.close()

                return_str = "입력 오류: "
                if valueErrorList == []:
                    return_str += "모두 정상 처리되었습니다."
                else:
                    return_str += ", ".join(valueErrorList)
                    self.logger.error("!Automation [Document|교육수료증명서] 미처리 항목: {}".format(return_str))

                return return_str

            except:
                return traceback.format_exc()

        elif doc_type == "대체실습확인서":
            try:
                where = "exam={}".format(exam)
                user_rs = self.inputChecker(self.DB.SELECT("id, name, RRN, phoneNumber, classNumber, classTime, trainingCreditHour, temporaryClassNumber", "user", where))
                valueErrorList = self.nullValueChecker(user_rs, 1)

                user_query_list = ["id", "name", "RRN", "phoneNumber", "classNumber", "classTime", "trainingCreditHour", "temporaryClassNumber"]
                item_dict = {}

                for rows in user_rs:
                    item_dict.clear()
                    for index in range(len(rows)):
                        item_dict[user_query_list[index]] = rows[index]

                    save_path = self.basePath + "{}\\{}{}\\{}".format(item_dict["classNumber"], item_dict["classNumber"], item_dict["classTime"], item_dict["name"])

                    self.wb = load_workbook(file_path)
                    self.ws = self.wb.active

                    teacher_rs = self.DB.SELECT("*", "teacher")
                    teacher_dict = {}
                    for rows in teacher_rs:
                        teacher_dict[rows[2]] = rows

                    where = "classNumber = '{}'".format(item_dict["temporaryClassNumber"])
                    temp_teacher_rs = self.DB.SELECT("teacherName", "temptrainingteacher", where=where)
                    temp_teacher_list = []
                    for rows in temp_teacher_rs:
                        temp_teacher_list.append(rows[0])

                    where = "classNumber='{}'".format(item_dict["temporaryClassNumber"])
                    tempInfo_rs = self.DB.SELECT("*", "temptraining", where, fetchone=True)

                    item_dict["startDate_temp"] = tempInfo_rs[1].strftime("%Y 년  %m 월  %d 일")
                    item_dict["endDate_temp"] = tempInfo_rs[2].strftime("%Y 년  %m 월  %d 일")
                    item_dict["awardDate"] = tempInfo_rs[3].strftime("%Y 년   %m 월    %d 일")

                    # 이름
                    string = item_dict["name"]
                    self.ws.cell(row=7, column=2).value = string

                    # 생년월일
                    DOB = item_dict["RRN"][:6]
                    string = DOB[:2] + ". " + DOB[2:4] + ". " + DOB[4:]
                    self.ws.cell(row=7, column=3).value = string

                    # 연락처
                    string = item_dict["phoneNumber"]
                    self.ws.cell(row=7, column=4).value = string

                    # 교육기관명
                    string = "남양노아요양보호사교육원"
                    self.ws.cell(row=7, column=5).value = string

                    # 교육과정명
                    string = " 요양보호사 {}".format(item_dict["classNumber"])
                    self.ws.cell(row=7, column=7).value = string

                    # 강사
                    teacher_start_row = 12
                    for teacher in temp_teacher_list:
                        for i in range(1, 7):
                            string = teacher_dict[teacher][i]
                            self.ws.cell(row=teacher_start_row, column=i + 1).value = string    
                        teacher_start_row += 1                    

                    # 대체실습 기간
                    string = "{}  ~  {}".format(item_dict["startDate_temp"], item_dict["endDate_temp"])
                    self.ws.cell(row=21, column=3).value = string

                    # 대체실습 시간
                    string = "  총     {}  시간".format(item_dict["trainingCreditHour"])
                    self.ws.cell(row=22, column=3).value = string

                    # 합격여부 
                    self.ws.cell(row=23, column=3).value = "합격"

                    # 자체시험 점수
                    temp_score = random.randint(85, 100)
                    self.ws.cell(row=23, column=6).value = temp_score

                    # 비고 

                    # 서명

                    # 수여일
                    string = "                                      {}".format(item_dict["awardDate"])
                    self.ws.cell(row=27, column=1).value = string

                    self.wb.save(save_path + "\\{}_{}.xlsx".format(item_dict["name"], doc_type))
                    self.logger.info("$Automation [Document|대체실습확인서][{}{} {}]작성".format(item_dict["classNumber"], item_dict["classTime"], item_dict["name"]))
                    self.wb.close()

                return_str = "입력 오류: "
                if valueErrorList == []:
                    return_str += "모두 정상 처리되었습니다."
                else:
                    return_str += ", ".join(valueErrorList)
                    self.logger.error("!Automation [Document|대체실습확인서] 미처리 항목: {}".format(return_str))

                return return_str

            except:
                return traceback.format_exc()

        elif doc_type == "요양보호사 자격증 발급,재발급 신청서":
            try:
                item_dict = {}
                exam_dict = {}

                where = "round={}".format(exam)
                exam_rs = self.DB.SELECT("*", "exam", where, fetchone=True)

                exam_dict["examDate"] = exam_rs[4].strftime("%Y년 %m월 %d일")
                exam_dict["passDate"] = exam_rs[5].strftime("%Y년 %m월 %d일")
                exam_dict["submitDate"] = exam_rs[6].strftime("     %Y  년     %m  월    %d   일    ")

                where = "exam={}".format(exam)
                user_rs = self.inputChecker(self.DB.SELECT("id, name, RRN, phoneNumber, address, classNumber, classTime, temporaryClassNumber", "user", where))
                valueErrorList = self.nullValueChecker(user_rs, 1)

                user_query_list = ["id", "name", "RRN", "phoneNumber", "address", "classNumber", "classTime", "temporaryClassNumber"]

                for rows in user_rs:
                    item_dict.clear()
                    for index in range(len(rows)):
                        item_dict[user_query_list[index]] = rows[index]

                    save_path = self.basePath + "{}\\{}{}\\{}".format(item_dict["classNumber"], item_dict["classNumber"], item_dict["classTime"], item_dict["name"])

                    self.wb = load_workbook(file_path)
                    self.ws = self.wb.active

                    where = "classNumber = '{}' and classTime = '{}'".format(item_dict["classNumber"], item_dict["classTime"])
                    classInfo_rs = self.DB.SELECT("*", "lecture", where, fetchone=True)

                    item_dict["startDate"] = classInfo_rs[2].strftime("%Y.%m.%d")
                    item_dict["endDate"] = classInfo_rs[3].strftime("%Y.%m.%d")

                    where = "classNumber='{}'".format(item_dict["temporaryClassNumber"])
                    tempInfo_rs = self.DB.SELECT("*", "temptraining", where, fetchone=True)

                    item_dict["startDate_temp"] = tempInfo_rs[1].strftime("%Y.%m.%d")
                    item_dict["endDate_temp"] = tempInfo_rs[2].strftime("%Y.%m.%d")
                    item_dict["awardDate"] = tempInfo_rs[3].strftime("%Y 년    %m 월     %d 일")
                    
                    """
                    시험 회차, 시험 시행일, 합격일, 신청 일자, 합격 예정일(?) 등 컬럼 맞춰서 table 생성하기!
                    """

                    # 사진
                    # D:\남양노아요양보호사교육원\교육생관리\7기\7기주간0503\7. 이윤옥
                    ## 사진 이름을 어떻게 짛을 지가 관건!
                    try:
                        student_picture = save_path + "\\{}.jpg".format(item_dict["name"])
                        img = Image(student_picture)
                    except:
                        student_picture = save_path + "\\{}{}_{}.jpg".format(item_dict["classNumber"], item_dict["classTime"], item_dict["name"])
                        img = Image(student_picture)
                    img.height = 142
                    img.width = 111
                    img.anchor = "G6"
                    self.ws.add_image(img)

                    # 이름
                    string = "성명(한자)   {}".format(item_dict["name"])
                    self.ws.cell(row=6, column=2).value = string

                    # 주민등록번호
                    string = "주민등록번호  {}".format(item_dict["RRN"])
                    self.ws.cell(row=7, column=2).value = string

                    # 주소
                    string = "주소   {}".format(item_dict["address"])
                    self.ws.cell(row=8, column=2).value = string

                    # 전화번호
                    string = "전화번호  {}".format(item_dict["phoneNumber"])
                    self.ws.cell(row=9, column=2).value = string

                    # 요양보호사 교육기간. 부터
                    string = item_dict["startDate"][2:]
                    self.ws.cell(row=12, column=2).value = string

                    # 요양보호사 교육기간. 까지
                    string = item_dict["endDate"][2:]
                    self.ws.cell(row=12, column=3).value = string

                    # 교육과정명
                    string = "요양보호사 {} (이론,실기)".format(item_dict["classNumber"])
                    self.ws.cell(row=12, column=4).value = string

                    # 교육기관명
                    string = "남양노아요양보호사교육원"
                    self.ws.cell(row=12, column=7).value = string

                    # 요양보호사 교육기간(실습). 부터
                    string = item_dict["startDate_temp"][2:]
                    self.ws.cell(row=13, column=2).value = string

                    # 요양보호사 교육기간(실습). 까지
                    string = item_dict["endDate_temp"][2:]
                    self.ws.cell(row=13, column=3).value = string

                    # 교육과정명(실습)
                    string = "요양보호사 (대체실습{})".format(item_dict["temporaryClassNumber"])
                    self.ws.cell(row=13, column=4).value = string

                    # 교육기관명(실습)
                    string = "남양노아요양보호사교육원"
                    self.ws.cell(row=13, column=7).value = string

                    # 시험 시행일
                    string = "시험시행일   {}".format(exam_dict["examDate"])
                    self.ws.cell(row=14, column=2).value = string

                    # 시험 합격일
                    string = "시험합격일   {}".format(exam_dict["passDate"])
                    self.ws.cell(row=14, column=5).value = string

                    # 신청 일자
                    string = exam_dict["submitDate"]
                    self.ws.cell(row=19, column=1).value = string

                    # 이름 / shift 는 keyDown(or Up) 에서 left 와 right 를 모두 입력해 주어야 정상작동 함 !!
                    string = "{} (서명 또는 인)".format(item_dict["name"])
                    self.ws.cell(row=20, column=4).value = string

                    self.wb.save(save_path + "\\{}_{}.xlsx".format(item_dict["name"], doc_type))
                    self.logger.info("$Automation [Document|요양보호사 자격증 발급,재발급 신청서][{}{} {}]작성".format(item_dict["classNumber"], item_dict["classTime"], item_dict["name"]))
                    self.wb.close()

                return_str = "입력 오류: "
                if valueErrorList == []:
                    return_str += "모두 정상 처리되었습니다."
                else:
                    return_str += ", ".join(valueErrorList)
                    self.logger.error("!Automation [Document|요양보호사 자격증 발급,재발급 신청서] 미처리 항목: {}".format(return_str))

                return return_str

            except:
                return traceback.format_exc()

    def report(self, doc_type, number, time=None, personal_dcit=None):
        """
        개강보고
        출석부
        대체실습 실시보고
        대체실습 수료보고
        """

        try:
            if time == "":
                time == None

            if doc_type == "개강보고":
                self.wb_imsi = Workbook()
                self.ws_imsi = self.wb_imsi.active
                rs = self.inputChecker(self.DB.SELECT("license, name, RRN, address, phoneNumber", "user", "classNumber='{}' and classTime='{}'".format(number, time), orderBy="FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"))
                for indexX, rows in enumerate(rs, start=1):
                    license = rows[0]
                    if license == "일반":
                        category = "신규\n(일반)"
                    else:
                        category = "자격증\n"
                        if license == "간호사" or license == "물리치료사":
                            category += "({})".format(license[:2])
                        else:
                            category += "({}{})".format(license[0], license[2])

                    name = rows[1]
                    DOB = rows[2][:6]
                    DOB = DOB[:2] + ". " + DOB[2:4] + ". " + DOB[4:]
                    address = rows[3]
                    phone = rows[4]

                    data_list = [category, name, DOB, address, phone]

                    self.ws_imsi.cell(row=indexX, column=1).value = indexX
                    self.ws_imsi.cell(row=indexX, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    self.ws_imsi.cell(row=indexX, column=1).font = Font(size=10)
                    for indexY, value in enumerate(data_list, start=2):
                        self.ws_imsi.cell(row=indexX, column=indexY).value = value
                        self.ws_imsi.cell(row=indexX, column=indexY).alignment = Alignment(horizontal="center", vertical="center")
                        self.ws_imsi.cell(row=indexX, column=indexY).font = Font(size=10)

            elif doc_type == "출석부":
                self.wb_imsi = Workbook()
                self.ws_imsi = self.wb_imsi.active
                rs = self.inputChecker(self.DB.SELECT("name, RRN", "user", "classNumber='{}' and classTime='{}'".format(number, time)))
                for idx, rows in enumerate(rs, start=1):
                    self.ws_imsi.cell(row=idx, column=1).value = idx
                    self.ws_imsi.cell(row=idx, column=2).value = rows[0]
                    DOB = rows[1][:6]
                    DOB = DOB[:2] + ". " + DOB[2:4] + ". " + DOB[4:]
                    self.ws_imsi.cell(row=idx, column=3).value = DOB

            elif doc_type == "대체실습 실시보고":
                self.wb_imsi = Workbook()
                self.ws_imsi = self.wb_imsi.active
                cnt = 0
                rs = self.inputChecker(self.DB.SELECT("license, name, RRN, phoneNumber, classTime, classNumber", "user", "temporaryClassNumber='{}'".format(number), orderBy="classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"))
                for indexX, rows in enumerate(rs, start=1):
                    license = rows[0]
                    if license == "일반":
                        category = "신규"
                        time = 80
                    else:
                        category = "자격증"
                        time = 8

                    name = rows[1]
                    DOB = rows[2][:6]
                    DOB = DOB[:2] + ". " + DOB[2:4] + ". " + DOB[4:]
                    phone = rows[3]
                    class_info = rows[4] + "반 " + rows[5]
                    rs_lecture = self.DB.SELECT("endDate", "lecture", "classNumber='{}' and classTime='{}'".format(rows[5], rows[4]))
                    class_end_date = rs_lecture[0][0].strftime("%Y.%m.%d")
                    time = str(time) + "시간"

                    data_list = [category, name, DOB, phone, class_info, class_end_date, time]

                    self.ws_imsi.cell(row=indexX, column=1).value = indexX
                    self.ws_imsi.cell(row=indexX, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    self.ws_imsi.cell(row=indexX, column=1).font = Font(size=10)
                    for indexY, value in enumerate(data_list, start=2):
                        self.ws_imsi.cell(row=indexX, column=indexY).value = value
                        self.ws_imsi.cell(row=indexX, column=indexY).alignment = Alignment(horizontal="center", vertical="center")
                        self.ws_imsi.cell(row=indexX, column=indexY).font = Font(size=10)

                    cnt = indexX
                cnt += 1
                # if personal_dcit != None:
                #     for items in personal_dcit.values():
                #         where = "id={} and name='{}'".format(items[0], items[1])
                #         rs = self.DB.SELECT("license, name, RRN, phoneNumber, classTime, classNumber", "user", where=where)



            elif doc_type == "대체실습 수료보고":
                self.wb_imsi = Workbook()
                self.ws_imsi = self.wb_imsi.active
                rs = self.inputChecker(self.DB.SELECT("classTime, classNumber, totalCreditHour, theoryCreditHour, practicalCreditHour, trainingCreditHour, name, RRN, address, phoneNumber", "user", "temporaryClassNumber='{}'".format(number), orderBy="classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"))
                rs_temp_lecture = self.DB.SELECT("endDate", "temptraining", "classNumber='{}'".format(number))
                endDate = rs_temp_lecture[0][0].strftime("%Y.%m.%d")
                for indexX, rows in enumerate(rs, start=1):
                    class_info = rows[0] + "반\n" + rows[1]

                    total_time = rows[2]
                    theory_time = rows[3]
                    practice_time = rows[4]
                    training_time = rows[5]
                    name = rows[6]
                    RRN = rows[7][:6] + "\n" + rows[7][6:]
                    address = rows[8]
                    phone = rows[9]

                    data_list = [class_info, total_time, theory_time, practice_time, "", training_time, name, RRN, address, phone, endDate]

                    self.ws_imsi.cell(row=indexX, column=1).value = indexX
                    self.ws_imsi.cell(row=indexX, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    self.ws_imsi.cell(row=indexX, column=1).font = Font(size=10)
                    for indexY, value in enumerate(data_list, start=2):
                        self.ws_imsi.cell(row=indexX, column=indexY).value = value
                        self.ws_imsi.cell(row=indexX, column=indexY).alignment = Alignment(horizontal="center", vertical="center")
                        self.ws_imsi.cell(row=indexX, column=indexY).font = Font(size=10)

            self.logger.info("$Automation [Report|{}]작성".format(doc_type))
            self.wb_imsi.save(self.imsi_workbook_path)
            self.wb_imsi.close()
            os.system(self.imsi_workbook_path)

            return "정상 처리"

        except:
            self.logger.error("!Automation [Report|{}]작성 에러 발생")
            return traceback.format_exc()


    def examPassList(self, exam):
        # NO	합격번호	시험시행기관	시험시행일	시험합격일	교육이수일자	교육시작일자	교육마감일자	대상구분	교육과정명	총교육시간	이론	실기	실습	자격/면허 취득 정보	자격면허코드	자격면허 번호	교부기관	교부기관코드	교부일자	주민등록번호	성명	주소	등록기준지(본적)	전화번호	핸드폰번호
        # 1     2           3             4          5          6              7              8              9          10         11          12     13      14      15                16             17             18         19             20         21             22      23     24                 25          26
        try:
            pass_list_path = self.docFilePath + "화성시-남양노아요양보호사교육원-00회합격자명단_작성용.xlsx"
            wb_pass = load_workbook(pass_list_path)
            ws_pass = wb_pass.active
            if not os.path.exists("D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용"):
                os.makedirs("D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용")

            if not os.path.exists("D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용\\화성시-남양노아요양보호사교육원-{}회합격자명단_제출용.xls".format(exam, exam)):
                shutil.copy("D:\\Master\\files\\화성시-남양노아요양보호사교육원-00회합격자명단_제출용.xls", "D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용\\화성시-남양노아요양보호사교육원-{}회합격자명단_제출용.xls".format(exam, exam))
            
            save_path = "D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용\\화성시-남양노아요양보호사교육원-{}회합격자명단_작성용.xlsx".format(exam, exam)

            where = "round={}".format(exam)
            exam_rs = self.DB.SELECT("*", "exam", where, fetchone=True)

            exam_dict = {}
            exam_dict["round"] = exam_rs[0]
            exam_dict["examDate"] = str(exam_rs[4]).replace("-", "")
            exam_dict["passDate"] = str(exam_rs[5]).replace("-", "")

            where = "exam={}".format(exam)
            user_rs = self.inputChecker(self.DB.SELECT("name, RRN, phoneNumber, license, address, originAddress, classNumber, classTime, totalCreditHour, theoryCreditHour, practicalCreditHour, trainingCreditHour, temporaryClassNumber, exam", "user", where))
            valueErrorList = self.nullValueChecker(user_rs, 0)

            user_query_list = ["name", "RRN", "phoneNumber", "license", "address", "originAddress", "classNumber", "classTime", "totalCreditHour", "theoryCreditHour", "practicalCreditHour", "trainingCreditHour", "temporaryClassNumber", "exam"]
            member_dict = {}

            for idx, rows in enumerate(user_rs, start=1):
                member_dict.clear()
                for index in range(len(rows)):
                    member_dict[user_query_list[index]] = rows[index]

                ws_pass.cell(row=idx + 4, column=1).value = idx
                ws_pass.cell(row=idx + 4, column=3).value = "한국의료보험인국가시험원"
                ws_pass.cell(row=idx + 4, column=4).value = exam_dict["examDate"]
                ws_pass.cell(row=idx + 4, column=5).value = exam_dict["passDate"]

                class_rs = self.DB.SELECT("endDate, startDate", "lecture", "classNumber='{}' and classTime='{}'".format(member_dict["classNumber"], member_dict["classTime"]), fetchone=True)
                ws_pass.cell(row=idx + 4, column=6).value = str(class_rs[0]).replace("-", "")
                ws_pass.cell(row=idx + 4, column=7).value = str(class_rs[1]).replace("-", "")

                temp_class_rs = self.DB.SELECT("endDate", "temptraining", "classNumber='{}'".format(member_dict["temporaryClassNumber"]), fetchone=True)
                ws_pass.cell(row=idx + 4, column=8).value = str(temp_class_rs[0]).replace("-", "")

                string_license = "일반교육 과정"
                license_code = ""
                ws_pass.cell(row=idx + 4, column=9).font = Font(name="맑은 고딕", size=12)
                if member_dict["license"] != "일반":
                    ws_pass.cell(row=idx + 4, column=9).font = Font(name="맑은 고딕", size=8)
                    string_license = "자격/면허 소지자 과정"
                    if member_dict["license"] == "사회복지사":
                        license_code = "25811"

                    elif member_dict["license"] == "간호조무사":
                        license_code = "24260"

                    elif member_dict["license"] == "물리치료사":
                        license_code = "24135"

                    elif member_dict["license"] == "작업치료사":
                        license_code = "24120"
                        
                    elif member_dict["license"] == "간호사":
                        license_code = "24060"

                    elif member_dict["license"] == "경력자":
                        string_license = "경력자 과정"
                        ws_pass.cell(row=idx + 4, column=9).font = Font(name="맑은 고딕", size=12)
                
                ws_pass.cell(row=idx + 4, column=9).value = string_license
                ws_pass.cell(row=idx + 4, column=10).value = "{}반 {}".format(member_dict["classTime"], member_dict["classNumber"])
                ws_pass.cell(row=idx + 4, column=11).value = member_dict["totalCreditHour"]
                ws_pass.cell(row=idx + 4, column=12).value = member_dict["theoryCreditHour"]
                ws_pass.cell(row=idx + 4, column=13).value = member_dict["practicalCreditHour"]
                ws_pass.cell(row=idx + 4, column=14).value = member_dict["trainingCreditHour"]

                if member_dict["license"] != "일반" or member_dict["license"] != "경력자":
                    ws_pass.cell(row=idx + 4, column=15).value = member_dict["license"]
                    ws_pass.cell(row=idx + 4, column=16).value = license_code

                nation = "내국인"
                if member_dict["originAddress"] == "외국인":
                    nation = "외국인"
                ws_pass.cell(row=idx + 4, column=32).value = nation
                ws_pass.cell(row=idx + 4, column=33).value = member_dict["RRN"].replace("-", "")
                ws_pass.cell(row=idx + 4, column=34).value = member_dict["name"]
                ws_pass.cell(row=idx + 4, column=35).value = member_dict["address"]
                ws_pass.cell(row=idx + 4, column=36).value = member_dict["originAddress"]

                ws_pass.cell(row=idx + 4, column=38).value = member_dict["phoneNumber"]

            wb_pass.save(save_path)
            self.logger.info("$Automation [{}회 합격자명단 작성]".format(exam))
            wb_pass.close()

            return_str = "입력 오류: "
            if valueErrorList == []:
                return_str += "모두 정상 처리되었습니다.\n" + save_path
            else:
                return_str += ", ".join(valueErrorList) + "\n" + save_path

            return return_str
        
        except:
            self.logger.error("!Automation[{}회 합격자명단 작성] 작성 중 오류발생".format(exam))
            return traceback.format_exc()

    def paymentList(self, class_number, class_time):
        try:
            self.wb_imsi = load_workbook("D:\\Master\\files\\00기0간_수강료 납부 대장.xlsx")
            self.ws_imsi = self.wb_imsi.active

            self.ws_imsi.cell(row=2, column=2).value = "{} {}".format(class_number, class_time)

            rs = self.inputChecker(self.DB.SELECT("name, RRN, phoneNumber", "user", "classNumber='{}' and classTime='{}'".format(class_number, class_time)))
            for idx, rows in enumerate(rs, start=5):
                self.ws_imsi.cell(row=idx, column=2).value = rows[0]
                DOB = rows[1][:6]
                DOB = DOB[:2] + ". " + DOB[2:4] + ". " + DOB[4:]
                self.ws_imsi.cell(row=idx, column=3).value = DOB
                self.ws_imsi.cell(row=idx, column=4).value = rows[2]

            self.wb_imsi.save(self.basePath + "{}\\{}{}\\{}{}_수강료 납부 대장.xlsx".format(class_number, class_number, class_time, class_number, class_time))
            self.wb_imsi.close()

            if not os.path.exists(self.basePath + "{}\\{}{}\\{}{} 교육기관 수강료 수납대장.hwp".format(class_number, class_number, class_time, class_number, class_time)):
                shutil.copy(r"D:\Master\files\00기0간 교육기관 수강료 수납대장.hwp", self.basePath + "{}\\{}{}\\{}{} 교육기관 수강료 수납대장.hwp".format(class_number, class_number, class_time, class_number, class_time))

            return self.basePath + "{}\\{}{}\\{}{}_수강료 납부 대장.xlsx".format(class_number, class_number, class_time, class_number, class_time)

        except:
            return traceback.format_exc()

    def locker(self, class_number, class_time):
        try:
            if class_time == "주간":
                self.wb_imsi = load_workbook(self.docFilePath + "사물함 주기_주간.xlsx")
            elif class_time == "야간":
                self.wb_imsi = load_workbook(self.docFilePath + "사물함 주기_야간.xlsx")

            self.ws_imsi = self.wb_imsi.active
            rs = self.inputChecker(self.DB.SELECT("name", "user", "classNumber='{}' and classTime='{}'".format(class_number, class_time)))
            rows = 2
            cols = 1
            for r in rs:
                self.ws_imsi.cell(row=rows, column=cols).value = r[0]
                cols += 1
                if cols == 3:
                    cols = 1
                    rows += 2

            self.wb_imsi.save(self.basePath + "{}\\{}{}\\{}{}_사물함 주기.xlsx".format(class_number, class_number, class_time, class_number, class_time))
            self.wb_imsi.close()

            return self.basePath + "{}\\{}{}\\{}{}_사물함 주기.xlsx".format(class_number, class_number, class_time, class_number, class_time)

        except:
            return traceback.format_exc()


    def accountList(self, exam):
        # seq, 이름, id, pw, 주민등록번호(생년월일), 전화번호(연락처), 주소, 가상계좌, 비고(입금완료)
        try:
            self.wb = load_workbook(self.docFilePath + "00회_응시접수명단.xlsx")
            self.ws = self.wb.active
            save_path = "D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용\\{}회 응시접수명단.xlsx".format(exam, exam)

            rs = self.inputChecker(self.DB.SELECT("name, RRN, phoneNumber, address, classNumber, classTime", "user", "exam={}".format(exam), orderBy="classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"))
            pre_class = ""
            class_cnt = 0
            for idx, rows in enumerate(rs, start=3):
                now_class = str(rows[4]) + str(rows[5])
                if pre_class != now_class:
                    pre_class = now_class
                    self.ws.merge_cells(start_row=idx + class_cnt, start_column=1, end_row=idx + class_cnt, end_column=9)
                    self.ws.cell(row=idx + class_cnt, column=1).font = Font(name="맑은 고딕", size=22, bold=True)
                    self.ws.cell(row=idx + class_cnt, column=1).fill = PatternFill(start_color="99CCFF", end_color="99CCFF", fill_type="solid")
                    self.ws.cell(row=idx + class_cnt, column=1).alignment = Alignment(horizontal="center", vertical="center")
                    self.ws.cell(row=idx + class_cnt, column=1).value = pre_class
                    class_cnt += 1
                    
                for i in range(1, 8):
                    if i <= 4:
                        self.ws.cell(row=idx + class_cnt, column=i).font = Font(name="맑은 고딕", size=15, bold=True)
                    else:
                        self.ws.cell(row=idx + class_cnt, column=i).font = Font(name="맑은 고딕", size=15)

                self.ws.cell(row=idx + class_cnt, column=1).value = idx - 2
                self.ws.cell(row=idx + class_cnt, column=1).fill = PatternFill(start_color="FF9999", end_color="FF9999", fill_type="solid")
                self.ws.cell(row=idx + class_cnt, column=2).value = rows[0]
                self.ws.cell(row=idx + class_cnt, column=5).value = rows[1]
                self.ws.cell(row=idx + class_cnt, column=6).value = rows[2]
                self.ws.cell(row=idx + class_cnt, column=7).value = rows[3]
                

            if not os.path.exists("D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용".format(exam)):
                os.makedirs("D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용".format(exam))
            

            self.wb.save(save_path)
            self.wb.close()

            return save_path

        except:
            return traceback.format_exc()

    def printDocument(self, exam, doc_type):
        try:
            non_file_list = []
            rs_user = self.inputChecker(self.DB.SELECT("classNumber, classTime, name", "user", where="exam={}".format(exam), orderBy="classNumber *1, FIELD(classTime, '주간', '야간'), FIELD(license, '일반', '사회복지사', '간호조무사', '간호사', '물리치료사'), id *1"))
            self.logger.info("$Automation [Document|{}][Exam|{}회] 서류 출력 진행".format(doc_type, exam))
            for rows in rs_user:
                doc_path = self.basePath + "{}\\{}{}\\{}\\{}_{}.xlsx".format(rows[0], rows[0], rows[1], rows[2], rows[2], doc_type)
                if not os.path.exists(doc_path):
                    non_file_list.append("{}{} {}".format(rows[0], rows[1], rows[2]))

                else:
                    os.startfile(doc_path, "print")

            if not(non_file_list == []):
                self.logger.error("!Automation [Document|{}][Exam|{}회] 서류 누락자 존재: {}".format(doc_type, exam, "/".join(non_file_list)))
                return "파일에러: {}명\n".format(len(non_file_list)) + " / ".join(non_file_list)

            else:
                return "없음"
        except:
            return traceback.format_exc()

    def gatherPictures(self, exam):
        try:
            non_file_list = []
            dir_path = "D:\\남양노아요양보호사교육원\\경기도청\\03_시험준비 및 자격증발급관련\\{}회_제출용\\자격증 사진".format(exam)
            if not os.path.exists(dir_path):
                os.makedirs(dir_path)
            rs_user = self.inputChecker(self.DB.SELECT("classNumber, classTime, name, RRN", "user", where="exam={}".format(exam)))
            for rows in rs_user:
                picture_path = self.basePath + "{}\\{}{}\\{}\\{}{}_{}.jpg".format(rows[0], rows[0], rows[1], rows[2], rows[0], rows[1], rows[2])
                re_name = rows[3].replace("-", "")
                try:
                    shutil.copy(picture_path, dir_path + "\\{}.jpg".format(re_name))
                except:
                    picture_path = self.basePath + "{}\\{}{}\\{}\\{}{}_{}.JPG".format(rows[0], rows[0], rows[1], rows[2], rows[0], rows[1], rows[2])
                    try:
                        shutil.copy(picture_path, dir_path + "\\{}.jpg".format(re_name))
                    except:
                        non_file_list.append("{}{} {}".format(rows[0], rows[1], rows[2]))


            if not(non_file_list == []):
                return dir_path, " / ".join(non_file_list)

            else:
                return dir_path, "없음"

        except:
            return traceback.format_exc(), "ERROR"







if __name__ == '__main__':
    a = Automation()
    # a.makeDocument(38, "대체실습확인서")