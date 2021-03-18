from selenium import webdriver
from selenium.webdriver.common.keys import Keys

from openpyxl import load_workbook
import pyperclip

import pyautogui
import keyboard

import time
import sys
from random import randint

import operate_data

pyautogui.PAUSE = 0.25

wb = wb_members = load_workbook(operate_data.ac_path["명단총정리"])
ws = wb.active

# 계정이 존재하지 않음
error_login = []
# 원서를 접수하지 않음
error_apply = []

browser = webdriver.Chrome('./Python Workspace/chromedriver.exe')

browser.maximize_window()

browser.get('https://www.kuksiwon.or.kr')

try:
    if browser.find_element_by_xpath('/html/body/div[1]/div[2]/a'):
        elem_home = browser.find_element_by_xpath('/html/body/div[1]/div[2]/a')
        elem_home.click()
except:
    pass

for idx, cell in enumerate(ws["D"], start=1):
    if ws.cell(row=idx, column=2).value == None or idx < 18:
        continue
    if cell.value == None:
        print(ws.cell(row=idx, column=2).value + "선생님 계정이 데이터베이스에 준비되어 있지 않습니다. 다음 선생님으로 넘어갑니다.")
        continue
    if ws.cell(row=idx, column=5).value != None:
        print(ws.cell(row=idx, column=2).value + "선생님은 이미 가상계좌가 입력되어 있습니다. 다음 선생님으로 넘어갑니다.")
        continue
    print("잠시 후에 " + ws.cell(row=idx, column=2).value + "선생님 가상계좌 입력을 시작합니다.")

    time.sleep(2)

    elem_login = browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[1]/a[1]')
    elem_login.click()

    pyperclip.copy(ws.cell(row=idx, column=3).value)

    elem_id = browser.find_element_by_id('headerLoginId')
    elem_id.send_keys(Keys.CONTROL, 'v')

    pyperclip.copy(ws.cell(row=idx, column=4).value)

    elem_pwd = browser.find_element_by_id('headerLoginPwd')
    elem_pwd.send_keys(Keys.CONTROL, 'v')
    elem_login_btn = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div/div[2]/div[1]/form/button')
    elem_login_btn.click()

    browser.implicitly_wait(3)

    try:
        if browser.find_element_by_xpath('/html/body/div[1]/div[2]/a'):
            elem_home = browser.find_element_by_xpath('/html/body/div[1]/div[2]/a')
            elem_home.click()
            print("홈페이지 이동 !!")
    except:
        pass

    try:
        if browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[2]/a[1]'):
            print("Log in 완료 !!")
            elem_mypage = browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[2]/a[1]')
            elem_mypage.click()
    except:
        print("\t\t! ! ! E . R . R . O . R ! ! !")
        print(ws.cell(row=idx, column=2).value + "선생님 계정이 존재하지 않습니다. 다음 순번으로 넘어갑니다.")
        error_login.append(ws.cell(row=idx, column=2).value)
        continue

    elem_mypage = browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[2]/a[1]')
    elem_mypage.click()
    elem_manage = browser.find_element_by_xpath('//*[@id="aside"]/ul/li[2]/a')
    elem_manage.click()

    try:
        if browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[2]/div/div/a/span'):
            elem_exam = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[2]/div/div/a/span')
            elem_exam.click()
    except:
        print(ws.cell(row=idx, column=2).value + "선생님 원서접수가 완료되지 않았습니다. 다음 순번으로 넘어갑니다.")
        error_apply.append(ws.cell(row=idx, column=2).value)
        elem_logout = browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[1]/a[1]')
        elem_logout.click()
        continue

    print("가상계좌가 존재하는지 확인합니다 . . .")

    try:
        if browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[2]/div/div[6]/div[2]/div[2]/span'):
            print("가상계좌가 존재합니다. 입력을 진행합니다.")

    except:
        print("가상계좌가 존재하지 않습니다. 결제하기를 진행한 후 입력을 진행합니다.")
        elem_payment = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[4]/div/a[2]')
        elem_payment.click()
        browser.implicitly_wait(3)
        print("page loading 완료")
        browser.switch_to.window(browser.window_handles[-1])
        elem_pay_exam = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[1]/div[1]/div/a')
        # //*[@id="content"]/div[2]/div[1]/div[1]/div/a
        # //*[@id="content"]/div[2]/div[1]/div[1]/div/a/div[1]/strong
        elem_pay_exam.click()
        elem_vraccount = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[3]/div[1]/div[2]/div[2]/span[2]/label')
        elem_vraccount.click()
        elem_paynow = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[3]/div[2]/div/a[1]')
        elem_paynow.click()

        print("가상계좌 결제를 시작합니다. 정보를 입력하세요.")

        time.sleep(3)

        fw = pyautogui.getActiveWindow()
        print(fw)

        # 전체동의
        pyautogui.click(fw.left + randint(493, 500) + 9, fw.top + randint(203, 210) + 9, duration=0.25)
        # 다음 1
        pyautogui.click(fw.left + randint(25, 580) + 9, fw.top + randint(835, 860) + 9, duration=0.25)
        # 은행선택(농협)
        pyautogui.click(fw.left + randint(168, 179) + 9, fw.top + randint(364, 369) + 9, duration=0.25)
        # 입금자명
        pyautogui.click(fw.left + randint(170, 365) + 9, fw.top + randint(455, 475) + 9, duration=0.25)
        keyboard.wait("enter")
        # 현금영수증:소득공제용
        pyautogui.click(fw.left + randint(170, 178) + 9, fw.top + randint(554, 561) + 9, duration=0.25)
        # 발급번호 종류
        pyautogui.click(fw.left + randint(170, 295) + 9, fw.top + randint(585, 600) + 9, duration=0.25)
        # 발급번호:전화번호
        pyautogui.click(fw.left + randint(170, 295) + 9, fw.top + randint(632, 640) + 9, duration=0.25)
        # 사업자 번호(전화번호) 입력 entry
        pyautogui.click(fw.left + randint(310, 500) + 9, fw.top + randint(585, 600) + 9, duration=0.25)
        keyboard.wait("enter")
        # 구매내역 동의
        pyautogui.click(fw.left + randint(43, 50) + 9, fw.top + randint(809, 816) + 9, duration=0.25)
        # 다음 2
        pyautogui.click(fw.left + randint(25, 580) + 9, fw.top + randint(835, 860) + 9, duration=0.25)
        keyboard.wait("enter")
        time.sleep(1)
        browser.close()
        browser.switch_to.window(browser.window_handles[0])
        browser.refresh()
        pass

    elem_bank = browser.find_element_by_xpath('//*[@id="content"]/div[2]/div[2]/div/div[6]/div[2]/div[2]/span')

    print(elem_bank.text)

    ws.cell(row=idx, column=5).value = elem_bank.text[6:]

    print(ws.cell(row=idx, column=2).value + "선생님 : " + elem_bank.text + "입력을 완료하였습니다.\n파일을 저장합니다.\n")
    wb.save("C:\\Users\\dldud\\Desktop\\info.xlsx")

    browser.switch_to.window(browser.window_handles[0])

    elem_logout = browser.find_element_by_xpath('//*[@id="header"]/div[1]/div/div[1]/a[1]')
    elem_logout.click()


wb.save("C:\\Users\\dldud\\Desktop\\info.xlsx")
print("저장 완료 !")

print(f"에러 발생 인원 : 총 {len(error_login) + len(error_apply)}명")

print("로그인 에러 :", error_login)

print("원서 부재 에러 :", error_apply)

browser.close()
wb.close()

# 전체동의 493,203 / 500,210
# 다음1 25,835 / 580,860
# 농협 168,364 / 179,369
# 입금자명 170,455 / 365,475
# 현금영수증:지출증빙용 269,554 / 275,561
# 현금영수증:소득공제용 170,554 / 178,561
# 발급번호 종류 170,585 / 295,600
# 발급번호:전화번호 170,632 / 295,640
# 사업자 번호(전화번호) 310,585 / 500,600
# 구매 내역 동의 43,809 / 50, 816
# 다음2 25,835 / 580,860