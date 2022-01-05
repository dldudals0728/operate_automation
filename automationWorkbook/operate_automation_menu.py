from ctypes import string_at
import sys
import shutil
# import time : 코드 내에서 time 변수가 다수 사용되어, time.sleep -> pyautogui.sleep 로 사용
import pyautogui
import pyperclip
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
import operate_data

# 대체실습 점수 랜덤지급
from random import randint

# 자격증 발급 필요 서류 자동 출력 : rpa_basic/11_file_system.py
import os
import fnmatch

#                                       example
# preform = automation()
# perform.auto_move_class(4, "야간")
# perform.auto_move_report()
# perform.automation_task_students(3, "주간", "자격증 발급,재발급 신청서")
# perform.automation_task_report(5, "주간", "개강보고") # 수료보고가 따로 없기 때문에 kind = 개강보고로 고정 !
# perform.automation_task_temporary(4, "수료보고")
# perform.(4, "주간", "교육수료증명서.hwp") # 뒤에 복사할 파일을 입력할 때 꼭 !!!!! 확장자 명까지 작성하기 ㅎㅎ
# perform.mkattendance(3, "야간")
# perform.update_attendance(4, "야간") 업데이트를 진행할 기수 + 시간

# 업데이트 목록
# 각 업무자동화 시스템 변수 초기화 하기(이름, 기수, 대체기수 ,,,etc.)
# 수여일 수정

class automation:
    def __init__(self): # slp : sleep / dura : duration / itv : interval / sleep, duration, interval 을 일괄적으로 시간을 정하기 위해 설정한다.
        print("Noa Automation Program")
        # 수행 전, 명단총정리 엑셀 파일을 불러와 자료를 복사할 준비
        self.ac_name = None
        self.db_path = None
        self.wb_members = None
        self.ws_members = None

        self.fw_epr = None

        self.input_database()
        self.init_workbook()

    def input_database(self):
        database_path_folder = "D:\\Master"
        database_path = "D:\\Master\\database.txt"
        if not os.path.exists(database_path_folder):
            os.makedirs(database_path_folder)

        if not os.path.exists(database_path):
            print("database 파일이 존재하지 않아 파일을 생성합니다 ...")
            f = open(database_path, "w", encoding="utf8")
            academy_name = input("요양보호사교육원 이름을 입력해주세요: ")
            file_database = input("데이터베이스(명단총정리)파일의 경로를 입력해주세요\nex)'D:\\Master\\남양노아요양보호사교육원_명단총정리.xlsx': ")
            f.write(f"acdemy_name\t{academy_name}\n")
            f.write(f"database_path\t{file_database}\n")
            f.close()

            check = f.readlines()

        else:
            f = open(database_path, "r", encoding="utf8")
            check = f.readlines()
            if check == []:
                print("database 파일이 작성되지 않아 파일 작성을 시작합니다 ...")
                f = open(database_path, "w", encoding="utf8")
                academy_name = input("요양보호사교육원 이름을 입력해주세요: ")
                file_database = input("데이터베이스(명단총정리)파일의 경로를 입력해주세요\nex)'D:\\Master\\남양노아요양보호사교육원_명단총정리.xlsx': ")
                f.write(f"acdemy_name\t{academy_name}\n")
                f.write(f"database_path\t{file_database}\n")
                f.close()

                check = f.readlines()

            else:
                print("database 파일이 존재합니다")

            self.ac_name = check[0]
            self.ac_name = self.ac_name.split('\t')
            self.ac_name = self.ac_name[1][:-1]
            print("교육원 명:", self.ac_name)

            self.db_path = check[1]
            self.db_path = self.db_path.split('\t')
            self.db_path = self.db_path[1][:-1]
            print("database 경로:", self.db_path)

    def modify_database(self, doc):
        # doc: 0 -> 학원 명 수정
        # doc: 1 -> database path 수정
        database_path = "D:\\Master\\database.txt"
        if doc == 0:
            mod_ac_name = input("변경된 교육원 명을 입력해 주세요:")
            mod_db_path = self.db_path

            
        elif doc == 1:
            mod_db_path = input("변경된 database 경로를 입력해 주세요:")
            mod_ac_name = self.ac_name

        f = open(database_path, "w", encoding="utf8")
        f.write(f"acdemy_name\t{mod_ac_name}\n")
        f.write(f"database_path\t{mod_db_path}\n")
        f.close()

        print("[변경 전]\t교육원 명:", self.ac_name, "\tdatabase path:", self.db_path)

        check = f.readlines()

        self.ac_name = check[0]
        self.ac_name = self.ac_name.split('\t')
        self.ac_name = self.ac_name[1][:-1]

        self.db_path = check[1]
        self.db_path = self.db_path.split('\t')
        self.db_path = self.db_path[1][:-1]

        print("[변경 후]\t교육원 명:", self.ac_name, "\tdatabase path:", self.db_path)
        


    def init_workbook(self):
        self.wb_members = load_workbook(self.db_path)
        self.ws_members = self.wb_members.active

    def automation_task_students(self, ordinal_num, time, task, version):
        # 사용방법 : x.automation_task_students(3, 주간, "교육수료증명서")

        wb_automation =  load_workbook("D:\\Master\\업무자동화.xlsx")
        ws_automation = wb_automation.active

        # excel 파일을 불러오기 위해 경로를 최신화 하기 위한 참조
        wb_automation = load_workbook("D:\\Master\\업무자동화.xlsx")
        ws_automation = wb_automation.active
        string_set = f"{ordinal_num}기{time}"
        switch = 0
        
        if task == "교육수료증명서":
            if version == True:
                i = 1
                for idx, cell in enumerate(self.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if switch == 0:
                        string_set = cell.value
                        switch = 1

                    if self.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간조\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_사복\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간호\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_completion = load_workbook(string_stu)
                    ws_completion = wb_completion.active

                    # 교육수료증명서 호수
                    string = f"    2021  년  제  {self.ws_members.cell(row=idx, column=2).value} 호"
                    ws_completion.cell(row=1, column=1).value = string

                    # 이름
                    string = f" {self.ws_members.cell(row=idx, column=18).value[0]} {self.ws_members.cell(row=idx, column=18).value[1]} {self.ws_members.cell(row=idx, column=18).value[2]}"
                    ws_completion.cell(row=4, column=3).value = string

                    # 주소
                    string = f" {self.ws_members.cell(row=idx, column=21).value}"
                    ws_completion.cell(row=5, column=3).value = string

                    # 주민등록번호
                    string = f" {self.ws_members.cell(row=idx, column=20).value[:6]} - {self.ws_members.cell(row=idx, column=20).value[7:]}"
                    ws_completion.cell(row=6, column=3).value = string

                    # 전화번호
                    string = f"{self.ws_members.cell(row=idx, column=19).value}"
                    ws_completion.cell(row=6, column=6).value = string

                    # 교육과정명
                    string = f" 요양보호사 {self.ws_members.cell(row=idx, column=4).value}"
                    ws_completion.cell(row=7, column=3).value = string

                    # 이론실기 이수기간 / 각 기수별로 기간 선정, 2020 년  11 월  16 일 ∼  21 년  01 월 15 일 형식으로, 끝기간은 년도수 두자리만 표시
                    string = f"{self.ws_members.cell(row=idx, column=6).value[:4]} 년  {self.ws_members.cell(row=idx, column=6).value[5:7]} 월  {self.ws_members.cell(row=idx, column=6).value[8:]} 일 ∼  {self.ws_members.cell(row=idx, column=7).value[2:4]} 년  {self.ws_members.cell(row=idx, column=7).value[5:7]} 월 {self.ws_members.cell(row=idx, column=7).value[8:]} 일 "
                    ws_completion.cell(row=9, column=3).value = string

                    # 이론실기 이수시간
                    string = f"        {str(int(self.ws_members.cell(row=idx, column=12).value) + int(self.ws_members.cell(row=idx, column=13).value))}  시간"
                    ws_completion.cell(row=9, column=7).value = string

                    # 실습기간 / 대체실습 각 기수별 or 실습기간 따로 만들기,  21년 01월 18일 ∼ 21년 03월 13일 형식으로, 년도수 두자리만 표시
                    string = f" {self.ws_members.cell(row=idx, column=9).value[2:4]}년 {self.ws_members.cell(row=idx, column=9).value[5:7]}월 {self.ws_members.cell(row=idx, column=9).value[8:]}일 ∼ {self.ws_members.cell(row=idx, column=10).value[2:4]}년 {self.ws_members.cell(row=idx, column=10).value[5:7]}월 {self.ws_members.cell(row=idx, column=10).value[8:]}일"
                    ws_completion.cell(row=12, column=4).value = string

                    # 대체실습이 종료되면, 각 사람마다 실습시간(각 기관) 이 달라짐. 업데이트 필요
                    # 실습시간
                    string = f"        {self.ws_members.cell(row=idx, column=14).value}  시간"

                    ws_completion.cell(row=12, column=7).value = string

                    # 총 실습시간
                    string = f"         {self.ws_members.cell(row=idx, column=14).value}  시간"
                    ws_completion.cell(row=18, column=7).value = string

                    # 총 이수시간
                    string = f"       {self.ws_members.cell(row=idx, column=11).value}  시간"
                    ws_completion.cell(row=19, column=7).value = string

                    # 수여일 / 각 인원 대체실습 기준 종료일 바로 다음 월요일 날짜로 지정
                    if "대체실습" in self.ws_members.cell(row=idx, column=8).value:
                        temp_string = self.ws_members.cell(row=idx, column=8).value
                        if len(temp_string) == 7:
                            gisu = int(self.ws_members.cell(row=idx, column=8).value[5])
                        elif len(temp_string) == 8:
                            gisu = int(self.ws_members.cell(row=idx, column=8).value[5:7])
                        if gisu <= 8:
                            string = f"                                      {ws_automation.cell(row=gisu + 2, column=3).value[:4]} 년   {ws_automation.cell(row=gisu + 2, column=3).value[5:7]} 월    {ws_automation.cell(row=gisu + 2, column=3).value[8:]} 일"
                        elif gisu <= 16:
                            string = f"                                      {ws_automation.cell(row=gisu - 6, column=5).value[:4]} 년   {ws_automation.cell(row=gisu - 6, column=5).value[5:7]} 월    {ws_automation.cell(row=gisu - 6, column=5).value[8:]} 일"
                        else:
                            string = f"                                      {ws_automation.cell(row=gisu - 14, column=7).value[:4]} 년   {ws_automation.cell(row=gisu - 14, column=7).value[5:7]} 월    {ws_automation.cell(row=gisu - 14, column=7).value[8:]} 일"
                        # if "1기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=3, column=3).value[:4]} 년    {ws_automation.cell(row=3, column=3).value[5:7]} 월     {ws_automation.cell(row=3, column=3).value[8:]} 일"
                        # elif "2기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=4, column=3).value[:4]} 년    {ws_automation.cell(row=4, column=3).value[5:7]} 월     {ws_automation.cell(row=4, column=3).value[8:]} 일"
                        # elif "3기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=5, column=3).value[:4]} 년    {ws_automation.cell(row=5, column=3).value[5:7]} 월     {ws_automation.cell(row=5, column=3).value[8:]} 일"
                        # elif "4기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=6, column=3).value[:4]} 년    {ws_automation.cell(row=6, column=3).value[5:7]} 월     {ws_automation.cell(row=6, column=3).value[8:]} 일"
                        # elif "5기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=7, column=3).value[:4]} 년    {ws_automation.cell(row=7, column=3).value[5:7]} 월     {ws_automation.cell(row=7, column=3).value[8:]} 일"
                        # elif "6기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=8, column=3).value[:4]} 년    {ws_automation.cell(row=8, column=3).value[5:7]} 월     {ws_automation.cell(row=8, column=3).value[8:]} 일"
                        # elif "7기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                               {ws_automation.cell(row=9, column=3).value[:4]} 년    {ws_automation.cell(row=9, column=3).value[5:7]} 월     {ws_automation.cell(row=9, column=3).value[8:]} 일"
                        # elif "8기" in self.ws_members.cell(row=idx, column=8).value:
                            # string = f"                               {ws_automation.cell(row=10, column=3).value[:4]} 년    {ws_automation.cell(row=10, column=3).value[5:7]} 월     {ws_automation.cell(row=10, column=3).value[8:]} 일"
                    ws_completion.cell(row=23, column=1).value = string

                    wb_completion.save(string_stu)
                    wb_completion.close()

                    i += 1

            elif version == False:
                print("한글 버전이 삭제되었습니다. 관리자에게 문의해주세요")
                
        elif task == "대체실습확인서":
            wb_temp_score = load_workbook("D:\\Master\\대체실습_점수.xlsx")
            ws_temp_score = wb_temp_score.active
            switch = 0
            if version == True:
                i = 1
                for idx, cell in enumerate(self.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    if switch == 0:
                        string_set = cell.value
                        switch = 1
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if self.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간조\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_사복\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간호\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_temp = load_workbook(string_stu)
                    ws_temp = wb_temp.active

                    # 대체실습확인서
                    # 이름
                    string = self.ws_members.cell(row=idx, column=18).value
                    ws_temp.cell(row=7, column=2).value = string

                    # 생년월일
                    string = f"{self.ws_members.cell(row=idx, column=20).value[:2]}. {self.ws_members.cell(row=idx, column=20).value[2:4]}. {self.ws_members.cell(row=idx, column=20).value[4:6]}"
                    ws_temp.cell(row=7, column=3).value = string

                    # 연락처
                    string = self.ws_members.cell(row=idx, column=19).value
                    ws_temp.cell(row=7, column=4).value = string

                    # 교육기관명
                    string = self.ws_members.cell(row=idx, column=3).value
                    ws_temp.cell(row=7, column=5).value = string

                    # 교육과정명
                    string = f" 요양보호사 {self.ws_members.cell(row=idx, column=4).value}"
                    ws_temp.cell(row=7, column=7).value = string

                    # 강사
                    if "대체실습" in self.ws_members.cell(row=idx, column=8).value:
                        if "1기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=17, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string

                        elif "2기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                        elif "3기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string
                        
                        elif "4기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string

                        elif "5기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string

                        elif "6기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string
                        
                        elif "7기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string
                        
                        elif "8기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                        elif "9기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=13, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string
                        elif "10기" in self.ws_members.cell(row=idx, column=8).value:
                            for j in range(1, 7):
                                string = ws_automation.cell(row=14, column=j).value
                                ws_temp.cell(row=12, column=j + 1).value = string
                            
                            for j in range(1, 7):
                                string = ws_automation.cell(row=15, column=j).value
                                ws_temp.cell(row=13, column=j + 1).value = string

                            for j in range(1, 7):
                                string = ws_automation.cell(row=16, column=j).value
                                ws_temp.cell(row=14, column=j + 1).value = string

                        elif "11기" in ws_automation.cell(row=idx, column=8).value:
                            pass

                        elif "12기" in ws_automation.cell(row=idx, column=8).value:
                            pass

                        elif "13기" in ws_automation.cell(row=idx, column=8).value:
                            pass

                        elif "14기" in ws_automation.cell(row=idx, column=8).value:
                            pass

                        elif "15기" in ws_automation.cell(row=idx, column=8).value:
                            pass
                        
                        elif "16기" in ws_automation.cell(row=idx, column=8).value:
                            pass

                    # 대체실습 기간
                    string = f"{self.ws_members.cell(row=idx, column = 9).value[:4]} 년  {self.ws_members.cell(row=idx, column = 9).value[5:7]} 월  {self.ws_members.cell(row=idx, column = 9).value[8:]} 일  ∼    {self.ws_members.cell(row=idx, column = 10).value[:4]} 년  {self.ws_members.cell(row=idx, column = 10).value[5:7]} 월  {self.ws_members.cell(row=idx, column = 10).value[8:]} 일"
                    ws_temp.cell(row=20, column=3).value = string

                    # 대체실습 시간
                    string = f"  총     {self.ws_members.cell(row=idx, column=14).value}  시간"
                    ws_temp.cell(row=21, column=3).value = string

                    # 합격여부 
                    ws_temp.cell(row=22, column=3).value = "합격"

                    # 자체시험 점수
                    name = self.ws_members.cell(row=idx, column=18).value
                    for cell in ws_temp_score["C"]:
                        if cell.value == name:
                            temp_row = cell.row
                    temp_score = ws_temp_score.cell(row=temp_row, column=7).value
                    if temp_score == None:
                        temp_score = randint(90, 100)
                    else:
                        pass
                    ws_temp.cell(row=22, column=6).value = temp_score

                    # 비고 

                    # 서명

                    # 수여일
                    if "대체실습" in self.ws_members.cell(row=idx, column=8).value:
                        temp_string = self.ws_members.cell(row=idx, column=8).value
                        if len(temp_string) == 7:
                            gisu = int(self.ws_members.cell(row=idx, column=8).value[5])
                        elif len(temp_string) == 8:
                            gisu = int(self.ws_members.cell(row=idx, column=8).value[5:7])
                        if gisu <= 8:
                            string = f"                                      {ws_automation.cell(row=gisu + 2, column=3).value[:4]} 년   {ws_automation.cell(row=gisu + 2, column=3).value[5:7]} 월    {ws_automation.cell(row=gisu + 2, column=3).value[8:]} 일"
                        else:
                            string = f"                                      {ws_automation.cell(row=gisu - 6, column=5).value[:4]} 년   {ws_automation.cell(row=gisu - 6, column=5).value[5:7]} 월    {ws_automation.cell(row=gisu - 6, column=5).value[8:]} 일"
                        # if "1기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=3, column=3).value[:4]} 년   {ws_automation.cell(row=3, column=3).value[5:7]} 월    {ws_automation.cell(row=3, column=3).value[8:]} 일"
                        # elif "2기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=4, column=3).value[:4]} 년   {ws_automation.cell(row=4, column=3).value[5:7]} 월    {ws_automation.cell(row=4, column=3).value[8:]} 일"
                        # elif "3기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=5, column=3).value[:4]} 년   {ws_automation.cell(row=5, column=3).value[5:7]} 월    {ws_automation.cell(row=5, column=3).value[8:]} 일"
                        # elif "4기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=6, column=3).value[:4]} 년   {ws_automation.cell(row=6, column=3).value[5:7]} 월    {ws_automation.cell(row=6, column=3).value[8:]} 일"
                        # elif "5기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=7, column=3).value[:4]} 년   {ws_automation.cell(row=7, column=3).value[5:7]} 월    {ws_automation.cell(row=7, column=3).value[8:]} 일"
                        # elif "6기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=8, column=3).value[:4]} 년   {ws_automation.cell(row=8, column=3).value[5:7]} 월    {ws_automation.cell(row=8, column=3).value[8:]} 일"
                        # elif "7기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=9, column=3).value[:4]} 년   {ws_automation.cell(row=9, column=3).value[5:7]} 월    {ws_automation.cell(row=9, column=3).value[8:]} 일"
                        # elif "8기" in self.ws_members.cell(row=idx, column=8).value:
                        #     string = f"                                      {ws_automation.cell(row=10, column=3).value[:4]} 년   {ws_automation.cell(row=10, column=3).value[5:7]} 월    {ws_automation.cell(row=10, column=3).value[8:]} 일"

                    ws_temp.cell(row=27, column=1).value = string

                    wb_temp.save(string_stu)
                    wb_temp.close()
                    i += 1

            elif version == False:
                print("한글 버전이 삭제되었습니다. 관리자에게 문의해주세요")

        elif task == "요양보호사 자격증 발급,재발급 신청서":
            switch = 0
            if version == True:
                i = 1
                for idx, cell in enumerate(self.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    if switch == 0:
                        string_set = cell.value
                        switch = 1
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if self.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간조\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_사복\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간호\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_certificate = load_workbook(string_stu)
                    ws_certificate = wb_certificate.active

                    # 요양보호사 자격증 발급,재발급 신청서
                    # 사진
                    # D:\남양노아요양보호사교육원\교육생관리\7기\7기주간0503\7. 이윤옥
                    if len(string_set) == 8:
                        class_info = string_set[:4]
                    elif len(string_set) == 9:
                        class_info = string_set[:5]
                    student_picture = string_stu.replace(f"{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx", f"{class_info}_{self.ws_members.cell(row=idx, column=18).value}.jpg")
                    img = Image(student_picture)
                    img.height = 142
                    img.width = 111
                    img.anchor = "G6"
                    ws_certificate.add_image(img)
                    # print(string_stu)
                    # print(os.path.isfile(string_stu))
                    # i += 1
                    # wb_certificate.save(string_stu)
                    # wb_certificate.close()
                    # continue
                    # 이름
                    string = f"성명(한자)   {self.ws_members.cell(row=idx, column=18).value}"
                    ws_certificate.cell(row=6, column=2).value = string

                    # 주민등록번호
                    string = f"주민등록번호  {self.ws_members.cell(row=idx, column=20).value}"
                    ws_certificate.cell(row=7, column=2).value = string

                    # 주소
                    string = f"주소   {self.ws_members.cell(row=idx, column=21).value}"
                    ws_certificate.cell(row=8, column=2).value = string

                    # 전화번호
                    string = f"전화번호  {self.ws_members.cell(row=idx, column=19).value}"
                    ws_certificate.cell(row=9, column=2).value = string

                    # 요양보호사 교육기간. 부터
                    string = f"{self.ws_members.cell(row=idx, column=6).value[2:4]}.{self.ws_members.cell(row=idx, column=6).value[5:7]}.{self.ws_members.cell(row=idx, column=6).value[8:]}"
                    ws_certificate.cell(row=12, column=2).value = string

                    # 요양보호사 교육기간. 까지
                    string = f"{self.ws_members.cell(row=idx, column=7).value[2:4]}.{self.ws_members.cell(row=idx, column=7).value[5:7]}.{self.ws_members.cell(row=idx, column=7).value[8:]}"
                    ws_certificate.cell(row=12, column=3).value = string

                    # 교육과정명
                    string = f"요양보호사 {self.ws_members.cell(row=idx, column=5).value[0]}기 (이론,실기)"
                    ws_certificate.cell(row=12, column=4).value = string

                    # 교육기관명
                    string = operate_data.ac_name
                    ws_certificate.cell(row=12, column=7).value = string

                    # 요양보호사 교육기간(실습). 부터
                    string = f"{self.ws_members.cell(row=idx, column=9).value[2:4]}.{self.ws_members.cell(row=idx, column=9).value[5:7]}.{self.ws_members.cell(row=idx, column=9).value[8:]}"
                    ws_certificate.cell(row=13, column=2).value = string

                    # 요양보호사 교육기간(실습). 까지
                    string = f"{self.ws_members.cell(row=idx, column=10).value[2:4]}.{self.ws_members.cell(row=idx, column=10).value[5:7]}.{self.ws_members.cell(row=idx, column=10).value[8:]}"
                    ws_certificate.cell(row=13, column=3).value = string

                    # 교육과정명(실습)
                    string = f"요양보호사 (대체실습{self.ws_members.cell(row=idx, column=8).value[5]}기)"
                    ws_certificate.cell(row=13, column=4).value = string

                    # 교육기관명(실습)
                    string = operate_data.ac_name
                    ws_certificate.cell(row=13, column=7).value = string

                    # 시험 시행일
                    if "34" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험시행일   2021년 02월 20일"
                    elif "35" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험시행일   2021년 05월 15일"
                    elif "36" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험시행일   2021년 08월 07일"
                    elif "37" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험시행일   2021년 11월 06일"
                    ws_certificate.cell(row=14, column=2).value = string

                    # 시험 합격일
                    if "34" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험합격일   2021년 03월 09일"
                    elif "35" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험합격일   2021년 06월 01일"
                    elif "36" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험합격일   2021년 08월 24일"
                    elif "37" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "시험합격일   2021년 11월 23일"
                    ws_certificate.cell(row=14, column=5).value = string

                    # 신청 일자
                    if "34" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "     2021  년     03  월    15   일    "
                    elif "35" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "     2021  년     06  월    07   일    "
                    elif "36" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "     2021  년     08  월    30   일    "
                    elif "37" in str(self.ws_members.cell(row=idx, column=24).value):
                        string = "     2021  년     11  월    29   일    "
                    ws_certificate.cell(row=19, column=1).value = string

                    # 이름 / shift 는 keyDown(or Up) 에서 left 와 right 를 모두 입력해 주어야 정상작동 함 !!
                    string = f"{self.ws_members.cell(row=idx, column=18).value} (서명 또는 인)"
                    ws_certificate.cell(row=20, column=4).value = string

                    wb_certificate.save(string_stu)
                    wb_certificate.close()
                    i += 1

            elif version == False:
                print("한글 버전이 삭제되었습니다. 관리자에게 문의해주세요")

        self.wb_members.close()

    def automation_task_report(self, ordinal_num, time, kind):
        # 사용방법 x.automation_task_report(3, "주간", "개강보고")
        print("mk_report_open_xlsx 함수로 기능이 변경되었습니다.")
        print("사용 방법 | (7기 주간 개강보고서 작성 시)\nmk_report_open_xlsx(7, 주간) (기수, 반)")

    def automation_task_temporary(self, ordinal_num, kind):
        print("mk_report_open_close_xlsx_temp 함수로 기능이 변경되었습니다.")
        print("사용 방법 | (대체실습 7기 실시보고서 작성 시)\mk_report_open_close_xlsx_temp(7, 0) (기수, class type)\n0: 실시보고 / 1: 수료보고")
    
    def mkattendance(self, ordinal_num, time):
        print("mk_attendance_xlsx 함수로 기능이 변경되었습니다.")
        print("사용 방법 | (대체실습 7기 출석부 명단 작성 시)\mk_attendance_xlsx(7, 2) (기수, class type)\n0: 주간 / 1: 야간 / 2: 대체실습")

    def mk_report_open_xlsx(self, ordinal_num, time):
        if os.path.isfile("D:\\Master\\imsi.xlsx"):
            os.remove("D:\\Master\\imsi.xlsx")
        wb_imsi = Workbook()
        ws_imsi = wb_imsi.active

        string_set = f"{ordinal_num}기{time}"
        i = 2
        ws_imsi.cell(row=1, column=1).value = f"{ordinal_num}기 {time}반 개강보고 명단"

        print("\n\n******************************")
        print(f"{ordinal_num}기 {time} 개강보고 명단을 작성합니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
        print("******************************\n\n")
        for idx, cell in enumerate(self.ws_members["E"], start=1):
            if idx < 5:
                continue
            if string_set not in cell.value:
                continue

            # 순번
            ws_imsi.cell(row=i, column=1).value = self.ws_members.cell(row=idx, column=1).value

            # 과정
            if self.ws_members.cell(row=idx, column=15).value == "일반":
                curriculum = "일반\n(신규)"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                curriculum = "자격증\n(사복)"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                curriculum = "자격증\n(간조)"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                curriculum = "자격증\n(간호)"
            ws_imsi.cell(row=i, column=2).value = curriculum

            # 이름
            ws_imsi.cell(row=i, column=3).value = self.ws_members.cell(row=idx, column=18).value

            # 생년월일
            ws_imsi.cell(row=i, column=4).value = f"{self.ws_members.cell(row=idx, column=20).value[:2]}. {self.ws_members.cell(row=idx, column=20).value[2:4]}. {self.ws_members.cell(row=idx, column=20).value[4:6]}"

            # 주소
            ws_imsi.cell(row=i, column=5).value = self.ws_members.cell(row=idx, column=21).value

            # 연락처
            ws_imsi.cell(row=i, column=6).value = self.ws_members.cell(row=idx, column=19).value


            i += 1

            print(f"{self.ws_members.cell(row=idx, column=1).value}. {ordinal_num}기{time} {self.ws_members.cell(row=idx, column=18).value} 작성 완료.")
            
        print("\n\n******************************")
        print(f"{ordinal_num}기 {time} 개강보고 명단 작성을 완료했습니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
        print("******************************\n\n")
        wb_imsi.save("D:\\Master\\imsi.xlsx")
        wb_imsi.close()

    def mk_report_open_close_xlsx_temp(self, ordinal_num, openClose):
        # openClose: 0 = 개강보고, 1 = 종강보고
        if os.path.isfile("D:\\Master\\imsi.xlsx"):
            os.remove("D:\\Master\\imsi.xlsx")
        wb_imsi = Workbook()
        ws_imsi = wb_imsi.active

        string_set = f"대체실습 {ordinal_num}기"
        i = 2

        if openClose == 0:
            ws_imsi.cell(row=1, column=1).value = f"대체실습 {ordinal_num}기 실시보고 명단"

            print("\n\n******************************")
            print(f"대체실습 {ordinal_num}기 실시보고 명단을 작성합니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
            print("******************************\n\n")

            for idx, cell in enumerate(self.ws_members["H"], start=1):
                if idx < 5:
                    continue
                if cell.value == None:
                    continue
                if string_set not in cell.value:
                    continue

                # 순번
                ws_imsi.cell(row=i, column=1).value = self.ws_members.cell(row=idx, column=1).value

                # 교육구분
                if self.ws_members.cell(row=idx, column=15).value == "일반":
                    curriculum = "(신규)"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                    curriculum = "자격증"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                    curriculum = "자격증"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                    curriculum = "자격증"
                ws_imsi.cell(row=i, column=2).value = curriculum

                # 성명
                ws_imsi.cell(row=i, column=3).value = self.ws_members.cell(row=idx, column=18).value

                # 생년월일
                ws_imsi.cell(row=i, column=4).value = f"{self.ws_members.cell(row=idx, column=20).value[:2]}.{self.ws_members.cell(row=idx, column=20).value[2:4]}.{self.ws_members.cell(row=idx, column=20).value[4:6]}"

                # 연락처
                ws_imsi.cell(row=i, column=5).value = self.ws_members.cell(row=idx, column=19).value

                # 교육기수
                ws_imsi.cell(row=i, column=6).value = f"{self.ws_members.cell(row=idx, column=4).value} {self.ws_members.cell(row=idx, column=5).value[2:4]}반"

                # 이론·실기 교육이수일
                ws_imsi.cell(row=i, column=7).value = f"{self.ws_members.cell(row=idx, column=7).value[:4]}. {self.ws_members.cell(row=idx, column=7).value[5:7]}. {self.ws_members.cell(row=idx, column=7).value[8:]}"

                # 대체실습 필요시간
                if self.ws_members.cell(row=idx, column=15).value == "일반":
                    needs_time = "80시간"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                    needs_time = "8시간"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                    needs_time = "8시간"
                elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                    needs_time = "8시간"
                ws_imsi.cell(row=i, column=8).value = needs_time

                i += 1

                print(f"{self.ws_members.cell(row=idx, column=1).value}. 대체실습 {ordinal_num}기 실시보고 {self.ws_members.cell(row=idx, column=18).value} 작성 완료.")

            print("\n\n******************************")
            print(f"대체실습 {ordinal_num}기 실시보고 명단 작성을 완료했습니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
            print("******************************\n\n")

        elif openClose == 1:
            ws_imsi.cell(row=1, column=1).value = f"대체실습 {ordinal_num}기 수료보고 명단"

            wb_automation = load_workbook("D:\\Master\\업무자동화.xlsx")
            ws_automation = wb_automation.active

            print("\n\n******************************")
            print(f"대체실습 {ordinal_num}기 수료보고 명단을 작성합니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
            print("******************************\n\n")

            for idx, cell in enumerate(self.ws_members["H"], start=1):
                if idx < 5:
                    continue
                if cell.value == None:
                    continue
                if string_set not in cell.value:
                    continue

                # 순번
                ws_imsi.cell(row=i, column=1).value = self.ws_members.cell(row=idx, column=1).value

                # 교육 과정명
                curriculum = f"{self.ws_members.cell(row=idx, column=5).value[2:4]}반\n{self.ws_members.cell(row=idx, column=4).value}"
                ws_imsi.cell(row=i, column=2).value = curriculum

                # 총 시간
                ws_imsi.cell(row=i, column=3).value = self.ws_members.cell(row=idx, column=11).value

                # 이론
                ws_imsi.cell(row=i, column=4).value = self.ws_members.cell(row=idx, column=12).value

                # 실기
                ws_imsi.cell(row=i, column=5).value = self.ws_members.cell(row=idx, column=13).value

                # 실습

                # 대체실습
                ws_imsi.cell(row=i, column=7).value = self.ws_members.cell(row=idx, column=14).value

                # 성명
                ws_imsi.cell(row=i, column=8).value = self.ws_members.cell(row=idx, column=18).value

                # 생년월일
                ws_imsi.cell(row=i, column=9).value = self.ws_members.cell(row=idx, column=20).value

                # 주소(도로명)
                ws_imsi.cell(row=i, column=10).value = self.ws_members.cell(row=idx, column=22).value

                # 연락처
                ws_imsi.cell(row=i, column=11).value = self.ws_members.cell(row=idx, column=19).value

                # 수료 연월일
                gisu = int(ordinal_num)
                if gisu <= 8:
                    complete_date = f"{ws_automation.cell(row=gisu + 2, column=3).value[:4]}.{ws_automation.cell(row=gisu + 2, column=3).value[5:7]}.{ws_automation.cell(row=gisu + 2, column=3).value[8:]}"
                else:
                    complete_date = f"{ws_automation.cell(row=gisu - 6, column=5).value[:4]}.{ws_automation.cell(row=gisu - 6, column=5).value[5:7]}.{ws_automation.cell(row=gisu - 6, column=5).value[8:]}"

                ws_imsi.cell(row=i, column=12).value = complete_date

                # 비고

                i += 1

                print(f"{self.ws_members.cell(row=idx, column=1).value}. 대체실습 {ordinal_num}기 수료보고 {self.ws_members.cell(row=idx, column=18).value} 작성 완료.")

            print("\n\n******************************")
            print(f"대체실습 {ordinal_num}기 수료보고 명단 작성을 완료했습니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
            print("******************************\n\n")
        wb_imsi.save("D:\\Master\\imsi.xlsx")
        wb_imsi.close()

    def mk_attendance_xlsx(self, ordinal_num, class_type):
        # class_type = 0: 주간 / 1: 야간 / 2:대체 실습
        if os.path.isfile("D:\\Master\\imsi.xlsx"):
            os.remove("D:\\Master\\imsi.xlsx")
        wb_imsi = Workbook()
        ws_imsi = wb_imsi.active

        if class_type < 2:
            if class_type == 0:
                time = "주간"
            else:
                time = "야간"
            string_set = f"{ordinal_num}기{time}"
            ws_imsi.cell(row=1, column=1).value = f"{ordinal_num}기 {time}반 출석부"
        elif class_type == 2:
            time = "대체실습"
            string_set = f"{time} {ordinal_num}기"
            ws_imsi.cell(row=1, column=1).value = f"{time} {ordinal_num}기 출석부"
        else:
            print("Error: ErrorCode: mk_attendance_xlsx definition has error code")
            return
        i = 2

        print("\n\n******************************")
        print(f"{ordinal_num}기 {time} 출석부 명단을 작성합니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
        print("******************************\n\n")

        if class_type < 2:
            for idx, cell in enumerate(self.ws_members["E"], start=1):
                if idx < 5:
                    continue
                if string_set not in cell.value:
                    continue

                # 순번
                ws_imsi.cell(row=i, column=1).value = self.ws_members.cell(row=idx, column=1).value

                # 이름
                ws_imsi.cell(row=i, column=2).value = self.ws_members.cell(row=idx, column=18).value

                # 생년월일
                ws_imsi.cell(row=i, column=3).value = f"{self.ws_members.cell(row=idx, column=20).value[:2]}. {self.ws_members.cell(row=idx, column=20).value[2:4]}. {self.ws_members.cell(row=idx, column=20).value[4:6]}"

                i += 1

                print(f"{self.ws_members.cell(row=idx, column=1).value}. {ordinal_num}기{time} {self.ws_members.cell(row=idx, column=18).value} 작성 완료.")

        else:
            for idx, cell in enumerate(self.ws_members["H"], start=1):
                if idx < 5:
                    continue
                if string_set not in cell.value:
                    continue

                # 순번
                ws_imsi.cell(row=i, column=1).value = i - 1

                # 이름
                ws_imsi.cell(row=i, column=2).value = self.ws_members.cell(row=idx, column=18).value

                # 생년월일
                ws_imsi.cell(row=i, column=3).value = f"{self.ws_members.cell(row=idx, column=20).value[:2]}. {self.ws_members.cell(row=idx, column=20).value[2:4]}. {self.ws_members.cell(row=idx, column=20).value[4:6]}"

                i += 1

                print(f"{self.ws_members.cell(row=idx, column=1).value}. {ordinal_num}기{time} {self.ws_members.cell(row=idx, column=18).value} 작성 완료.")
            
        print("\n\n******************************")
        print(f"{ordinal_num}기 {time} 출석부 명단 작성을 완료했습니다.\n저장파일경로: 'D:\\Master\\imsi.xlsx'")
        print("******************************\n\n")
        wb_imsi.save("D:\\Master\\imsi.xlsx")
        wb_imsi.close()


    def list_pass(self, exam_round, exist=0):
        # excel 호환모드(.xls)는 openpyxl 로 다룰 수 없음.
        wb_pass = load_workbook("D:\\"+operate_data.ac_name+f"\\경기도청\\자격증발급\\{exam_round}회_제출용\\화성시-"+operate_data.ac_name+f"-{exam_round}회합격자명단.xlsx")
        # D:\남양노아요양보호사교육원\경기도청\자격증발급\35회_제출용\화성시-남양노아요양보호사교육원-35회합격자명단
        ws_pass = wb_pass.active
        starting_row = 5 + exist
        for idx, cell in enumerate(self.ws_members["X"], start=1):
            if cell.value != exam_round:
                continue
            # 순번
            ws_pass.cell(row=starting_row, column=1).value = starting_row - 4
            
            # 합격번호

            #시행기관
            ws_pass.cell(row=starting_row, column=3).value = "한국의료보험인국가시험원"

            # 시험 시행일
            if cell.value == 34:
                string = "20210220"
            elif cell.value == 35:
                string = "20210515"
            elif cell.value == 36:
                string = "20210807"
            elif cell.value == 37:
                string = "20211106"
            ws_pass.cell(row=starting_row, column=4).value = string

            # 시험 합격일
            if cell.value == 34:
                string = "20210309"
            elif cell.value == 35:
                string = "20210601"
            elif cell.value == 36:
                string = "20210824"
            elif cell.value == 37:
                string = "20211123"
            ws_pass.cell(row=starting_row, column=5).value = string

            # 교육이수일자
            string = str(self.ws_members.cell(row=idx, column=10).value)
            string = string.replace("-","")
            ws_pass.cell(row=starting_row, column=6).value = string

            # 교육시작일자
            string = str(self.ws_members.cell(row=idx, column=6).value)
            string = string.replace("-","")
            ws_pass.cell(row=starting_row, column=7).value = string

            # 교육마감일자
            string = str(self.ws_members.cell(row=idx, column=10).value)
            string = string.replace("-","")
            ws_pass.cell(row=starting_row, column=8).value = string

            # 대상구분
            if self.ws_members.cell(row=idx, column=15).value == "일반":
                string = "일반교육과정"
            elif self.ws_members.cell(row=idx, column=15).value == "경력자":
                string = "경력자과정"
            else:
                string = "자격/면허 소지자 과정"
            ws_pass.cell(row=starting_row, column=9).value = string

            # 교육과정명
            string = self.ws_members.cell(row=idx, column=5).value[2:4] + "반 " + self.ws_members.cell(row=idx, column=4).value
            ws_pass.cell(row=starting_row, column=10).value = string

            # 총 교육시간
            ws_pass.cell(row=starting_row, column=11).value = self.ws_members.cell(row=idx, column=11).value

            # 이론
            ws_pass.cell(row=starting_row, column=12).value = self.ws_members.cell(row=idx, column=12).value

            # 실기
            ws_pass.cell(row=starting_row, column=13).value = self.ws_members.cell(row=idx, column=13).value

            # 실습
            ws_pass.cell(row=starting_row, column=14).value = self.ws_members.cell(row=idx, column=14).value

            # 자격/면허 취득 정보(자격증반 해당)
            if self.ws_members.cell(row=idx, column=15).value == "일반" or self.ws_members.cell(row=idx, column=15).value == "경력자":
                pass
            else:
                if self.ws_members.cell(row=idx, column=15).value[4:6] == "간조":
                    string = "간호조무사"
                elif self.ws_members.cell(row=idx, column=15).value[4:6] == "사복":
                    string = "사회복지사"
                elif self.ws_members.cell(row=idx, column=15).value[4:6] == "간호":
                    string = "간호사"
                ws_pass.cell(row=starting_row, column=15).value = string

            # 관련자격정보 + 경력사항 = pass

            # 내외국인 구분 : 이거는 잘 체크 해야됨 !
            ws_pass.cell(row=starting_row, column=32).value = "내국인"

            # 주민등록번호
            ws_pass.cell(row=starting_row, column=33).value = self.ws_members.cell(row=idx, column=20).value

            # 성명
            ws_pass.cell(row=starting_row, column=34).value = self.ws_members.cell(row=idx, column=18).value

            # 주소
            ws_pass.cell(row=starting_row, column=35).value = self.ws_members.cell(row=idx, column=22).value

            # 등록기준지(본적)
            ws_pass.cell(row=starting_row, column=36).value = self.ws_members.cell(row=idx, column=23).value

            # 전화번호

            # 핸드폰번호
            ws_pass.cell(row=starting_row, column=38).value = self.ws_members.cell(row=idx, column=19).value
            
            starting_row += 1

        wb_pass.save("D:\\"+operate_data.ac_name+f"\\경기도청\\자격증발급\\{exam_round}회_제출용\\화성시-"+operate_data.ac_name+f"-{exam_round}회합격자명단.xlsx")
        wb_pass.close()


    def update_attendance(self, update_num, update_time):
        # how to use ? : update_attendance(4, "야간") 업데이트를 진행할 기수 + 시간
        name = f"{update_num}기{update_time}_출석부(기관장)"
        wb_update = load_workbook("D:\\"+operate_data.ac_name+f"\\교육생관리\\출석부_기관장용\\{name}.xlsx")
        ws_update = wb_update.active

        string_set = f"{update_num}기{update_time}"

        # 출석부 파일의 시작점과 각 결석일수를 참조하기 위한 초기화 작업
        if update_time == "주간":
                num = 7
                start_set = 3
        elif update_time == "야간":
            num = 4
            start_set = 0
        member = 1

        for idx, cell in enumerate(self.ws_members["E"], start=1):
            if not string_set in str(cell.value):
                continue

            # 실습 시간이 안적혀 있는 경우, 0시간으로 초기화하여 오류 방지
            if self.ws_members.cell(row=idx, column=14).value == None:
                self.ws_members.cell(row=idx, column=14).value = 0
            
            # 출석부 결석시간이 안적혀 있는 경우, 0시간으로 초기화하여 오류 방지
            if ws_update.cell(row=(member * num) - start_set, column=4).value == None:
                ws_update.cell(row=(member * num) - start_set, column=4).value = 0
            if ws_update.cell(row=(member * num) - start_set, column=5).value == None:
                ws_update.cell(row=(member * num) - start_set, column=5).value = 0
            
            if self.ws_members.cell(row=idx, column=15).value == "일반":
                self.ws_members.cell(row=idx, column=12).value = 80 - ws_update.cell(row=(member * num) - start_set, column=4).value
                self.ws_members.cell(row=idx, column=13).value = 80 - ws_update.cell(row=(member * num) - start_set, column=5).value
                self.ws_members.cell(row=idx, column=11).value = self.ws_members.cell(row=idx, column=12).value + self.ws_members.cell(row=idx, column=13).value + self.ws_members.cell(row=idx, column=14).value
            
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                self.ws_members.cell(row=idx, column=12).value = 31 - ws_update.cell(row=(member * num) - start_set, column=4).value
                self.ws_members.cell(row=idx, column=13).value = 11 - ws_update.cell(row=(member * num) - start_set, column=5).value
                self.ws_members.cell(row=idx, column=11).value = self.ws_members.cell(row=idx, column=12).value + self.ws_members.cell(row=idx, column=13).value + self.ws_members.cell(row=idx, column=14).value
            
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                self.ws_members.cell(row=idx, column=12).value = 32 - ws_update.cell(row=(member * num) - start_set, column=4).value
                self.ws_members.cell(row=idx, column=13).value = 10 - ws_update.cell(row=(member * num) - start_set, column=5).value
                self.ws_members.cell(row=idx, column=11).value = self.ws_members.cell(row=idx, column=12).value + self.ws_members.cell(row=idx, column=13).value + self.ws_members.cell(row=idx, column=14).value
            
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                self.ws_members.cell(row=idx, column=12).value = 26 - ws_update.cell(row=(member * num) - start_set, column=4).value
                self.ws_members.cell(row=idx, column=13).value = 6 - ws_update.cell(row=(member * num) - start_set, column=5).value
                self.ws_members.cell(row=idx, column=11).value = self.ws_members.cell(row=idx, column=12).value + self.ws_members.cell(row=idx, column=13).value + self.ws_members.cell(row=idx, column=14).value

            member += 1

        self.wb_members.save(operate_data.database_path)


    def mkfile(self, new_path_num, new_path_time, file_name):
        # new_path 는 기수를 설정하여, 각 기수의 멤버를 받아 파일을 복사한다.
        # ex) original_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\4기\\4기주간1207\\1.abc\\abc_요양보호사 자격증 발급,재발급 신청서.hwp" / new_path =  "4기야간0201"
        # how to use ? : x.mkfile(4, "주간", "교육수료증명서.hwp") !!! 주의 !!! 꼭 확장자 명을 작성할 것 !
        string_set = f"{new_path_num}기{new_path_time}"
        print(string_set)

        switch = 0

        i = 1
        original_path = f"D:\\Master\\mkfile\\{file_name}"
        print("파일을 복사합니다.\n원본파일 :", original_path)
        for idx, cell in enumerate(self.ws_members["E"], start=1):
            if idx < 5:
                continue
            if string_set not in cell.value:
                continue
            if switch == 0:
                string_set = cell.value
                switch = 1
            if self.ws_members.cell(row=idx, column=15).value == "일반":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_사복"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간조"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간호"
            
            if not os.path.isdir(path):
                print(path, " 폴더가 존재하지 않아 폴더를 생성합니다.")
                os.mkdir(path)

            if self.ws_members.cell(row=idx, column=15).value == "일반":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_사복\\{self.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간조\\{self.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif self.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {self.ws_members.cell(row=idx, column=18).value}_간호\\{self.ws_members.cell(row=idx, column=18).value}_{file_name}"
            shutil.copyfile(original_path, path)
            print("파일이 복사되었습니다 :", path)
            i += 1

    def print_file(self, ordinal_num, time, task, option):
        if option == 0:
            i = 0
            string_set = f"{ordinal_num}기{time}"

            non_member = None
            non_member_list = []
            checker = True
            while(True):
                non_member = input(f"{ordinal_num}기 {time}반 중 제외할 사람을 입력해 주세요(q를 누르면 종료합니다.):")
                if not(non_member == "q" or non_member == "Q"):
                    non_member_list.append(non_member)
                    print(non_member_list)

                elif(non_member == "q" or non_member == "Q"):
                    break

            while(True):
                if string_set in self.ws_members.cell(row=5+i, column=5).value:
                    string_set = self.ws_members.cell(row=5+i, column=5).value
                    break
                else:
                    i += 1
            
            print(string_set)
            # D:\남양노아요양보호사교육원\교육생관리\7기\7기주간0503\8. 윤지숙
            dir_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\" + str(ordinal_num) + "기\\" + string_set
            print(dir_path)
            os.chdir(dir_path)
            pattern = "*_" + task + ".xlsx"
            result = []
            for root, dirs, files in os.walk(os.getcwd()):
                # [a.txt, b.txt, c.txt, 11_file_system.py, ...]
                print("files:", files)
                for name in files:
                    checker = True
                    for non_member in non_member_list:
                        if non_member in name:
                            checker = False
                            print(non_member, "출력 X")
                            continue
                    
                    if checker == True:
                        print("name:", name)
                        if fnmatch.fnmatch(name, pattern): # 이름과 패턴이 일치하면
                            result.append(os.path.join(root, name))

            print(result)            
            print("이 프로그램은 1번 -> 10번대 -> 2번 -> 20번대 ,,, 순으로 출력합니다.")

            for file_name in result:
                print(file_name + " 출력을 시작합니다.")
                os.startfile(file_name, "print")

        if option == 1:
            string_set = f"{ordinal_num}기{time}"

    def copy_picture(self, exam):
        # 시험 차수: 36 <class 'int'>
        not_exist = []
        exsist = []
        # D:\남양노아요양보호사교육원\경기도청\자격증발급\35회_제출용\자격증 사진
        copy_path = "D:\\"+operate_data.ac_name+f"\\경기도청\\자격증발급\\{exam}회_제출용\\자격증 사진"
        if os.path.exists(copy_path):
            print(copy_path, "폴더가 존재하여 삭제합니다.")
            shutil.rmtree(copy_path)
            os.mkdir(copy_path)
        else:
            print(copy_path, "폴더가 존재하지 않아 폴더를 생성합니다.")
            os.mkdir(copy_path)
        for idx, cell in enumerate(self.ws_members["X"], start=1):
            if idx <= 4:
                continue
            if cell.value == None:
                continue
            if not exam == int(cell.value):
                continue
            print("start")
            # D:\남양노아요양보호사교육원\교육생관리\6기\6기주간0315\10. 이순희\6기주간_이순희.jpg
            name = self.ws_members.cell(row=idx, column=18).value
            ordinal_num = self.ws_members.cell(row=idx, column=4).value
            time = self.ws_members.cell(row=idx, column=5).value
            folder_order = self.ws_members.cell(row=idx, column=1).value
            file_name = str(self.ws_members.cell(row=idx, column=20).value[:6]) + str(self.ws_members.cell(row=idx, column=20).value[7:])
            print(f"{name} : {time[:4]}")
            
            if self.ws_members.cell(row=idx, column=15).value == "일반":
                value = "일반"
                original_path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(ordinal_num)}\\{time}\\{folder_order}. {name}\\{time[:4]}_{name}.jpg"
                if not os.path.isfile(original_path):
                    print("\n파일이 존재하지 않습니다\n" + original_path + "\n\n")
                    not_exist.append(name)
            else:
                value = self.ws_members.cell(row=idx, column=15).value[4:6]
                original_path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(ordinal_num)}\\{time}\\{folder_order}. {name}_{value}\\{time[:4]}_{name}_{value}.jpg"
                if os.path.isfile(original_path):
                    pass
                else:
                    original_path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(ordinal_num)}\\{time}\\{folder_order}. {name}_{value}\\{time[:4]}_{name}.jpg"
                    if not os.path.isfile(original_path):
                        print("\n파일이 존재하지 않습니다\n" + original_path + "\n\n")
                        not_exist.append(name)

            copy_path = "D:\\"+operate_data.ac_name+f"\\경기도청\\자격증발급\\{exam}회_제출용\\자격증 사진\\{file_name}.jpg"
            shutil.copy2(original_path, copy_path)
            print("파일이 복사되었습니다 :", name, "->", file_name)

        if not not_exist == []:
            print("존재하지 않는 파일:", not_exist)

    def test(self, exam, task):
        i = 0
        exam_num = exam
        names = ["진유정", "양혜주", "구미정", "조미영", "나은미", "신세연", "김주리", "김희정"]
        printMemberIndex = {0:[], 4:[], 5:[], 6:[], 7:[], 8:[], 9:[], 10:[]}
        printPath = []
        hasNoFile = []
        totalMember = 0

        print(f"{exam}회차 인원")
        for idx in range(1, 300):
            # print(str(self.ws_members.cell(row=idx, column=24)).strip())
            
            if(str(self.ws_members.cell(row=idx, column=24).value).strip() == "37"):
                dir_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\" + str(self.ws_members.cell(row=idx, column=4).value).strip() + "\\" + str(self.ws_members.cell(row=idx, column=5).value).strip()
                
                if dir_path not in printPath:
                    print("path 추가")
                    if os.path.isdir(dir_path):
                        printPath.append(dir_path)
                    else:
                        print("path가 존재하지 않습니다.")


                if(self.ws_members.cell(row=idx, column=15).value.strip() == "일반"):
                    print_path = dir_path + f"\\{self.ws_members.cell(row=idx, column=1).value}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                else:
                    print_path = dir_path + f"\\{self.ws_members.cell(row=idx, column=1).value}. {self.ws_members.cell(row=idx, column=18).value}_{self.ws_members.cell(row=idx, column=15).value[4:6]}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"


                print(f"기수: {self.ws_members.cell(row=idx, column=4).value}\t반: {self.ws_members.cell(row=idx, column=5).value}\t이름: {self.ws_members.cell(row=idx, column=18).value}")
                print("path: ", print_path)
                inner = 0
                if(idx <= 222):
                    inner = int(self.ws_members.cell(row=idx, column=4).value[0])
                else:
                    inner = int(self.ws_members.cell(row=idx, column=4).value[:2])

                # printMemberIndex[inner].append(self.ws_members.cell(row=idx, column=18).value)
                printMemberIndex[inner].append(print_path)
                totalMember += 1

                if not(os.path.isfile(print_path)):
                    print("파일이 존재하지 않습니다.")
                    hasNoFile.append(self.ws_members.cell(row=idx, column=18).value)

            elif(str(self.ws_members.cell(row=idx, column=18).value).strip() in names):
                dir_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\" + str(self.ws_members.cell(row=idx, column=4).value).strip() + "\\" + str(self.ws_members.cell(row=idx, column=5).value).strip()
                
                if dir_path not in printPath:
                    print("path 추가")
                    if os.path.isdir(dir_path):
                        printPath.append(dir_path)
                    else:
                        print("path가 존재하지 않습니다.")


                if(self.ws_members.cell(row=idx, column=15).value.strip() == "일반"):
                    print_path = dir_path + f"\\{self.ws_members.cell(row=idx, column=1).value}. {self.ws_members.cell(row=idx, column=18).value}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                else:
                    print_path = dir_path + f"\\{self.ws_members.cell(row=idx, column=1).value}. {self.ws_members.cell(row=idx, column=18).value}_{self.ws_members.cell(row=idx, column=15).value[4:6]}\\{self.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                print(f"기수: {self.ws_members.cell(row=idx, column=4).value}\t반: {self.ws_members.cell(row=idx, column=5).value}\t이름: {self.ws_members.cell(row=idx, column=18).value}")
                print("path: ", print_path)

                inner = 0
                if(idx <= 222):
                    inner = int(self.ws_members.cell(row=idx, column=4).value[0])
                else:
                    inner = int(self.ws_members.cell(row=idx, column=4).value[:2])

                # printMemberIndex[inner].append(self.ws_members.cell(row=idx, column=18).value)
                printMemberIndex[inner].append(print_path)
                totalMember += 1

                if not(os.path.isfile(print_path)):
                    print("파일이 존재하지 않습니다.")
                    hasNoFile.append(self.ws_members.cell(row=idx, column=18).value)

        print(printMemberIndex)
        print(printPath)
        print("total: ", totalMember)
        print(hasNoFile)
        sys.exit()

        for mList in printMemberIndex:
            if(mList >= 6):
                print("테스트를 위해 진행하지 않습니다.")
                continue
            for file in mList:
                os.startfile(file, "print")






        non_member = None
        non_member_list = []
        checker = True

        print(string_set)
        # D:\남양노아요양보호사교육원\교육생관리\7기\7기주간0503\8. 윤지숙
        for idx in printMemberIndex.keys():
            for i in idx:
                dir_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\" + str(self.ws_members.cell(row=idx, column=4).value).strip() + "\\" + str(self.ws_members.cell(row=idx, column=5).value).strip()
        print(dir_path)
        os.chdir(dir_path)
        pattern = "*_" + task + ".xlsx"
        result = []
        for root, dirs, files in os.walk(os.getcwd()):
            # [a.txt, b.txt, c.txt, 11_file_system.py, ...]
            print("files:", files)
            for name in files:
                checker = True
                for non_member in non_member_list:
                    if non_member in name:
                        checker = False
                        print(non_member, "출력 X")
                        continue
                
                if checker == True:
                    print("name:", name)
                    if fnmatch.fnmatch(name, pattern): # 이름과 패턴이 일치하면
                        result.append(os.path.join(root, name))

        print(result)            
        print("이 프로그램은 1번 -> 10번대 -> 2번 -> 20번대 ,,, 순으로 출력합니다.")

        for file_name in result:
            print(file_name + " 출력을 시작합니다.")
            # os.startfile(file_name, "print")



# file_position_detail : x = 335, y = 230, +- = 21
# file_position_big_icorn : x = 330, y = 242, +- = 110
# folder_goto_back : 20,165
# win + left 기준 폰트 : 615,150 / 맑은고딕 : 700,465
# file_path : alt + D
# pyautogui.scroll() = 한 틱당 약 500

# 교육수료증명서(최 우측, win + left 기준)
# 호수 : 300,320 / 이름 : 715,410 / 주소 :  715,455 / 주민등록번호 : 475,505 / 전화번호 : 715,505 / 교육과정명 : 715,545 / 이론실기 이수기간 : 595,635 / 이론실기 이수시간 : 715,635 / 
# 실습기간 : 600,730 / 실습시간 : 715,730 / 총 실습시간 : 715,890 / 총 이수시간 : 715,945 / -> pyautogui.scroll(-2000) -> / 수여일 : 600,765 / 닫기 : 942,12 

# 대체실습확인서(최 우측, win + left 기준)
# 이름 220,475 / 생년월일 305,475 / 교육기관명 565,475 / 교육과정명 685,475 / 구분 : 215,610(640) *** 이름 ~ 도 승인일자 틀림 수정필요 *** / 이름 : 300,555 / 생년월일 : 395,555 / 
# 자격사항 490,555 / 경력기간 590,555 / 도 승인일자 685,555 대체실습 기간 675,785 / 대체실습 시간 680,825 / 합격여부 375,890 / 자체시험 점수 680,890 / 비고 380,930 / 교수이름 680,930 /
# 수여 날짜 705,780

# 요양보호사 자격증 발급, 재발급 신청서(최 우측, win + left 기준)
# 이름 600,420 / 요양보호사 교육기간(이론, 실기) 320,580 / 발급신청일(표 중앙) 500,795 // 나머지는 alt + right 로 커버 ^^
# x = automation()
# x.auto_move_class(3, "야간")
# x.automation_task("edu_completion")

# 대체실습 수료보고
# 제목 615,450 / 2페이지 연번 915,600 / 18번 스크롤(2250) 한 후 3페이지 연번(11) 140,435 / 4페이지 연번(25) 1220,435 / 14번 스크롤 한 후 5페이지 연번(39) 140,430 / 
#  교육과정명 710,570 / 교육인원/수료인원 710,610 / 이론/실기 교육시간 710,690 / 실습 교육기간 710,755 / 수료일 710,795 / 1번 교육과정명 975,600 / 
# 2페이지 11번 교육과정명 195,445

# 개강보고
# ac 교육대상자 명단
# 과정구분 1095,345

# 대체실습 실시보고 scroll(-2500) 기준
# [교육생명단] 교육구분 370,395

# 출석부
# 주간 355,480 / 야간 1140,515

# 한글 꿀팁
# 서식복사 : alt + C


# pyautogui.mouseInfo()

