import sys
import shutil
# import time : 코드 내에서 time 변수가 다수 사용되어, time.sleep -> pyautogui.sleep 로 사용
import pyautogui
import pyperclip
from openpyxl import load_workbook
import operate_data

#                                       example
# preform = automation()
# perform.auto_move_class(4, "야간")
# perform.auto_move_report()
# perform.automation_task_students(3, "주간", "자격증 발급,재발급 신청서")
# perform.automation_task_report(5, "주간", "개강보고") # 수료보고가 따로 없기 때문에 kind = 개강보고로 고정 !
# perform.automation_task_temporary(4, "수료보고")
# perform.(4, "주간", "교육수료증명서.hwp") # 뒤에 복사할 파일을 입력할 때 꼭 !!!!! 확장자 명까지 작성하기 ㅎㅎ
# perform.mkattendance(3, "야간")
# perform.update_attendance(4, "야간") 업데이트를 진행할 기수 + 시간

# 업데이트 목록
# 각 업무자동화 시스템 변수 초기화 하기(이름, 기수, 대체기수 ,,,etc.)
# 수여일 수정

class automation:
    

    # 수행 전, 명단총정리 엑셀 파일을 불러와 자료를 복사할 준비
    wb_members = load_workbook("D:\\Master\\"+operate_data.ac_name+"_명단총정리.xlsx")
    ws_members = wb_members.active


    def __init__(self): # slp : sleep / dura : duration / itv : interval / sleep, duration, interval 을 일괄적으로 시간을 정하기 위해 설정한다.
        print("Noa Automation Program")

    def open_file_explorer(self):
        global fw_epr

        # pyautogui.PAUSE = 0.3
        pyautogui.hotkey("win", "r")
        pyautogui.sleep(0.5)
        pyautogui.write("explorer", interval=0.05)
        pyautogui.press("enter")

        # epr : explorer
        fw_epr = pyautogui.getWindowsWithTitle("파일 탐색기")
        while fw_epr == []:
            fw_epr = pyautogui.getWindowsWithTitle("파일 탐색기")
        fw_epr = pyautogui.getWindowsWithTitle("파일 탐색기")[0]
        if fw_epr.isActive == False:
            fw_epr.activate()
        
        pyautogui.sleep(0.3)

    def auto_move_class(self, ordinal_num, time):
        automation.open_file_explorer(self)
        self.ordinal_num = ordinal_num
        self.time = time

        if ordinal_num >5:
            # print("현재 기수는 5기까지 있습니다.")
            # print("입력값 : " + str(ordinal_num))
            pyautogui.alert("현재 기수는 5기까지 있습니다.\n입력값 : " + str(ordinal_num), "오류")
            return
        
        if time != "주간" and time != "야간":
            # print("현재 반은 \"주간(주간)\", 야간\"(야간)\"으로 구성되어 있습니다.")
            # print("입력값 : " + str(time))
            pyautogui.alert("현재 반은 \"주간\", 야간\"\"으로 구성되어 있습니다.\n입력값 : " + str(time))
            return

        # click to disk_d
        pyautogui.sleep(0.1)
        pyautogui.click(fw_epr.left + 100, fw_epr.top + 695, duration=0.35)
        # pyautogui.sleep(0.1)
        # ac_location
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 314, duration=0.35)
        # 교육생 관리
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 314, duration=0.35)
        # 1기
        if ordinal_num == 1:
            # pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
            print("현재 3기부터 이용 가능합니다. (업데이트 필요)")
            return
            # if time == "주간":
            #     pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            # elif time == "야간":
            #     pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
        # 2기
        elif ordinal_num == 2:
            # pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            print("현재 3기부터 이용 가능합니다. (업데이트 필요)")
            return
            # if time == "주간":
            #     pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            # elif time == "야간":
            #     pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
        # 3기
        elif ordinal_num == 3:
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 272, duration=0.35)
            if time == "주간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            elif time == "야간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
        # 4기
        elif ordinal_num == 4:
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 293, duration=0.35)
            if time == "주간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            elif time == "야간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
        # 5기
        elif ordinal_num == 5:
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 314, duration=0.35)
            if time == "주간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 251, duration=0.35)
            elif time == "야간":
                pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)

        pyautogui.sleep(0.5)

    def auto_move_report(self):
        automation.open_file_explorer(self)
        # click to disk_d
        pyautogui.sleep(0.1)
        pyautogui.click(fw_epr.left + 100, fw_epr.top + 695, duration=0.35)
        # pyautogui.sleep(0.1)
        # ac_location
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 293, duration=0.35)
        # 경기도청
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 377, duration=0.35)

        pyautogui.sleep(0.5)

    def automation_task_students(self, ordinal_num, time, task, version):
        # 사용방법 : x.automation_task_students(3, 주간, "교육수료증명서")
        global fw_epr
        self.ordinal_num = ordinal_num
        self.time = time
        self.task = task
        self.version = version

        is_ready = pyautogui.confirm("자동화 프로그램을 시작하시기 전에\n1. 모든 각 문서는 D:\\Master\\rpa_basic_file 에서 복사된 파일이어야 합니다.\n(각 자료가 입력되어 있어야 함)\n2. 한글, 파일 탐색기(폴더) 가 하나라도 실행되어 있어선 안됩니다.\n준비가 되었으면 확인, 안되어있으면 취소를 눌러주세요.", "경고")
        if is_ready == "OK":
            pyautogui.alert("자동화 프로그램을 시작합니다.\n프로그램을 강제로 종료하고 싶으실 경우 마우스를 각 꼭짓점에 가져가주세요.", "실행")
        elif is_ready == "Cancel":
            pyautogui.alert("프로그램을 종료합니다.\n준비작업을 마치신 후 다시 실행해 주세요.", "종료")
            return

        wb_automation = load_workbook("D:\\Master\\업무자동화.xlsx")
        ws_automation = wb_automation.active

        # # 컴퓨터 특성 상 처음여는 파일 또는 작업은 열리는 시간이 오래걸리기 때문에 duration 을 직접 설정하여 작업 시간을 설정할 수도 있다.
        # self.dura = dura

        # fw_epr(파일 탐색기를 통해 열린 폴더) 이 전역변수로 되지 않아 매 함수마다 최신화
        # excel 모드가 True 일 경우, explorer 를 실행 할 필요가 없기 때문에 False 일 경우에만 실행
        if version == False:
            while fw_epr.isActive == False:
                fw_epr.activate()

        # 1기
        if ordinal_num == 1:
            pyautogui.alert("이 프로그램은 3기 주간반 부터 자동화가 가능합니다.\n(업데이트 필요)", "프로그램 오류")
            return
            if self.time == "주간":
                members = 6
                
            elif self.time == "야간":
                members = 16
        # 2기
        elif ordinal_num == 2:
            pyautogui.alert("이 프로그램은 3기 주간반 부터 자동화가 가능합니다.\n(업데이트 필요)", "프로그램 오류")
            return
            if self.time == "주간":
                members = 6
            elif self.time == "야간":
                members = 9
        # 3기
        elif ordinal_num == 3:
            if self.time == "주간":
                members = 20
            elif self.time == "야간":
                members = 12
        # 4기
        elif ordinal_num == 4:
            if self.time == "주간":
                members = 15
            elif self.time == "야간":
                members = 1
        # 5기
        elif ordinal_num == 5:
            if self.time == "주간":
                members = 1
            elif self.time == "야간":
                members = 1
        
        # excel 파일을 불러오기 위해 경로를 최신화 하기 위한 참조
        wb_automation = load_workbook("D:\\Master\\업무자동화.xlsx")
        ws_automation = wb_automation.active
        string_set = f"{ordinal_num}기{time}"
        
        if task == "교육수료증명서":
            if version == True:
                i = 1
                for idx, cell in enumerate(automation.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if time == "주간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=2).value
                    elif time == "야간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=3).value
                    if automation.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간조\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_사복\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간호\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_completion = load_workbook(string_stu)
                    ws_completion = wb_completion.active

                    # 교육수료증명서 호수
                    string = f"    2021  년  제  {automation.ws_members.cell(row=idx, column=2).value} 호"
                    ws_completion.cell(row=1, column=1).value = string

                    # 이름
                    string = f" {automation.ws_members.cell(row=idx, column=18).value[0]} {automation.ws_members.cell(row=idx, column=18).value[1]} {automation.ws_members.cell(row=idx, column=18).value[2]}"
                    ws_completion.cell(row=4, column=3).value = string

                    # 주소
                    string = f" {automation.ws_members.cell(row=idx, column=21).value}"
                    ws_completion.cell(row=5, column=3).value = string

                    # 주민등록번호
                    string = f" {automation.ws_members.cell(row=idx, column=20).value[:6]} - {automation.ws_members.cell(row=idx, column=20).value[7:]}"
                    ws_completion.cell(row=6, column=3).value = string

                    # 전화번호
                    string = f"{automation.ws_members.cell(row=idx, column=19).value}"
                    ws_completion.cell(row=6, column=6).value = string

                    # 교육과정명
                    string = f" 요양보호사 {automation.ws_members.cell(row=idx, column=4).value}"
                    ws_completion.cell(row=7, column=3).value = string

                    # 이론실기 이수기간 / 각 기수별로 기간 선정, 2020 년  11 월  16 일 ∼  21 년  01 월 15 일 형식으로, 끝기간은 년도수 두자리만 표시
                    string = f"{automation.ws_members.cell(row=idx, column=6).value[:4]} 년  {automation.ws_members.cell(row=idx, column=6).value[5:7]} 월  {automation.ws_members.cell(row=idx, column=6).value[8:]} 일 ∼  {automation.ws_members.cell(row=idx, column=7).value[2:4]} 년  {automation.ws_members.cell(row=idx, column=7).value[5:7]} 월 {automation.ws_members.cell(row=idx, column=7).value[8:]} 일 "
                    ws_completion.cell(row=9, column=3).value = string

                    # 이론실기 이수시간
                    string = f"        {str(int(automation.ws_members.cell(row=idx, column=12).value) + int(automation.ws_members.cell(row=idx, column=13).value))}  시간"
                    ws_completion.cell(row=9, column=7).value = string

                    # 실습기간 / 대체실습 각 기수별 or 실습기간 따로 만들기,  21년 01월 18일 ∼ 21년 03월 13일 형식으로, 년도수 두자리만 표시
                    string = f" {automation.ws_members.cell(row=idx, column=9).value[2:4]}년 {automation.ws_members.cell(row=idx, column=9).value[5:7]}월 {automation.ws_members.cell(row=idx, column=9).value[8:]}일 ∼ {automation.ws_members.cell(row=idx, column=10).value[2:4]}년 {automation.ws_members.cell(row=idx, column=10).value[5:7]}월 {automation.ws_members.cell(row=idx, column=10).value[8:]}일"
                    ws_completion.cell(row=12, column=4).value = string

                    # 대체실습이 종료되면, 각 사람마다 실습시간(각 기관) 이 달라짐. 업데이트 필요
                    # 실습시간
                    string = f"        {automation.ws_members.cell(row=idx, column=14).value}  시간"

                    ws_completion.cell(row=12, column=7).value = string

                    # 총 실습시간
                    string = f"         {automation.ws_members.cell(row=idx, column=14).value}  시간"
                    ws_completion.cell(row=18, column=7).value = string

                    # 총 이수시간
                    string = f"       {automation.ws_members.cell(row=idx, column=11).value}  시간"
                    ws_completion.cell(row=19, column=7).value = string

                    # 수여일 / 각 인원 대체실습 기준 종료일 바로 다음 월요일 날짜로 지정
                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        string = f"                               {ws_automation.cell(row=3, column=3).value[:4]} 년    {ws_automation.cell(row=3, column=3).value[5:7]} 월     {ws_automation.cell(row=3, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        string = f"                               {ws_automation.cell(row=4, column=3).value[:4]} 년    {ws_automation.cell(row=4, column=3).value[5:7]} 월     {ws_automation.cell(row=4, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        string = f"                               {ws_automation.cell(row=5, column=3).value[:4]} 년    {ws_automation.cell(row=5, column=3).value[5:7]} 월     {ws_automation.cell(row=5, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        string = f"                               {ws_automation.cell(row=6, column=3).value[:4]} 년    {ws_automation.cell(row=6, column=3).value[5:7]} 월     {ws_automation.cell(row=6, column=3).value[8:]} 일"
                    ws_completion.cell(row=23, column=1).value = string

                    wb_completion.save(string_stu)
                    wb_completion.close()

                    i += 1

            elif version == False:
                i = 0
                for idx, cell in enumerate(automation.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    fw_epr = pyautogui.getActiveWindow()

                    # 학생 폴더 선택
                    pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230 + (i * 21), duration=0.35)
                    # pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230, duration=0.35)
                    # 큰 아이콘으로 만들기
                    pyautogui.sleep(0.5)
                    pyautogui.click(fw_epr.left + 175, fw_epr.top + 35, duration=0.35)
                    pyautogui.click(fw_epr.left + 311, fw_epr.top + 72, duration=0.35)
                    # 교육수료 증명서 실행
                    pyautogui.doubleClick(fw_epr.left + 440, fw_epr.top + 242, duration=0.35)
                    fw_completion = pyautogui.getWindowsWithTitle("교육수료증명서")
                    while fw_completion == []:
                        fw_completion = pyautogui.getWindowsWithTitle("교육수료증명서")
                    fw_completion = pyautogui.getWindowsWithTitle("교육수료증명서")[0]
                    pyautogui.sleep(0.5)
                    fw_completion.activate()

                    # hotkey 를 통해 좌측으로 미룰 때, 이미 좌측에 있으면 맨 우측으로 넘어가는 것을 방지
                    if fw_completion.isMaximized == False:
                        fw_completion.maximize()
                    pyautogui.hotkey("win", "left")
                    pyautogui.press("esc")
                    pyautogui.sleep(1.5)
                    # 파일에 스크롤이 내려가 있을 수 있기 때문에 한번 쭉 올려준다
                    pyautogui.moveTo(fw_completion.left + 440, fw_completion.top + 500)
                    pyautogui.scroll(5000)

                    # 교육수료증명서 호수
                    pyautogui.click(fw_completion.left + 300, fw_completion.top + 320)
                    pyautogui.sleep(1)
                    string = f"    2021  년  제  {automation.ws_members.cell(row=idx, column=2).value} 호"
                    pyperclip.copy(string)
                    pyautogui.click()
                    pyautogui.click()
                    pyautogui.click()
                    pyautogui.hotkey("ctrl", "v")
                    # 수여일 형식속 enter 을 지운 후 다시 복구하기 위한 press
                    pyautogui.press("enter")

                    for j in range(3):
                        pyautogui.hotkey("alt", "right")

                    # 이름
                    string = f" {automation.ws_members.cell(row=idx, column=18).value[0]} {automation.ws_members.cell(row=idx, column=18).value[1]} \
                        {automation.ws_members.cell(row=idx, column=18).value[2]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(3):
                        pyautogui.hotkey("alt", "right")
                    
                    # 주소
                    string = f" {automation.ws_members.cell(row=idx, column=21).value}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(3):
                        pyautogui.hotkey("alt", "right")

                    # 주민등록번호
                    string = f" {automation.ws_members.cell(row=idx, column=20).value[:6]} - {automation.ws_members.cell(row=idx, column=20).value[7:]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(2):
                        pyautogui.hotkey("alt", "right")

                    # 전화번호
                    string = f"{automation.ws_members.cell(row=idx, column=19).value}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(3):
                        pyautogui.hotkey("alt", "right")

                    # 교육과정명
                    string = f" 요양보호사 {automation.ws_members.cell(row=idx, column=4).value}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(7):
                        pyautogui.hotkey("alt", "right")

                    # 이론실기 이수기간 / 각 기수별로 기간 선정, 2020 년  11 월  16 일 ∼  21 년  01 월 15 일 형식으로, 끝기간은 년도수 두자리만 표시
                    string = f"{automation.ws_members.cell(row=idx, column=6).value[:4]} 년  {automation.ws_members.cell(row=idx, column=6).value[5:7]} 월  \
                        {automation.ws_members.cell(row=idx, column=6).value[8:]} 일 ∼  {automation.ws_members.cell(row=idx, column=7).value[2:4]} 년  \
                            {automation.ws_members.cell(row=idx, column=7).value[5:7]} 월 {automation.ws_members.cell(row=idx, column=7).value[8:]} 일 "
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    pyautogui.hotkey("alt", "right")

                    # 이론실기 이수시간
                    string = f"        {str(int(automation.ws_members.cell(row=idx, column=12).value) + int(automation.ws_members.cell(row=idx, column=13).value))}  시간"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(9):
                        pyautogui.hotkey("alt", "right")

                    # 실습기간 / 대체실습 각 기수별 or 실습기간 따로 만들기,  21년 01월 18일 ∼ 21년 03월 13일 형식으로, 년도수 두자리만 표시
                    string = f" {automation.ws_members.cell(row=idx, column=9).value[2:4]}년 {automation.ws_members.cell(row=idx, column=9).value[5:7]}월 \
                        {automation.ws_members.cell(row=idx, column=9).value[8:]}일 ∼ {automation.ws_members.cell(row=idx, column=10).value[2:4]}년 \
                            {automation.ws_members.cell(row=idx, column=10).value[5:7]}월 {automation.ws_members.cell(row=idx, column=10).value[8:]}일"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    pyautogui.hotkey("alt", "right")

                    # 대체실습이 종료되면, 각 사람마다 실습시간(각 기관) 이 달라짐. 업데이트 필요
                    # 실습시간
                    string = f"        {automation.ws_members.cell(row=idx, column=14).value}  시간"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(14):
                        pyautogui.hotkey("alt", "right")

                    # 총 실습시간
                    string = f"         {automation.ws_members.cell(row=idx, column=14).value}  시간"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    for j in range(3):
                        pyautogui.hotkey("alt", "right")

                    # 총 이수시간
                    string = f"       {automation.ws_members.cell(row=idx, column=11).value}  시간"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    # 클릭 ver.:
                        # # 주소
                        # string = " {}" .format(automation.ws_members.cell(row=idx, column=21).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 455, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 주민등록번호
                        # string = " {} - {}" .format(automation.ws_members.cell(row=idx, column=20).value[:6], automation.ws_members.cell(row=idx, column=20).value[7:])
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 475, fw_completion.top + 505, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 전화번호
                        # string = "{}" .format(automation.ws_members.cell(row=idx, column=19).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 505, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 교육과정명
                        # string = " 요양보호사 {}" .format(automation.ws_members.cell(row=idx, column=4).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 545, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 이론실기 이수기간 / 각 기수별로 기간 선정, 2020 년  11 월  16 일 ∼  21 년  01 월 15 일 형식으로, 끝기간은 년도수 두자리만 표시
                        # string = "{} 년  {} 월  {} 일 ∼  {} 년  {} 월 {} 일 " .format(automation.ws_members.cell(row=idx, column=6).value[:4], \
                        #     automation.ws_members.cell(row=idx, column=6).value[5:7], automation.ws_members.cell(row=idx, column=6).value[8:], \
                        #         automation.ws_members.cell(row=idx, column=7).value[2:4], automation.ws_members.cell(row=idx, column=7).value[5:7], \
                        #             automation.ws_members.cell(row=idx, column=7).value[8:])
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 595, fw_completion.top + 635, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 이론실기 이수시간
                        # string = "        {}  시간" .format(str(int(automation.ws_members.cell(row=idx, column=12).value) + int(automation.ws_members.cell(row=idx, column=13).value)))
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 635, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 실습기간 / 대체실습 각 기수별 or 실습기간 따로 만들기,  21년 01월 18일 ∼ 21년 03월 13일 형식으로, 년도수 두자리만 표시
                        # string = " {}년 {}월 {}일 ∼ {}년 {}월 {}일" .format(automation.ws_members.cell(row=idx, column=9).value[2:4], \
                        #     automation.ws_members.cell(row=idx, column=9).value[5:7], automation.ws_members.cell(row=idx, column=9).value[8:], \
                        #         automation.ws_members.cell(row=idx, column=10).value[2:4], automation.ws_members.cell(row=idx, column=10).value[5:7], \
                        #             automation.ws_members.cell(row=idx, column=10).value[8:])
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 600, fw_completion.top + 730, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 대체실습이 종료되면, 각 사람마다 실습시간(각 기관) 이 달라짐. 업데이트 필요
                        # # 실습시간
                        # string = "        {}  시간" .format(automation.ws_members.cell(row=idx, column=14).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 730, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 총 실습시간
                        # string = "         {}  시간" .format(automation.ws_members.cell(row=idx, column=14).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 890, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 총 이수시간
                        # string = "       {}  시간" .format(automation.ws_members.cell(row=idx, column=11).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_completion.left + 715, fw_completion.top + 945, duration=0.35)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                    # 스크롤
                    pyautogui.scroll(-2000)

                    # 수여일 / 각 인원 대체실습 기준 종료일 바로 다음 월요일 날짜로 지정
                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        string = f"                               {ws_automation.cell(row=3, column=3).value[:4]} 년    {ws_automation.cell(row=3, column=3).value[5:7]} 월     \
                            {ws_automation.cell(row=3, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        string = f"                               {ws_automation.cell(row=4, column=3).value[:4]} 년    {ws_automation.cell(row=4, column=3).value[5:7]} 월     \
                            {ws_automation.cell(row=4, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        string = f"                               {ws_automation.cell(row=5, column=3).value[:4]} 년    {ws_automation.cell(row=5, column=3).value[5:7]} 월     \
                            {ws_automation.cell(row=5, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        string = f"                               {ws_automation.cell(row=6, column=3).value[:4]} 년    {ws_automation.cell(row=6, column=3).value[5:7]} 월     \
                            {ws_automation.cell(row=6, column=3).value[8:]} 일"
                    pyperclip.copy(string)
                    pyautogui.moveTo(fw_completion.left + 600, fw_completion.top + 765, duration=0.35)
                    # pyautogui.hotkey("ctrl", "a") 로 할 경우, 모든 데이터가 선택되지만, 고쳐야할 줄은 한 줄이기 때문에 click 을 세번 사용하여 한 줄 드래그
                    pyautogui.click()
                    pyautogui.click()
                    pyautogui.click()
                    pyautogui.hotkey("ctrl", "v")
                    # 수여일 형식속 enter 을 지운 후 다시 복구하기 위한 press
                    pyautogui.press("enter")

                    # win + left 된 상태에서 닫기버튼을 누른 후 저장에 enter
                    pyautogui.click(fw_completion.left + 940, fw_completion.top + 12)
                    pyautogui.press("enter")

                    # 폴더 뒤로가기
                    pyautogui.sleep(1.5)
                    pyautogui.press("backspace")
                    fw_epr = pyautogui.getActiveWindow()
                    print(fw_epr)
                    pyautogui.moveTo(fw_epr.left + 20, fw_epr.top + 165, duration=0.35)

                    print(i)

                    i += 1

        elif task == "대체실습확인서":
            wb_temp_score = load_workbook("D:\\Master\\대체실습_점수.xlsx")
            ws_temp_score = wb_temp_score.active
            if version == True:
                i = 1
                for idx, cell in enumerate(automation.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if time == "주간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=2).value
                    elif time == "야간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=3).value
                    if automation.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간조\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_사복\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간호\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_temp = load_workbook(string_stu)
                    ws_temp = wb_temp.active

                    # 대체실습확인서
                    # 이름
                    string = automation.ws_members.cell(row=idx, column=18).value
                    ws_temp.cell(row=7, column=2).value = string

                    # 생년월일
                    string = f"{automation.ws_members.cell(row=idx, column=20).value[:2]}. {automation.ws_members.cell(row=idx, column=20).value[2:4]}. {automation.ws_members.cell(row=idx, column=20).value[4:6]}"
                    ws_temp.cell(row=7, column=3).value = string

                    # 연락처
                    string = automation.ws_members.cell(row=idx, column=19).value
                    ws_temp.cell(row=7, column=4).value = string

                    # 교육기관명
                    string = automation.ws_members.cell(row=idx, column=3).value
                    ws_temp.cell(row=7, column=5).value = string

                    # 교육과정명
                    string = f" 요양보호사 {automation.ws_members.cell(row=idx, column=4).value}"
                    ws_temp.cell(row=7, column=7).value = string

                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        # operate_data.teacher[1]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=14, column=j).value
                            ws_temp.cell(row=12, column=j + 1).value = string

                        # operate_data.teacher[4]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=17, column=j).value
                            ws_temp.cell(row=13, column=j + 1).value = string
                        
                        # operate_data.teacher[2]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=16, column=j).value
                            ws_temp.cell(row=14, column=j + 1).value = string

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        # operate_data.teacher[0]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=13, column=j).value
                            ws_temp.cell(row=12, column=j + 1).value = string

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        # operate_data.teacher[0]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=13, column=j).value
                            ws_temp.cell(row=12, column=j + 1).value = string

                        # operate_data.teacher[1]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=14, column=j).value
                            ws_temp.cell(row=13, column=j + 1).value = string
                        
                        # operate_data.teacher[2]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=16, column=j).value
                            ws_temp.cell(row=14, column=j + 1).value = string

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        # operate_data.teacher[0]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=13, column=j).value
                            ws_temp.cell(row=12, column=j + 1).value = string

                        # operate_data.teacher[1]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=14, column=j).value
                            ws_temp.cell(row=13, column=j + 1).value = string
                        
                        # operate_data.teacher[2]
                        for j in range(1, 7):
                            string = ws_automation.cell(row=16, column=j).value
                            ws_temp.cell(row=14, column=j + 1).value = string

                    # 대체실습 기간
                    string = f"{automation.ws_members.cell(row=idx, column = 9).value[:4]} 년  {automation.ws_members.cell(row=idx, column = 9).value[5:7]} 월  {automation.ws_members.cell(row=idx, column = 9).value[8:]} 일  ∼    {automation.ws_members.cell(row=idx, column = 10).value[:4]} 년  {automation.ws_members.cell(row=idx, column = 10).value[5:7]} 월  {automation.ws_members.cell(row=idx, column = 10).value[8:]} 일"
                    ws_temp.cell(row=20, column=3).value = string

                    # 대체실습 시간
                    string = f"  총     {automation.ws_members.cell(row=idx, column=14).value}  시간"
                    ws_temp.cell(row=21, column=3).value = string

                    # 합격여부 
                    ws_temp.cell(row=22, column=3).value = "합격"

                    # 자체시험 점수
                    name = automation.ws_members.cell(row=idx, column=18).value
                    for cell in ws_temp_score["C"]:
                        if cell.value == name:
                            temp_row = cell.row
                    temp_score = ws_temp_score.cell(row=temp_row, column=7).value
                    if temp_score == None:
                        temp_score = 99
                    else:
                        pass
                    ws_temp.cell(row=22, column=6).value = temp_score

                    # 비고 

                    # 서명

                    # 수여일
                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        string = f"                                      {ws_automation.cell(row=3, column=3).value[:4]} 년   {ws_automation.cell(row=3, column=3).value[5:7]} 월    {ws_automation.cell(row=3, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        string = f"                                      {ws_automation.cell(row=4, column=3).value[:4]} 년   {ws_automation.cell(row=4, column=3).value[5:7]} 월    {ws_automation.cell(row=4, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        string = f"                                      {ws_automation.cell(row=5, column=3).value[:4]} 년   {ws_automation.cell(row=5, column=3).value[5:7]} 월    {ws_automation.cell(row=5, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        string = f"                                      {ws_automation.cell(row=6, column=3).value[:4]} 년   {ws_automation.cell(row=6, column=3).value[5:7]} 월    {ws_automation.cell(row=6, column=3).value[8:]} 일"
                    ws_temp.cell(row=27, column=1).value = string

                    wb_temp.save(string_stu)
                    wb_temp.close()
                    i += 1

            elif version == False:
                i = 0
                for idx, cell in enumerate(automation.ws_members["E"]):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    
                    fw_epr = pyautogui.getActiveWindow()
                    # 1번 학생
                    pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230 + (i * 21), duration=0.35)
                    # 큰 아이콘으로 만들기
                    pyautogui.sleep(0.5)
                    pyautogui.click(fw_epr.left + 175, fw_epr.top + 35, duration=0.35)
                    pyautogui.click(fw_epr.left + 311, fw_epr.top + 72, duration=0.35)
                    # 대체실습확인서 실행
                    pyautogui.doubleClick(fw_epr.left + 660, fw_epr.top + 242, duration=0.35)
                    # 한글 실행시간 대기
                        # pyautogui.sleep(10)
                        # pyautogui.click() 을 통해 한글 파일을 getActiveWindow() 로 가져올 수 있도록 함
                        # pyautogui.click()
                        # fw_completion = pyautogui.getActiveWindow()
                    fw_training = pyautogui.getWindowsWithTitle("대체실습확인서")
                    while fw_training == []:
                        fw_training = pyautogui.getWindowsWithTitle("대체실습확인서")
                    fw_training = pyautogui.getWindowsWithTitle("대체실습확인서")[0]
                    pyautogui.sleep(0.5)

                    # hotkey 를 통해 좌측으로 미룰 때, 이미 좌측에 있으면 맨 우측으로 넘어가는 것을 방지
                    if fw_training.isMaximized == False:
                        fw_training.maximize()
                    pyautogui.hotkey("win", "left")
                    pyautogui.press("esc")
                    pyautogui.sleep(1)
                    # 파일에 스크롤이 내려가 있을 수 있기 때문에 한번 쭉 올려준다
                    pyautogui.moveTo(fw_training.left + 440, fw_training.top + 500)
                    pyautogui.scroll(5000)

                    # 이름
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
                    pyautogui.click(fw_training.left + 215, fw_training.top + 475, duration=0.35)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 생년월일
                    string = f"{automation.ws_members.cell(row=idx, column=20).value[:2]}. {automation.ws_members.cell(row=idx, column=20).value[2:4]}. {automation.ws_members.cell(row=idx, column=20).value[4:6]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 연락처
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=19).value)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육기관명
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=3).value)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육과정명
                    string = f" 요양보호사 {automation.ws_members.cell(row=idx, column=4).value}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    # # 글씨체 변경 + 가운데 정렬
                        # # 폰트 : 615,150 / 맑은고딕 : 710,460
                        # for i in range(2):
                        #     pyautogui.press("F5")
                        # for i in range(4):
                        #     pyautogui.press("left")
                        # pyautogui.hotkey("ctrl", "shift", "c")
                        # pyautogui.click(fw_training.left + 615, fw_training.top + 150, duration=0.35)
                        # pyautogui.moveTo(fw_training.left + 710, fw_training.top + 460, duration=0.35)
                        # 스크롤이 너무 느림 ,,,
                        # pyautogui.scroll(10000)
                        # pyautogui.sleep(0.5)
                        # pyautogui.click()

                    # 가리기(대체실습 이수자 기본사항 mouse.ver):
                        # # 생년월일
                        # string = "{}" .format(automation.ws_members.cell(row=idx, column=23).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_training.left + 305, fw_training.top + 475)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 교육기관명
                        # string = "{}" .format(automation.ws_members.cell(row=idx, column=3).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_training.left + 565, fw_training.top + 475)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                        # # 교육과정명
                        # string = " 요양보호사 {}" .format(automation.ws_members.cell(row=idx, column=4).value)
                        # pyperclip.copy(string)
                        # pyautogui.click(fw_training.left + 685, fw_training.top + 475)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")

                    # 대체실습 교육지도자 줄 수는 3줄을 기준으로 한다. / but 3기 주간부터 수정할때는 3개 기준으로 한다 + 오류파일들 적어놓고 나중에 수정
                    # 대체실습 교육지도자 / if 가리기(구분 : 대체실습 교육지도자 keyboard.fer)
                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        # operate_data.teacher[1]
                        pyautogui.click(fw_training.left + 215, fw_training.top + 610, duration=0.35)
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=14, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")  

                        # 가리기(아래 내용을 for 문으로 넣음, 2기 ~ 도 마찬가지):
                            # 구분
                            # pyperclip.copy(ws_automation.cell(row=17, column=1).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")

                            # # 이름
                            # pyperclip.copy(ws_automation.cell(row=17, column=2).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")

                            # # 생년월일
                            # pyperclip.copy(ws_automation.cell(row=17, column=3).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")

                            # #자격사항
                            # pyperclip.copy(ws_automation.cell(row=17, column=4).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")

                            # # 경력 기간
                            # pyperclip.copy(ws_automation.cell(row=17, column=5).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")

                            # # 도 승인일자
                            # pyperclip.copy(ws_automation.cell(row=17, column=6).value)
                            # pyautogui.hotkey("ctrl", "a")
                            # pyautogui.hotkey("ctrl", "v")
                            # pyautogui.hotkey("alt", "right")
                        # operate_data.teacher[4]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=17, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")
                        
                        # operate_data.teacher[2]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=16, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        # operate_data.teacher[0]
                        pyautogui.click(fw_training.left + 215, fw_training.top + 610, duration=0.35)
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=13, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")
                        
                        # 남은 줄(빈칸) 삭제 (대체실습 2기는 교육지도자가 1명)
                        for i in range(2):
                            pyautogui.press("F5")
                        for i in range(5):
                            pyautogui.press("right")
                        pyautogui.press("down")
                        pyautogui.press("delete")
                        pyautogui.press("n")

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        # operate_data.teacher[0]
                        pyautogui.click(fw_training.left + 215, fw_training.top + 610, duration=0.35)
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=13, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")

                        # operate_data.teacher[1]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=14, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")
                        
                        # operate_data.teacher[2]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=16, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")

                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        # operate_data.teacher[0]
                        pyautogui.click(fw_training.left + 215, fw_training.top + 610, duration=0.35)
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=13, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")

                        # operate_data.teacher[1]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=14, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")
                        
                        # operate_data.teacher[2]
                        for i in range(1, 7):
                            pyperclip.copy(ws_automation.cell(row=16, column=i).value)
                            pyautogui.hotkey("ctrl", "a")
                            pyautogui.hotkey("ctrl", "v")
                            pyautogui.hotkey("alt", "right")

                        #대체실습 기간 위치를 맞추기 위한 작업
                        # pyautogui.press("down")
                        # pyautogui.press("end")
                        # pyautogui.press("enter")
                    
                    # 대체실습 기간
                    pyautogui.click(fw_training.left + 675, fw_training.top + 785, duration=0.35)
                    string = f"{automation.ws_members.cell(row=idx, column = 9).value[:4]} 년  {automation.ws_members.cell(row=idx, column = 9).value[5:7]} 월  \
                        {automation.ws_members.cell(row=idx, column = 9).value[8:]} 일  ∼    {automation.ws_members.cell(row=idx, column = 10).value[:4]} 년  \
                            {automation.ws_members.cell(row=idx, column = 10).value[5:7]} 월  {automation.ws_members.cell(row=idx, column = 10).value[8:]} 일"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    # 대체실습 시간
                    # pyautogui.click(fw_training.left + 680, fw_training.top + 825, duration=0.35)
                    pyautogui.hotkey("alt", "right")
                    pyautogui.hotkey("alt", "right")
                    string = f"  총     {automation.ws_members.cell(row=idx, column=14).value}  시간"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")

                    # 합격여부 
                    pyautogui.hotkey("alt", "right")
                    pyautogui.hotkey("alt", "right")
                    pyperclip.copy("합격")
                    pyautogui.hotkey("ctrl", "v")

                    # 자체시험 점수
                    pyautogui.hotkey("alt", "right")
                    pyautogui.hotkey("alt", "right")
                    name = automation.ws_members.cell(row=idx, column=18).value
                    for cell in ws_temp_score["C"]:
                        if cell.value == name:
                            temp_row = cell.row
                    temp_score = ws_temp_score.cell(row=temp_row, column=7).value
                    if temp_score == None:
                        temp_score = 99
                    else:
                        pass
                    pyperclip.copy(temp_score)
                    pyautogui.hotkey("ctrl", "v")

                    # 비고 
                    pyautogui.hotkey("alt", "right")
                    pyautogui.hotkey("alt", "right")
                    
                    # 서명
                    pyautogui.hotkey("alt", "right")
                    pyautogui.hotkey("alt", "right")
                    pyperclip.copy(operate_data.teacher[0])
                    pyautogui.hotkey("ctrl", "v")

                    # 수여일
                    if automation.ws_members.cell(row=idx, column=8).value == "대체실습 1기":
                        string = f"                                      {ws_automation.cell(row=3, column=3).value[:4]} 년   \
                            {ws_automation.cell(row=3, column=3).value[5:7]} 월    {ws_automation.cell(row=3, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 2기":
                        string = f"                                      {ws_automation.cell(row=4, column=3).value[:4]} 년   \
                            {ws_automation.cell(row=4, column=3).value[5:7]} 월    {ws_automation.cell(row=4, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 3기":
                        string = f"                                      {ws_automation.cell(row=5, column=3).value[:4]} 년   \
                            {ws_automation.cell(row=5, column=3).value[5:7]} 월    {ws_automation.cell(row=5, column=3).value[8:]} 일"
                    elif automation.ws_members.cell(row=idx, column=8).value == "대체실습 4기":
                        string = f"                                      {ws_automation.cell(row=6, column=3).value[:4]} 년   \
                            {ws_automation.cell(row=6, column=3).value[5:7]} 월    {ws_automation.cell(row=6, column=3).value[8:]} 일"
                    pyperclip.copy(string)
                    pyautogui.scroll(-2000)
                    pyautogui.moveTo(fw_training.left + 705, fw_training.top + 780, duration=0.35)
                    for i in range(3):
                        pyautogui.click()
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.press("enter")

                    # win + left 된 상태에서 닫기버튼을 누른 후 저장에 enter
                    pyautogui.click(fw_training.left + 940, fw_training.top + 12)
                    pyautogui.press("enter")

                    # 폴더 뒤로가기
                    pyautogui.sleep(1.5)
                    pyautogui.press("backspace")
                    fw_epr = pyautogui.getActiveWindow()
                    print(fw_epr)
                    pyautogui.moveTo(fw_epr.left + 20, fw_epr.top + 165, duration=0.35)

                    i += 1

        elif task == "요양보호사 자격증 발급,재발급 신청서":
            if version == True:
                i = 1
                for idx, cell in enumerate(automation.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue
                    # !!!caution!!! f-string 사용 시에는 \enter(줄바꿈) 사용하면 X ! 그대로 입력됨
                    if time == "주간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=2).value
                    elif time == "야간":
                        startingdate = ws_automation.cell(row=112 + ordinal_num, column=3).value
                    if automation.ws_members.cell(row=idx, column=15).value == "일반":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간조\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_사복\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                        string_stu = f"D:\\"+operate_data.ac_name+f"\\교육생관리\\{ordinal_num}기\\{ordinal_num}기{time}{startingdate}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간호\\{automation.ws_members.cell(row=idx, column=18).value}_{task}.xlsx"
                    print(string_stu)
                    wb_certificate = load_workbook(string_stu)
                    ws_certificate = wb_certificate.active

                    # 요양보호사 자격증 발급,재발급 신청서
                    # 이름
                    string = f"성명(한자)   {automation.ws_members.cell(row=idx, column=18).value}"
                    ws_certificate.cell(row=6, column=2).value = string

                    # 주민등록번호
                    string = f"주민등록번호  {automation.ws_members.cell(row=idx, column=20).value}"
                    ws_certificate.cell(row=7, column=2).value = string

                    # 주소
                    string = f"주소   {automation.ws_members.cell(row=idx, column=21).value}"
                    ws_certificate.cell(row=8, column=2).value = string

                    # 전화번호
                    string = f"전화번호  {automation.ws_members.cell(row=idx, column=19).value}"
                    ws_certificate.cell(row=9, column=2).value = string

                    # 요양보호사 교육기간. 부터
                    string = f"{automation.ws_members.cell(row=idx, column=6).value[2:4]}.{automation.ws_members.cell(row=idx, column=6).value[5:7]}.{automation.ws_members.cell(row=idx, column=6).value[8:]}"
                    ws_certificate.cell(row=12, column=2).value = string

                    # 요양보호사 교육기간. 까지
                    string = f"{automation.ws_members.cell(row=idx, column=7).value[2:4]}.{automation.ws_members.cell(row=idx, column=7).value[5:7]}.{automation.ws_members.cell(row=idx, column=7).value[8:]}"
                    ws_certificate.cell(row=12, column=3).value = string

                    # 교육과정명
                    string = f"요양보호사 {automation.ws_members.cell(row=idx, column=5).value[0]}기 (이론,실기)"
                    ws_certificate.cell(row=12, column=4).value = string

                    # 교육기관명
                    string = "남양노아요양보호사교육원"
                    ws_certificate.cell(row=12, column=7).value = string

                    # 요양보호사 교육기간(실습). 부터
                    string = f"{automation.ws_members.cell(row=idx, column=9).value[2:4]}.{automation.ws_members.cell(row=idx, column=9).value[5:7]}.{automation.ws_members.cell(row=idx, column=9).value[8:]}"
                    ws_certificate.cell(row=13, column=2).value = string

                    # 요양보호사 교육기간(실습). 까지
                    string = f"{automation.ws_members.cell(row=idx, column=10).value[2:4]}.{automation.ws_members.cell(row=idx, column=10).value[5:7]}.{automation.ws_members.cell(row=idx, column=10).value[8:]}"
                    ws_certificate.cell(row=13, column=3).value = string

                    # 교육과정명(실습)
                    string = f"요양보호사 (대체실습{automation.ws_members.cell(row=idx, column=8).value[5]}기)"
                    ws_certificate.cell(row=13, column=4).value = string

                    # 교육기관명(실습)
                    string = "남양노아요양보호사교육원"
                    ws_certificate.cell(row=13, column=7).value = string

                    # 시험 시행일
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "시험시행일   2021년 05월 15일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "시험시행일   2021년 02월 20일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "시험시행일   2021년 05월 15일"
                    ws_certificate.cell(row=14, column=2).value = string

                    # 시험 합격일
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "시험합격일   2021년 06월 01일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "시험합격일   2021년 03월 09일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "시험합격일   2021년 06월 01일"
                    ws_certificate.cell(row=14, column=5).value = string

                    # 신청 일자
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "     2021  년     03  월    15   일    "
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "     2021  년     03  월    15   일    "
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "     2021  년     03  월    15   일    "
                    ws_certificate.cell(row=19, column=1).value = string

                    # 이름 / shift 는 keyDown(or Up) 에서 left 와 right 를 모두 입력해 주어야 정상작동 함 !!
                    string = f"{automation.ws_members.cell(row=idx, column=18).value} (서명 또는 인)"
                    ws_certificate.cell(row=20, column=4).value = string

                    wb_certificate.save(string_stu)
                    wb_certificate.close()
                    i += 1

            elif version == False:
                i = 0
                for idx, cell in enumerate(automation.ws_members["E"], start=1):
                    if not string_set in str(cell.value):
                        continue
                    if idx <= 4:
                        continue

                    fw_epr = pyautogui.getActiveWindow()

                    # 1번 학생
                    pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230 + (i * 21), duration=0.35)
                    # 큰 아이콘으로 만들기
                    pyautogui.sleep(0.5)
                    pyautogui.click(fw_epr.left + 175, fw_epr.top + 35, duration=0.35)
                    pyautogui.click(fw_epr.left + 311, fw_epr.top + 72, duration=0.35)
                    # 자격증 발급 신청서 실행
                    pyautogui.doubleClick(fw_epr.left + 770, fw_epr.top + 242, duration=0.35)

                    fw_certificate = pyautogui.getWindowsWithTitle("자격증 발급,재발급 신청서")
                    while fw_certificate == []:
                        fw_certificate = pyautogui.getWindowsWithTitle("자격증 발급,재발급 신청서")
                    fw_certificate = pyautogui.getWindowsWithTitle("자격증 발급,재발급 신청서")[0]
                    pyautogui.sleep(0.5)

                    # hotkey 를 통해 좌측으로 미룰 때, 이미 좌측에 있으면 맨 우측으로 넘어가는 것을 방지
                    if fw_certificate.isMaximized == False:
                        fw_certificate.maximize()
                    pyautogui.hotkey("win", "left")
                    pyautogui.press("esc")
                    pyautogui.sleep(1)
                    # 파일에 스크롤이 내려가 있을 수 있기 때문에 한번 쭉 올려준다
                    pyautogui.moveTo(fw_certificate.left + 440, fw_certificate.top + 500)
                    pyautogui.scroll(5000)

                    # 글씨체가 변경됨 !!!
                        # # 이름
                        # pyautogui.click(fw_certificate.left + 600, fw_certificate.top + 420, duration=0.35)
                        # string = "성명(한자)   {}" .format(automation.ws_members.cell(row=idx, column=18).value)
                        # pyperclip.copy(string)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")
                        # pyautogui.hotkey("alt", "right")

                        # for j in range(2):
                        #     pyautogui.hotkey("alt", "right")

                        # # 주민등록번호
                        # string = "주민등록번호  {}" .format(automation.ws_members.cell(row=idx, column=20).value)
                        # pyperclip.copy(string)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")
                        # pyautogui.hotkey("alt", "right")

                        # for j in range(2):
                        #     pyautogui.hotkey("alt", "right")

                        # # 주소
                        # string = "주소   {}" .format(automation.ws_members.cell(row=idx, column=21).value)
                        # pyperclip.copy(string)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")
                        # pyautogui.hotkey("alt", "right")

                        # for j in range(2):
                        #     pyautogui.hotkey("alt", "right")

                        # # 전화번호
                        # string = "전화번호  {}" .format(automation.ws_members.cell(row=idx, column=19).value)
                        # pyperclip.copy(string)
                        # pyautogui.hotkey("ctrl", "a")
                        # pyautogui.hotkey("ctrl", "v")
                        # pyautogui.hotkey("alt", "right")

                        # for j in range(8):
                        #     pyautogui.hotkey("alt", "right")

                    # 글씨체 변경 방지를 위한 업데이트
                    # 이름
                    pyautogui.click(fw_certificate.left + 600, fw_certificate.top + 420, duration=0.35)
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)

                    # 글씨체 변경 X (but 처리할 내용이 많음)
                    # pyautogui.hotkey("ctrl", "a") ->
                    # pyautogui.press("end"), ~ pyautogui.keyUp("shiftright")
                    pyautogui.press("end")
                    pyautogui.keyDown("shiftleft")
                    pyautogui.keyDown("shiftright")

                    pyautogui.press("home")
                    pyautogui.keyDown("ctrl")
                    pyautogui.press("right")

                    pyautogui.keyUp("ctrl")
                    pyautogui.keyUp("shiftleft")
                    pyautogui.keyUp("shiftright")

                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    for j in range(2):
                        pyautogui.hotkey("alt", "right")

                    # 주민등록번호
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=20).value)
                    pyautogui.press("end")
                    pyautogui.keyDown("shiftleft")
                    pyautogui.keyDown("shiftright")

                    pyautogui.press("home")
                    pyautogui.keyDown("ctrl")
                    pyautogui.press("right")

                    pyautogui.keyUp("ctrl")
                    pyautogui.keyUp("shiftleft")
                    pyautogui.keyUp("shiftright")

                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    for j in range(2):
                        pyautogui.hotkey("alt", "right")

                    # 주소
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=21).value)
                    pyautogui.press("end")
                    pyautogui.keyDown("shiftleft")
                    pyautogui.keyDown("shiftright")

                    pyautogui.press("home")
                    pyautogui.keyDown("ctrl")
                    pyautogui.press("right")

                    pyautogui.keyUp("ctrl")
                    pyautogui.keyUp("shiftleft")
                    pyautogui.keyUp("shiftright")

                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    for j in range(2):
                        pyautogui.hotkey("alt", "right")

                    # 전화번호
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=19).value)
                    pyautogui.press("end")
                    pyautogui.keyDown("shiftleft")
                    pyautogui.keyDown("shiftright")

                    pyautogui.press("home")
                    pyautogui.keyDown("ctrl")
                    pyautogui.press("right")

                    pyautogui.keyUp("ctrl")
                    pyautogui.keyUp("shiftleft")
                    pyautogui.keyUp("shiftright")

                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    for j in range(8):
                        pyautogui.hotkey("alt", "right")

                    # 요양보호사 교육기간. 부터
                    string = f"{automation.ws_members.cell(row=idx, column=6).value[2:4]}.{automation.ws_members.cell(row=idx, column=6).value[5:7]}.\
                        {automation.ws_members.cell(row=idx, column=6).value[8:]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 요양보호사 교육기간. 까지
                    string = f"{automation.ws_members.cell(row=idx, column=7).value[2:4]}.{automation.ws_members.cell(row=idx, column=7).value[5:7]}.\
                        {automation.ws_members.cell(row=idx, column=7).value[8:]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육과정명
                    string = f"요양보호사 {automation.ws_members.cell(row=idx, column=5).value[0]}기 (이론,실기)"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육기관명
                    string = operate_data.ac_name
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    pyautogui.hotkey("alt", "right")
                    
                    # 요양보호사 교육기간(실습). 부터
                    string = f"{automation.ws_members.cell(row=idx, column=9).value[2:4]}.{automation.ws_members.cell(row=idx, column=9).value[5:7]}.\
                        {automation.ws_members.cell(row=idx, column=9).value[8:]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 요양보호사 교육기간(실습). 까지
                    string = f"{automation.ws_members.cell(row=idx, column=10).value[2:4]}.{automation.ws_members.cell(row=idx, column=10).value[5:7]}.\
                        {automation.ws_members.cell(row=idx, column=10).value[8:]}"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육과정명(실습)
                    string = f"요양보호사 (대체실습{automation.ws_members.cell(row=idx, column=8).value[5]}기)"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 교육기관명(실습)
                    string = operate_data.ac_name
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    pyautogui.hotkey("alt", "right")

                    # 시험 시행일
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "시험시행일   2021년 05월 15일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "시험시행일   2021년 02월 20일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "시험시행일   2021년 05월 15일"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    # 시험 합격일
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "시험합격일   2021년 06월 01일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "시험합격일   2021년 03월 09일"
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "시험합격일   2021년 06월 01일"
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    for j in range(5):
                        pyautogui.hotkey("alt", "right")

                    # 신청 일자
                    if automation.ws_members.cell(row=idx, column=18).value == "오연숙":
                        string = "     2021  년     03  월    15   일    "
                    elif automation.ws_members.cell(row=idx, column=5).value == "3기주간1019" or "3기야간1116" or "4기주간1207":
                        string = "     2021  년     03  월    15   일    "
                    elif automation.ws_members.cell(row=idx, column=5).value == "4기야간0201" or "5기주간0201":
                        string = "     2021  년     03  월    15   일    "
                    pyperclip.copy(string)
                    pyautogui.hotkey("ctrl", "a")
                    pyautogui.hotkey("ctrl", "v")
                    pyautogui.hotkey("alt", "right")

                    pyautogui.hotkey("alt", "right")

                    # 이름 / shift 는 keyDown(or Up) 에서 left 와 right 를 모두 입력해 주어야 정상작동 함 !!
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
                    pyautogui.keyDown("shiftleft")
                    pyautogui.keyDown("shiftright")
                    for j in range(3):
                        pyautogui.press("right")
                    pyautogui.keyUp("shiftleft")
                    pyautogui.keyUp("shiftright")
                    pyautogui.hotkey("ctrl", "v")

                    # win + left 된 상태에서 닫기버튼을 누른 후 저장에 enter
                    pyautogui.click(fw_certificate.left + 940, fw_certificate.top + 12)
                    pyautogui.press("enter")

                    # 폴더 뒤로가기
                    pyautogui.sleep(1.5)
                    pyautogui.press("backspace")
                    fw_epr = pyautogui.getActiveWindow()
                    print(fw_epr)
                    pyautogui.moveTo(fw_epr.left + 20, fw_epr.top + 165, duration=0.35)

                    i += 1

        automation.wb_members.close()

    def automation_task_report(self, ordinal_num, time, kind):
        # 사용방법 x.automation_task_report(3, "주간", "개강보고")
        self.ordinal_num = ordinal_num
        self.time = time
        self.kind = kind # kind = 개강보고

        is_ready = pyautogui.confirm("자동화 프로그램을 시작하시기 전에\n1. 모든 각 문서는 D:\\Master\\rpa_basic_file 에서 복사된 파일이어야 합니다.\n(각 자료가 입력되어 있어야 함)\n2. 한글, 파일 탐색기(폴더) 가 하나라도 실행되어 있어선 안됩니다.\n준비가 되었으면 확인, 안되어있으면 취소를 눌러주세요.", "경고")
        if is_ready == "OK":
            pyautogui.alert("자동화 프로그램을 시작합니다.\n프로그램을 강제로 종료하고 싶으실 경우 마우스를 각 꼭짓점에 가져가주세요.", "실행")
        elif is_ready == "Cancel":
            pyautogui.alert("프로그램을 종료합니다.\n준비작업을 마치신 후 다시 실행해 주세요.", "종료")
            return

        # 4기야갼 개강보고 자동화
        # 파일 경로 이동
        fw_epr.activate()
        # 4기보고
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230 + ((ordinal_num - 1) * 21), duration=0.35)
        # 4기야간 보고서 실행
        if time == "주간" :
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 314, duration=0.35)
        elif time == "야간":
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 293, duration=0.35)
        title = f"{ordinal_num}기{time}{kind}"
        fw_report = pyautogui.getWindowsWithTitle(title)
        while fw_report == []:
            fw_report = pyautogui.getWindowsWithTitle(title)
        fw_report = pyautogui.getWindowsWithTitle(title)[0]
        pyautogui.sleep(0.5)

        if fw_report.isMaximized == False:
            fw_report.maximize()
        
        # 자동화 셋팅
        string_set = f"{ordinal_num}기{time}"
        # 중앙으로 이동후 스크롤바 위치 표준화
        pyautogui.moveTo(fw_report.center.x, fw_report.center.y, duration=0.35)
        pyautogui.sleep(1)
        pyautogui.scroll(10000)
        pyautogui.sleep(1)

        # 과정구분 1095,360
        pyautogui.click(fw_report.left + 8 + 1085, fw_report.top + 8 + 345, duration=0.35)
        pyautogui.sleep(1)
        for idx, cell in enumerate(automation.ws_members["E"], start=1):
            # print("시작한다. ", idx)      지렸다 ,, 니네가 날 살렸다 ,, 사랑한다 ,,
            # print(string, cell.value)
            if not string_set in str(cell.value):
                continue

            # 과정구분
            if automation.ws_members.cell(row=idx, column=15).value == "일반":
                string = f"{automation.ws_members.cell(row=idx, column=15).value}\n(신규)"
            else:
                string = f"{automation.ws_members.cell(row=idx, column=15).value[:3]}\n{automation.ws_members.cell(row=idx, column=15).value[3:]}"
            pyperclip.copy(string)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")


            # 이름
            pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            # 생년월일
            string = f"{automation.ws_members.cell(row=idx, column=20).value[:2]}. {automation.ws_members.cell(row=idx, column=20).value[2:4]}. \
                {automation.ws_members.cell(row=idx, column=20).value[4:6]}"
            pyperclip.copy(string)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            # 주소
            pyperclip.copy(automation.ws_members.cell(row=idx, column=21).value)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            # 연락처
            pyperclip.copy(automation.ws_members.cell(row=idx, column=19).value)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            # 비고
            pyautogui.hotkey("alt", "right")
            pyautogui.hotkey("alt", "right")

        automation.wb_members.close()
        

    def automation_task_temporary(self, ordinal_num, kind):
        # 사용방법 : x.automation_task_temporary(3, "실시보고")
        self.ordinal_num = ordinal_num
        self.kind = kind
        is_ready = pyautogui.confirm("자동화 프로그램을 시작하시기 전에\n1. 모든 각 문서는 D:\\Master\\rpa_basic_file 에서 복사된 파일이어야 합니다.\n(각 자료가 입력되어 있어야 함)\n2. 한글, 파일 탐색기(폴더) 가 하나라도 실행되어 있어선 안됩니다.\n준비가 되었으면 확인, 안되어있으면 취소를 눌러주세요.", "경고")
        if is_ready == "OK":
            pyautogui.alert("자동화 프로그램을 시작합니다.\n프로그램을 강제로 종료하고 싶으실 경우 마우스를 각 꼭짓점에 가져가주세요.", "실행")
        elif is_ready == "Cancel":
            pyautogui.alert("프로그램을 종료합니다.\n준비작업을 마치신 후 다시 실행해 주세요.", "종료")
            return

        fw_epr.activate()

        title = f"대체실습 {ordinal_num}기 교육{kind}"

        # 파일 경로 이동
        # 대체실습
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 356, duration=0.35)
        # 실시 수료 보고
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 272, duration=0.35)
        if kind == "실시보고":
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + (230 + ((ordinal_num - 1) * 21)), duration=0.35)
        elif kind == "수료보고":
            # 대체실습 0기 수료보고
            pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + (356 + ((ordinal_num - 1) * 21)), duration=0.35)

        fw_temporary = pyautogui.getWindowsWithTitle(title)
        while fw_temporary == []:
            fw_temporary = pyautogui.getWindowsWithTitle(title)
        fw_temporary = pyautogui.getWindowsWithTitle(title)[0]
        pyautogui.sleep(0.5)
        
        if fw_temporary.isMaximized == False:
            fw_temporary.maximize()

        pyautogui.moveTo(fw_temporary.center.x, fw_temporary.center.y, duration=0.35)
        pyautogui.scroll(10000)
        pyautogui.sleep(0.5)

        if kind == "실시보고":
            # 실시보고는 명단이 중아에 있기 때문에 맞추기 위한 스크롤 내리기
            pyautogui.scroll(-2500) # 한 틱당 약 125
            # [교육생명단] 교육구분 370,395
            pyautogui.click(fw_temporary.top + 8 + 370, fw_temporary.left + 8 + 395, duration=0.35)
            for idx, cell in enumerate(automation.ws_members["H"]):
                if cell.value != f"대체실습 {ordinal_num}기":
                    continue

                # 교육구분
                if automation.ws_members.cell(row=idx, column=15).value == "일반":
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=15).value)
                elif "자격증" in str(automation.ws_members.cell(row=idx, column=15).value):
                    pyperclip.copy(automation.ws_members.cell(row=idx, column=15).value[:3])
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 성명
                pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 생년월일
                string = f"{automation.ws_members.cell(row=idx, column=20).value[:2]}. {automation.ws_members.cell(row=idx, column=20).value[2:4]}. \
                    {automation.ws_members.cell(row=idx, column=20).value[4:6]}"
                pyperclip.copy(string)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 연락처
                pyperclip.copy(automation.ws_members.cell(row=idx, column=19).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 교육기수
                string = f"{automation.ws_members.cell(row=idx, column=5).value[:2]} {automation.ws_members.cell(row=idx, column=5).value[2:4]}반"
                pyperclip.copy(string)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 이론, 실기 교육이수일
                string = f"{automation.ws_members.cell(row=idx, column=7).value[2:4]}.{automation.ws_members.cell(row=idx, column=7).value[5:7]}.\
                    {automation.ws_members.cell(row=idx, column=7).value[8:]}"
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 대체실습 필요시간
                if automation.ws_members.cell(row=idx, column=15).value == "일반":
                    pyperclip.copy("80시간")
                elif "자격증" in str(automation.ws_members.cell(row=idx, column=15).value):
                    pyperclip.copy("8시간")
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 순번 넘기기
                pyautogui.hotkey("alt", "right")
                

        elif kind == "수료보고":
            cnt = 0
            number = 1
            # 제목 615,450 / 2페이지 연번 915,600 / 18번 스크롤(2250) 한 후 3페이지 연번(11) 140,460 / 4페이지 연번(25) 1220,460 / 14번 스크롤 한 후 5페이지 연번(39) 140,430 / 

            # 연번 클릭
            pyautogui.click(fw_temporary.left + 8 + 915, fw_temporary.top + 8 + 600, duration=0.35)
            for idx, cell in enumerate(automation.ws_members["H"], start=1):
                if cell.value != f"대체실습 {ordinal_num}기":
                    continue

                # 연번
                pyperclip.copy(number)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 교육과정 명
                string = f"{automation.ws_members.cell(row=idx, column=5).value[2:4]}반\n{automation.ws_members.cell(row=idx, column=5).value[:2]}"
                pyperclip.copy(string)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 이수시간(총 시간)
                pyperclip.copy(automation.ws_members.cell(row=idx, column=11).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 이수시간(이론)
                pyperclip.copy(automation.ws_members.cell(row=idx, column=12).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 이수시간(실기)
                pyperclip.copy(automation.ws_members.cell(row=idx, column=13).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 이수시간(실습, 공백)
                pyautogui.hotkey("alt", "right")

                # 이수시간(대체실습)
                pyperclip.copy(automation.ws_members.cell(row=idx, column=14).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 성명
                pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 주민번호
                string = f"{automation.ws_members.cell(row=idx, column=20).value[:6]}\n{automation.ws_members.cell(row=idx, column=20).value[6:]}"
                pyperclip.copy(string)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 주소(도로명)
                pyperclip.copy(automation.ws_members.cell(row=idx, column=22).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                # 연락처
                pyperclip.copy(automation.ws_members.cell(row=idx, column=19).value)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                #수료 연월일
                string = f"{automation.ws_members.cell(row=idx, column=10).value[2:4]}.{automation.ws_members.cell(row=idx, column=10).value[5:7]}.\
                    {automation.ws_members.cell(row=idx, column=10).value[8:]}"
                pyperclip.copy(string)
                pyautogui.hotkey("ctrl", "a")
                pyautogui.hotkey("ctrl", "v")
                pyautogui.hotkey("alt", "right")

                #비고
                pyautogui.hotkey("alt", "right")

                number += 1
                cnt += 1

                if cnt == 10:
                    pyautogui.scroll(-2250)
                    pyautogui.click(fw_temporary.left + 8 + 140, fw_temporary.top + 8 + 435, duration=0.35)
                
                if cnt == 25:
                    pyautogui.click(fw_temporary.left + 8 + 1220, fw_temporary.top + 8 + 435, duration=0.35)

        automation.wb_members.close()

        # if task == "bigging_class":
        # if task == "certificate":

    
    def mkattendance(self, ordinal_num, time):
        self.ordinal_num = ordinal_num
        self.time = time

        is_ready = pyautogui.confirm("자동화 프로그램을 시작하시기 전에\n1. 모든 각 문서는 D:\\Master\\rpa_basic_file 에서 복사된 파일이어야 합니다.\n(각 자료가 입력되어 있어야 함)\n2. 한글, 파일 탐색기(폴더) 가 하나라도 실행되어 있어선 안됩니다.\n준비가 되었으면 확인, 안되어있으면 취소를 눌러주세요.", "경고")
        if is_ready == "OK":
            pyautogui.alert("자동화 프로그램을 시작합니다.\n프로그램을 강제로 종료하고 싶으실 경우 마우스를 각 꼭짓점에 가져가주세요.", "실행")
        elif is_ready == "Cancel":
            pyautogui.alert("프로그램을 종료합니다.\n준비작업을 마치신 후 다시 실행해 주세요.", "종료")
            return

        fw_epr.activate()

        number = 1

        string_set = f"{ordinal_num}기{time}"

        # 기수 폴더 클릭
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 230 + ((ordinal_num - 1) * 21), duration=0.35)

        # 출석부 폴더
        pyautogui.doubleClick(fw_epr.left + 330, fw_epr.top + 335, duration=0.35)

        fw_attendance = pyautogui.getWindowsWithTitle("출석부")
        while fw_attendance == []:
            fw_attendance = pyautogui.getWindowsWithTitle("출석부")
        fw_attendance = pyautogui.getWindowsWithTitle("출석부")[0]
        pyautogui.sleep(1)

        if fw_attendance.isMaximized == False:
            fw_attendance.maximize()

        pyautogui.moveTo(fw_attendance.center.x, fw_attendance.center.y, duration=0.35)
        pyautogui.scroll(10000)
        pyautogui.sleep(0.5)

        if time == "주간":
            pyautogui.click(fw_attendance.top + 8 + 275, fw_attendance.top + 8 + 470)

        elif time == "야간":
            pyautogui.click(fw_attendance.top + 8 + 1040, fw_attendance.top + 8 + 475)

        pyautogui.sleep(0.5)

        for idx, cell in enumerate(automation.ws_members["E"], start=1):
            if not string_set in str(cell.value):
                continue

            pyperclip.copy(number)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            pyperclip.copy(automation.ws_members.cell(row=idx, column=18).value)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            string = f"{automation.ws_members.cell(row=idx, column=20).value[:2]}. {automation.ws_members.cell(row=idx, column=20).value[2:4]}.\
                 {automation.ws_members.cell(row=idx, column=20).value[4:6]}"
            pyperclip.copy(string)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.hotkey("ctrl", "v")
            pyautogui.hotkey("alt", "right")

            if time == "주간":
                for i in range(7):
                    pyautogui.hotkey("alt", "right")
            
            elif time == "야간":
                for i in range(4):
                    pyautogui.hotkey("alt", "right")
            
            number += 1

        automation.wb_members.close()

    def update_attendance(self, update_num, update_time):
        # how to use ? : update_attendance(4, "야간") 업데이트를 진행할 기수 + 시간
        name = f"{update_num}기{update_time}_출석부(기관장)"
        wb_update = load_workbook("D:\\"+operate_data.ac_name+f"\\교육생관리\\출석부_기관장용\\{name}.xlsx")
        ws_update = wb_update.active

        string_set = f"{update_num}기{update_time}"

        # 출석부 파일의 시작점과 각 결석일수를 참조하기 위한 초기화 작업
        if update_time == "주간":
                num = 7
                start_set = 3
        elif update_time == "야간":
            num = 4
            start_set = 0
        member = 1

        for idx, cell in enumerate(automation.ws_members["E"], start=1):
            if not string_set in str(cell.value):
                continue

            # 실습 시간이 안적혀 있는 경우, 0시간으로 초기화하여 오류 방지
            if automation.ws_members.cell(row=idx, column=14).value == None:
                automation.ws_members.cell(row=idx, column=14).value = 0
            
            # 출석부 결석시간이 안적혀 있는 경우, 0시간으로 초기화하여 오류 방지
            if ws_update.cell(row=(member * num) - start_set, column=4).value == None:
                ws_update.cell(row=(member * num) - start_set, column=4).value = 0
            if ws_update.cell(row=(member * num) - start_set, column=5).value == None:
                ws_update.cell(row=(member * num) - start_set, column=5).value = 0
            
            if automation.ws_members.cell(row=idx, column=15).value == "일반":
                automation.ws_members.cell(row=idx, column=12).value = 80 - ws_update.cell(row=(member * num) - start_set, column=4).value
                automation.ws_members.cell(row=idx, column=13).value = 80 - ws_update.cell(row=(member * num) - start_set, column=5).value
                automation.ws_members.cell(row=idx, column=11).value = automation.ws_members.cell(row=idx, column=12).value + automation.ws_members.cell(row=idx, column=13).value + automation.ws_members.cell(row=idx, column=14).value
            
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                automation.ws_members.cell(row=idx, column=12).value = 31 - ws_update.cell(row=(member * num) - start_set, column=4).value
                automation.ws_members.cell(row=idx, column=13).value = 11 - ws_update.cell(row=(member * num) - start_set, column=5).value
                automation.ws_members.cell(row=idx, column=11).value = automation.ws_members.cell(row=idx, column=12).value + automation.ws_members.cell(row=idx, column=13).value + automation.ws_members.cell(row=idx, column=14).value
            
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                automation.ws_members.cell(row=idx, column=12).value = 32 - ws_update.cell(row=(member * num) - start_set, column=4).value
                automation.ws_members.cell(row=idx, column=13).value = 10 - ws_update.cell(row=(member * num) - start_set, column=5).value
                automation.ws_members.cell(row=idx, column=11).value = automation.ws_members.cell(row=idx, column=12).value + automation.ws_members.cell(row=idx, column=13).value + automation.ws_members.cell(row=idx, column=14).value
            
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                automation.ws_members.cell(row=idx, column=12).value = 26 - ws_update.cell(row=(member * num) - start_set, column=4).value
                automation.ws_members.cell(row=idx, column=13).value = 6 - ws_update.cell(row=(member * num) - start_set, column=5).value
                automation.ws_members.cell(row=idx, column=11).value = automation.ws_members.cell(row=idx, column=12).value + automation.ws_members.cell(row=idx, column=13).value + automation.ws_members.cell(row=idx, column=14).value

            member += 1

        automation.wb_members.save("D:\\Master\\"+operate_data.ac_name+"_명단총정리.xlsx")


    def mkfile(self, new_path_num, new_path_time, file_name):
        # new_path 는 기수를 설정하여, 각 기수의 멤버를 받아 파일을 복사한다.
        # ex) original_path = "D:\\"+operate_data.ac_name+"\\교육생관리\\4기\\4기주간1207\\1.abc\\abc_요양보호사 자격증 발급,재발급 신청서.hwp" / new_path =  "4기야간0201"
        # how to use ? : x.mkfile(4, "주간", "교육수료증명서.hwp") !!! 주의 !!! 꼭 확장자 명을 작성할 것 !
        self.new_path_num = new_path_num
        self.new_path_time = new_path_time
        self.file_name = file_name

        wb_automation = load_workbook("D:\\Master\\업무자동화.xlsx")
        ws_automation = wb_automation.active
        if new_path_time == "주간":
            startingdate = ws_automation.cell(row=112 + new_path_num, column=2).value
        elif new_path_time == "야간":
            startingdate = ws_automation.cell(row=112 + new_path_num, column=3).value
        string_set = f"{new_path_num}기{new_path_time}{startingdate}"
        print(string_set)

        i = 1
        original_path = f"D:\\Master\\mkfile\\{file_name}"
        print("파일을 복사합니다.\n원본파일 :", original_path)
        for idx, cell in enumerate(automation.ws_members["E"], start=1):
            if string_set != cell.value:
                continue
            if automation.ws_members.cell(row=idx, column=15).value == "일반":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}\\{automation.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(사복)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_사복\\{automation.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간조)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간조\\{automation.ws_members.cell(row=idx, column=18).value}_{file_name}"
            elif automation.ws_members.cell(row=idx, column=15).value == "자격증(간호)":
                path = "D:\\"+operate_data.ac_name+f"\\교육생관리\\{str(new_path_num)}기\\{string_set}\\{i}. {automation.ws_members.cell(row=idx, column=18).value}_간호\\{automation.ws_members.cell(row=idx, column=18).value}_{file_name}"
            shutil.copyfile(original_path, path)
            print("파일이 복사되었습니다 :", path)
            i += 1



        
            
# file_position_detail : x = 335, y = 230, +- = 21
# file_position_big_icorn : x = 330, y = 242, +- = 110
# folder_goto_back : 20,165
# win + left 기준 폰트 : 615,150 / 맑은고딕 : 700,465
# file_path : alt + D
# pyautogui.scroll() = 한 틱당 약 500

# 교육수료증명서(최 우측, win + left 기준)
# 호수 : 300,320 / 이름 : 715,410 / 주소 :  715,455 / 주민등록번호 : 475,505 / 전화번호 : 715,505 / 교육과정명 : 715,545 / 이론실기 이수기간 : 595,635 / 이론실기 이수시간 : 715,635 / 
# 실습기간 : 600,730 / 실습시간 : 715,730 / 총 실습시간 : 715,890 / 총 이수시간 : 715,945 / -> pyautogui.scroll(-2000) -> / 수여일 : 600,765 / 닫기 : 942,12 

# 대체실습확인서(최 우측, win + left 기준)
# 이름 220,475 / 생년월일 305,475 / 교육기관명 565,475 / 교육과정명 685,475 / 구분 : 215,610(640) *** 이름 ~ 도 승인일자 틀림 수정필요 *** / 이름 : 300,555 / 생년월일 : 395,555 / 
# 자격사항 490,555 / 경력기간 590,555 / 도 승인일자 685,555 대체실습 기간 675,785 / 대체실습 시간 680,825 / 합격여부 375,890 / 자체시험 점수 680,890 / 비고 380,930 / 교수이름 680,930 /
# 수여 날짜 705,780

# 요양보호사 자격증 발급, 재발급 신청서(최 우측, win + left 기준)
# 이름 600,420 / 요양보호사 교육기간(이론, 실기) 320,580 / 발급신청일(표 중앙) 500,795 // 나머지는 alt + right 로 커버 ^^
# x = automation()
# x.auto_move_class(3, "야간")
# x.automation_task("edu_completion")

# 대체실습 수료보고
# 제목 615,450 / 2페이지 연번 915,600 / 18번 스크롤(2250) 한 후 3페이지 연번(11) 140,435 / 4페이지 연번(25) 1220,435 / 14번 스크롤 한 후 5페이지 연번(39) 140,430 / 
#  교육과정명 710,570 / 교육인원/수료인원 710,610 / 이론/실기 교육시간 710,690 / 실습 교육기간 710,755 / 수료일 710,795 / 1번 교육과정명 975,600 / 
# 2페이지 11번 교육과정명 195,445

# 개강보고
# ac 교육대상자 명단
# 과정구분 1095,345

# 대체실습 실시보고 scroll(-2500) 기준
# [교육생명단] 교육구분 370,395

# 출석부
# 주간 355,480 / 야간 1140,515

# 한글 꿀팁
# 서식복사 : alt + C


# pyautogui.mouseInfo()

